# pages/1_Report.py — WORKING HOUR REPORT (rebuilt on the real face-scan
# timesheet via timesheet_db). Full-width, branded table; leave-type
# selection; month filter; Permanent/Contract & OT-rate column toggles.
import io
import datetime as dt
import calendar
import streamlit as st
st.set_page_config(layout="wide", initial_sidebar_state="expanded")
from lib import theme as _theme
from lib.auth import require_capability, has_capability, current_user
from lib import timesheet_db as ts

_theme.inject()
require_capability("report.view")
user = current_user()
view_all = has_capability("report.view_all")

_theme.header("รายงานชั่วโมงทำงาน", "Working Hour Report",
              "กำลังคน • ชั่วโมงทำงาน • OT ตามอัตรา • การลาแยกประเภท ตามกลุ่ม/หน้าที่")

# ───────────────────────── timesheet source (admin-managed) ────────────
up = ts.latest_upload()
if not up:
    st.info("ยังไม่มีข้อมูลบันทึกเวลา — ผู้ดูแลระบบอัปโหลดไฟล์ได้ที่ "
            "**ผู้ดูแลระบบ → ข้อมูล & อัปโหลด** · No timesheet yet; an admin "
            "uploads it in **Admin → Data & Uploads**.")
    st.stop()

# ───────────────────────── settings (collapsible) ──────────────────────
df_from = dt.date.fromisoformat(up["date_from"]) if up["date_from"] else None
df_to = dt.date.fromisoformat(up["date_to"]) if up["date_to"] else None
months = ts.months(up["id"])
with st.expander("⚙️ ตั้งค่ารายงาน · Report settings", expanded=False):
    c1, c2 = st.columns(2)
    month = c1.selectbox("เดือน · Month", ["(ทุกเดือน · all months)"] + months)
    density = c2.radio("ความหนาแน่น · Density",
                       ["ปกติ · Normal", "กระชับ · Compact"], horizontal=True)
    _adv = st.toggle(
        "⚙️ ตั้งชั่วโมงมาตรฐานแยกรายวัน/ประเภท (ขั้นสูง) · per-weekday & "
        "per-type standard hours", value=False)
    weekday_hours = None
    type_hours = None
    if not _adv:
        std_day_hours = st.number_input(
            "⏱️ ชั่วโมงทำงานมาตรฐานต่อวัน · Standard hours per working day "
            "(ใช้แปลงวันลา/ขาดงาน → ชั่วโมง · converts leave/absence days → "
            "hours)", min_value=1.0, max_value=24.0, value=8.0, step=0.5)
    else:
        std_day_hours = 8.0
        st.caption("ชั่วโมงมาตรฐานของแต่ละวันในสัปดาห์ (เช่น เสาร์ครึ่งวัน) — "
                   "วันลา/ขาดงานของวันใดจะถูกแปลงเป็นชั่วโมงตามมาตรฐานวันนั้น · "
                   "standard hours per weekday; each leave/absence day converts "
                   "using that weekday's standard.")
        _wd = st.columns(7)
        _wl = ["จ·Mon", "อ·Tue", "พ·Wed", "พฤ·Thu", "ศ·Fri", "ส·Sat", "อา·Sun"]
        weekday_hours = {}
        for _i, _lab in enumerate(_wl):
            weekday_hours[_i] = _wd[_i].number_input(
                _lab, 0.0, 24.0, 8.0 if _i < 5 else 0.0, 0.5, key=f"wh{_i}")
        st.caption("override รายประเภท (>0 = บังคับชั่วโมงนั้นสำหรับประเภทนั้น "
                   "ทุกวัน, 0 = ใช้มาตรฐานรายวันด้านบน) · per-type override "
                   "(>0 forces that many hours/day for that type; 0 = use the "
                   "weekday standard above).")
        _tc = st.columns(4)
        type_hours = {}
        for _col, _t, _lab in ((_tc[0], "absent", "ขาดงาน·Absent"),
                               (_tc[1], "sick", "ลาป่วย·Sick"),
                               (_tc[2], "personal", "ลากิจ·Personal"),
                               (_tc[3], "annual", "ลาพักร้อน·Annual")):
            _v = _col.number_input(_lab, 0.0, 24.0, 0.0, 0.5, key=f"th_{_t}")
            if _v > 0:
                type_hours[_t] = _v
    st.markdown("**นับวันลาประเภทใดเป็น 'การขาดงาน' · count which leave types "
                "as absence**")
    st.caption("ไฟล์บันทึกเวลาจับ 4 ประเภทนี้ — เลือกได้ว่าให้ประเภทใดรวมในอัตราขาด"
               "งาน (ลาพักร้อนยังแสดงแยกเป็นคอลัมน์ AL เสมอ) · the timesheet "
               "captures these 4 categories; tick which count toward the absence "
               "rate (annual is also always shown separately as AL).")
    la, lb, lc, ld = st.columns(4)
    inc_absent = la.checkbox("ขาดงาน · No-show", value=True)
    inc_sick = lb.checkbox("ลาป่วย · Sick", value=True)
    inc_personal = lc.checkbox("ลากิจ · Personal", value=True)
    inc_annual = ld.checkbox("ลาพักร้อน · Annual", value=False)
    t1, t2 = st.columns(2)
    show_class = t1.toggle("แยก ประจำ/ซับ · Permanent/Contract split",
                           value=False)
    show_ot = t2.toggle("แยกอัตรา OT · OT-rate breakdown (×1/×1.5/×2/×3)",
                        value=True)
    show_days = st.toggle("แสดงจำนวนวันลาแยกประเภท · show leave days by type",
                          value=False)

absent_types = set()
if inc_absent:
    absent_types.add("absent")
if inc_sick:
    absent_types.add("sick")
if inc_personal:
    absent_types.add("personal")
if inc_annual:
    absent_types.add("annual")

if month.startswith("("):
    d_from, d_to = up["date_from"], up["date_to"]
    period_label = f"{up['date_from']} → {up['date_to']}"
else:
    y, m = map(int, month.split("-"))
    d_from = f"{y:04d}-{m:02d}-01"
    d_to = f"{y:04d}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"
    period_label = month

rows = ts.report_rows(up["id"], date_from=d_from, date_to=d_to,
                      absent_types=absent_types, day_hours=std_day_hours,
                      weekday_hours=weekday_hours, type_hours=type_hours)

# ───────────────────────── aggregation helpers ─────────────────────────
GROUPS = ["SG&A", "MANU Support", "MANU"]
SUM_KEYS = ["hc", "perm_hc", "con_hc", "work", "perm_work", "con_work",
            "absent_h", "perm_absent_h", "con_absent_h", "al_h", "perm_al_h",
            "con_al_h", "ot1", "ot15", "ot2", "ot3", "total_ot", "sick_d",
            "personal_d", "absent_d", "annual_d"]


def _agg(rs):
    d = {k: sum(r.get(k, 0) for r in rs) for k in SUM_KEYS}
    base = d["work"] + d["absent_h"]
    pbase = d["perm_work"] + d["perm_absent_h"]
    cbase = d["con_work"] + d["con_absent_h"]
    d["pct_absent"] = (d["absent_h"] / base * 100) if base else 0
    d["pct_perm_absent"] = (d["perm_absent_h"] / pbase * 100) if pbase else 0
    d["pct_con_absent"] = (d["con_absent_h"] / cbase * 100) if cbase else 0
    d["pct_ot"] = (d["total_ot"] / d["work"] * 100) if d["work"] else 0
    return d


grand = _agg(rows)

# previous-month aggregate for month-over-month comparison (req. 1.5)
grand_prev = None
if not month.startswith("("):
    _py, _pm = (y, m - 1) if m > 1 else (y - 1, 12)
    _pf = f"{_py:04d}-{_pm:02d}-01"
    _pt = f"{_py:04d}-{_pm:02d}-{calendar.monthrange(_py, _pm)[1]:02d}"
    _prev_rows = ts.report_rows(up["id"], date_from=_pf, date_to=_pt,
                                absent_types=absent_types,
                                day_hours=std_day_hours,
                                weekday_hours=weekday_hours,
                                type_hours=type_hours)
    if _prev_rows:
        grand_prev = _agg(_prev_rows)

# ───────────────────────── summary cards (TOP) ─────────────────────────
st.caption(f"ช่วงข้อมูล · Period: **{period_label}** • "
           f"นับขาดงานจาก · absence from: "
           f"{', '.join(sorted(absent_types)) or '—'} • ลาพักร้อนแยกต่างหาก")
def _d(key, fmt="{:+,.0f}"):
    """Month-over-month delta string, or None when no prior month."""
    if grand_prev is None:
        return None
    return fmt.format(grand[key] - grand_prev[key]) + " vs prev mo"


_VS = " vs prev mo"
m = st.columns(6)
m[0].metric("กำลังคน · Headcount", f"{int(grand['hc']):,}",
            _d("hc"), delta_color="off")
m[1].metric("ชม.ทำงาน · Working (hrs)", f"{grand['work']:,.0f}",
            _d("work"), delta_color="off")
m[2].metric(f"ขาดงาน · Absent — {grand['pct_absent']:.2f}%",
            f"{grand['absent_h']:,.0f} ชม.",
            (f"{grand['pct_absent'] - grand_prev['pct_absent']:+.2f} pp{_VS}"
             if grand_prev else None), delta_color="inverse")
m[3].metric(f"OT รวม · OT total — {grand['pct_ot']:.1f}%",
            f"{grand['total_ot']:,.0f} ชม.",
            (f"{grand['pct_ot'] - grand_prev['pct_ot']:+.1f} pp{_VS}"
             if grand_prev else None), delta_color="inverse")
m[4].metric("ลาพักร้อน · Annual leave (hrs)", f"{grand['al_h']:,.0f}",
            _d("al_h"), delta_color="off")
_sp = grand["sick_d"] + grand["personal_d"]
m[5].metric("ลาป่วย+ลากิจ · Sick+Personal (days)", f"{_sp:,.0f}",
            (f"{_sp - (grand_prev['sick_d'] + grand_prev['personal_d']):+,.0f}"
             f"{_VS}" if grand_prev else None), delta_color="inverse")
if grand_prev is None and not month.startswith("("):
    st.caption("ℹ️ ไม่มีข้อมูลเดือนก่อนหน้าสำหรับเปรียบเทียบ · no prior-month "
               "data to compare against.")

# ───────────────────────── column spec ─────────────────────────────────
# (key, header_th, header_en, kind)  kind: lbl | int | hr | pct
COLS = [("group", "กลุ่ม", "Group", "lbl"),
        ("function", "หน้าที่", "Function", "lbl"),
        ("hc", "HC รวม", "Total HC", "int")]
if show_class:
    COLS += [("perm_hc", "ประจำ", "Perm HC", "int"),
             ("con_hc", "ซับ", "Contract HC", "int")]
COLS += [("work", "ชม.ทำงาน", "Working Hrs", "hr")]
if show_class:
    COLS += [("perm_work", "ทำงาน(ประจำ)", "Perm Working", "hr"),
             ("con_work", "ทำงาน(ซับ)", "Contract Working", "hr")]
COLS += [("absent_h", "ขาดงาน(ชม.)", "Absent Hrs", "hr")]
if show_class:
    COLS += [("perm_absent_h", "ขาด(ประจำ)", "Perm Absent", "hr"),
             ("con_absent_h", "ขาด(ซับ)", "Contract Absent", "hr")]
COLS += [("pct_absent", "% ขาด", "% Absent", "pct")]
if show_class:
    COLS += [("pct_perm_absent", "% ขาด(ประจำ)", "% Perm Absent", "pct"),
             ("pct_con_absent", "% ขาด(ซับ)", "% Contract Absent", "pct")]
if show_days:
    COLS += [("sick_d", "ลาป่วย(วัน)", "Sick(d)", "hr"),
             ("personal_d", "ลากิจ(วัน)", "Personal(d)", "hr"),
             ("absent_d", "ขาด(วัน)", "No-show(d)", "hr")]
if show_ot:
    COLS += [("ot1", "OT×1", "OT×1", "hr"), ("ot15", "OT×1.5", "OT×1.5", "hr"),
             ("ot2", "OT×2", "OT×2", "hr"), ("ot3", "OT×3", "OT×3", "hr")]
COLS += [("total_ot", "OT รวม", "Total OT", "hr"),
         ("pct_ot", "% OT", "% OT", "pct"),
         ("al_h", "ลาพักร้อน(ชม.)", "AL Hrs", "hr")]
if show_class:
    COLS += [("perm_al_h", "AL(ประจำ)", "Perm AL", "hr"),
             ("con_al_h", "AL(ซับ)", "Contract AL", "hr")]

_lang = _theme.i18n.cur_lang()
NUMK = {"int", "hr", "pct"}


def _fmt(v, kind):
    if kind == "int":
        return f"{int(round(v)):,}"
    if kind == "hr":
        return f"{v:,.1f}"
    if kind == "pct":
        return f"{v:.2f}%"
    return str(v)


# ───────────────────────── build full-width HTML table ─────────────────
GROUP_ACCENT = {"SG&A": "#009ADE", "MANU Support": "#715091", "MANU": "#E31D93"}


def _cell(val, kind, cls=""):
    align = "right" if kind in NUMK else "left"
    return f'<td class="{cls}" style="text-align:{align}">{val}</td>'


def _row_html(r, accent, cls="", show_group=True):
    tds = []
    for k, _th, _en, kind in COLS:
        if k == "group":
            v = r.get("group", "") if show_group else ""
            tds.append(f'<td class="{cls} grpcell" style="border-left:4px '
                       f'solid {accent}">{v}</td>')
        elif k == "function":
            tds.append(f'<td class="{cls}" style="text-align:left">'
                       f'<b>{r.get("function","")}</b></td>')
        else:
            tds.append(_cell(_fmt(r.get(k, 0), kind), kind, cls))
    return "<tr>" + "".join(tds) + "</tr>"


head = "".join(f'<th style="text-align:{"right" if kind in NUMK else "left"}">'
               f'{(_th if _lang=="th" else _en)}</th>'
               for k, _th, _en, kind in COLS)
body = []
for g in GROUPS:
    grs = [r for r in rows if r["group"] == g]
    if not grs:
        continue
    grs.sort(key=lambda r: r["function"])
    accent = GROUP_ACCENT[g]
    for r in grs:
        body.append(_row_html(r, accent))
    sub = _agg(grs); sub["group"] = g; sub["function"] = f"รวม {g} · Total {g}"
    body.append(_row_html(sub, accent, cls="subtot"))
gt = dict(grand); gt["group"] = ""; gt["function"] = "GRAND TOTAL"
body.append(_row_html(gt, "#1f2937", cls="grand"))

px = 8 if density.startswith("กระชับ") else 11
table_html = f"""
<style>
.whr-wrap{{width:100%;overflow-x:auto;border:1px solid #e6e9f2;
  border-radius:16px;box-shadow:0 6px 22px rgba(15,23,42,.06);margin:6px 0 4px}}
.whr{{border-collapse:separate;border-spacing:0;width:100%;font-size:12.5px;
  font-variant-numeric:tabular-nums;min-width:640px}}
.whr thead th{{position:sticky;top:0;z-index:2;background:linear-gradient(135deg,
  #2A2F45,#4b3b63);color:#fff;font-weight:700;padding:10px {px}px;
  white-space:nowrap;font-size:11.5px;text-align:right;border-bottom:0}}
.whr thead th:nth-child(1),.whr thead th:nth-child(2){{text-align:left}}
.whr tbody td{{padding:{px-2}px {px}px;border-bottom:1px solid #eef1f7;
  white-space:nowrap;color:#26303E}}
.whr tbody tr:nth-child(even) td{{background:#fafbff}}
.whr tbody tr:hover td{{background:#eef6fd}}
.whr .grpcell{{color:#5b6472;font-size:11px}}
.whr tbody tr td.subtot{{background:#efeafa !important;font-weight:800;
  color:#3d2a66;border-top:2px solid #d8cdf0;border-bottom:2px solid #d8cdf0}}
.whr tbody tr td.grand{{background:#1f2937 !important;color:#fff;
  font-weight:800;font-size:13px}}
</style>
<div class="whr-wrap"><table class="whr"><thead><tr>{head}</tr></thead>
<tbody>{''.join(body)}</tbody></table></div>
"""
st.markdown(table_html, unsafe_allow_html=True)
st.caption("เลื่อนซ้าย-ขวาเพื่อดูทุกคอลัมน์ · scroll horizontally to see all "
           "columns • เปิด 'ตั้งค่ารายงาน' เพื่อเลือกคอลัมน์/ประเภทการลา")

# ───────────────────────── export ──────────────────────────────────────
st.markdown("##### ⬇️ ส่งออก · Export")
hdrs = [(_th if _lang == "th" else _en) for k, _th, _en, kind in COLS]


def _table_matrix():
    mat = [hdrs]
    for g in GROUPS:
        grs = sorted([r for r in rows if r["group"] == g],
                     key=lambda r: r["function"])
        for r in grs:
            mat.append([r.get(k, "") if kind == "lbl" else
                        round(r.get(k, 0), 2) for k, _t, _e, kind in COLS])
        if grs:
            sub = _agg(grs); sub["group"] = g
            sub["function"] = f"Total {g}"
            mat.append([sub.get(k, "") if kind == "lbl" else
                        round(sub.get(k, 0), 2) for k, _t, _e, kind in COLS])
    g2 = dict(grand); g2["group"] = ""; g2["function"] = "GRAND TOTAL"
    mat.append([g2.get(k, "") if kind == "lbl" else round(g2.get(k, 0), 2)
                for k, _t, _e, kind in COLS])
    return mat


e1, e2 = st.columns(2)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb = Workbook(); ws = wb.active; ws.title = "Working hours"
mat = _table_matrix()
for ri, row in enumerate(mat):
    ws.append(row)
    for ci in range(1, len(row) + 1):
        cell = ws.cell(ri + 1, ci)
        if ri == 0:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2A2F45")
        elif str(row[1]).startswith("Total "):
            cell.fill = PatternFill("solid", fgColor="EFEAFA")
            cell.font = Font(bold=True)
        elif row[1] == "GRAND TOTAL":
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(bold=True, color="FFFFFF")
buf = io.BytesIO(); wb.save(buf)
e1.download_button("📊 Excel (.xlsx)", buf.getvalue(),
                   file_name=f"Working_hour_report_{period_label}.xlsx",
                   use_container_width=True)
# PNG generated inline — no "prepare" step; the download button just appears.
_png_bytes = None
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    # register bundled Thai font so any Thai cell text renders (not tofu)
    try:
        import os as _os
        import matplotlib.font_manager as _fm
        _fpath = "assets/fonts/Sarabun-Regular.ttf"
        if not _os.path.exists(_fpath):
            _fpath = _os.path.join(_os.path.dirname(__file__), "..",
                                   "assets", "fonts", "Sarabun-Regular.ttf")
        if _os.path.exists(_fpath):
            _fm.fontManager.addfont(_fpath)
            matplotlib.rcParams["font.family"] = \
                _fm.FontProperties(fname=_fpath).get_name()
    except Exception:
        pass
    # PNG column headers are ALWAYS English (Thai headers render poorly here)
    _png_hdrs = [_en for k, _t, _en, kind in COLS]
    fig, ax = plt.subplots(figsize=(min(24, 1.0 * len(COLS)),
                                    0.34 * (len(mat) + 1)))
    ax.axis("off")
    tbl = ax.table(cellText=[[str(c) for c in r] for r in mat[1:]],
                   colLabels=_png_hdrs, loc="center", cellLoc="center")
    tbl.auto_set_font_size(False); tbl.set_fontsize(7); tbl.scale(1, 1.3)
    for j in range(len(_png_hdrs)):
        tbl[0, j].set_facecolor("#2A2F45")
        tbl[0, j].set_text_props(color="white", fontweight="bold")
    _pb = io.BytesIO(); fig.savefig(_pb, dpi=150, bbox_inches="tight")
    plt.close(fig); _png_bytes = _pb.getvalue()
except Exception:
    _png_bytes = None
if _png_bytes:
    e2.download_button("🖼️ PNG image", _png_bytes,
                       file_name=f"Working_hour_report_{period_label}.png",
                       mime="image/png", use_container_width=True)
else:
    e2.caption("PNG ไม่พร้อมในเครื่องนี้ · PNG unavailable here")
