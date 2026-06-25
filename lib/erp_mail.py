# lib/erp_mail.py
# ============================================================================
# Outlook integration for the purchase workflow (§1.2-1.3) — NO IT setup:
# we generate a .eml DRAFT file. Opening it launches Outlook with
# To / Subject / bilingual body filled and the Excel file ALREADY ATTACHED;
# the user just presses Send.  (X-Unsent: 1 = "open as unsent draft".)
#   - request_eml():  to the current approver, on submit / each level
#   - supplier_eml(): to the supplier, after approval (RFQ / ใบเสนอราคา)
# The Excel attachment follows the agreed layout: header, requester block
# (Staff ID / name / cost centre / position / department), item lines
# (item no., detail, qty, price), totals + approval line + signatures.
# ============================================================================
import datetime as dt
import io
from email.message import EmailMessage

from lib import erp_db
from lib import approval_db as adb
from lib import employee_db as edb

XLSX_MIME = ("application/"
             "vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------- Excel
def build_po_xlsx(po, lines, requester_rec=None, approvals=None):
    """Returns bytes of the PO Excel per blueprint §1.3."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook(); ws = wb.active; ws.title = "Purchase Request"
    PURPLE = "715091"; BLUE = "009ADE"
    H = Font(bold=True, color="FFFFFF", name="Tahoma", size=10)
    B = Font(name="Tahoma", size=10)
    BB = Font(name="Tahoma", size=10, bold=True)
    FILL_P = PatternFill("solid", fgColor=PURPLE)
    FILL_B = PatternFill("solid", fgColor=BLUE)
    TH = Border(*[Side(style="thin")] * 4)

    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = ("ANCA Manufacturing Solutions (Thailand) — "
                             "ใบขอสั่งซื้อ / Purchase Request")
    c.font = Font(name="Tahoma", size=13, bold=True, color=PURPLE)

    hdr = [("PO No.", po.get("po_no")),
           ("Date", str(po.get("created_at") or "")[:10]),
           ("Status", po.get("status")),
           ("Supplier", po.get("supplier_key") or "-"),
           ("Purpose / วัตถุประสงค์", po.get("purpose") or "-")]
    r = 3
    for k, v in hdr:
        ws.cell(row=r, column=1, value=k).font = BB
        ws.cell(row=r, column=2, value=str(v)).font = B
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="ผู้ขอ / Requester").font = H
    ws.cell(row=r, column=1).fill = FILL_P
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    req = requester_rec or {}
    cost_centre = req.get("cost_center") or req.get("cost_centre") or "-"
    for k, v in [("Staff ID", po.get("requester_emp_no")),
                 ("Name", po.get("requester_name")),
                 ("Name (TH)", req.get("emp_name_th") or "-"),
                 ("Cost centre", cost_centre),
                 ("Position", req.get("title") or "-"),
                 ("Department", po.get("department"))]:
        ws.cell(row=r, column=1, value=k).font = BB
        ws.cell(row=r, column=2, value=str(v or "-")).font = B
        r += 1

    r += 1
    cols = ["#", "Item no. / รหัส", "Item detail / รายการ", "Brand",
            "Qty", "Unit price ฿", "Line total ฿"]
    for j, cname in enumerate(cols, 1):
        cc = ws.cell(row=r, column=j, value=cname)
        cc.font = H; cc.fill = FILL_B; cc.border = TH
        cc.alignment = Alignment(horizontal="center")
    r += 1
    total = 0.0
    prods = {p["product_code"]: p for p in
             erp_db.search_products(limit=100000)} if lines else {}
    for i, l in enumerate(lines, 1):
        lt = (l.get("unit_price_thb") or 0) * (l.get("qty") or 0)
        total += lt
        brand = (prods.get(l.get("product_code")) or {}).get("brand") or ""
        vals = [i, l.get("product_code") or "-", l.get("description") or "",
                brand, l.get("qty"), l.get("unit_price_thb"), lt]
        for j, v in enumerate(vals, 1):
            cc = ws.cell(row=r, column=j, value=v)
            cc.font = B; cc.border = TH
            if j >= 5:
                cc.number_format = "#,##0.00" if j > 5 else "0.##"
        r += 1
    ws.cell(row=r, column=6, value="รวม / Grand total").font = BB
    cc = ws.cell(row=r, column=7, value=total)
    cc.font = BB; cc.number_format = "#,##0.00"
    r += 2

    ws.cell(row=r, column=1, value="สายอนุมัติ / Approval line").font = H
    ws.cell(row=r, column=1).fill = FILL_P
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    for a in (approvals or []):
        ws.cell(row=r, column=1,
                value=f"L{a['level']}  {a['approver_name']}").font = B
        ws.cell(row=r, column=3, value=a["status"]).font = B
        ws.cell(row=r, column=4,
                value=str(a.get("acted_at") or "")).font = B
        r += 1
    r += 1
    for label in ("ผู้ขอ / Requested by", "ผู้อนุมัติ / Approved by"):
        ws.cell(row=r, column=1, value=label).font = BB
        ws.cell(row=r, column=3,
                value="ลงชื่อ ____________________  วันที่ ________").font = B
        r += 2

    for col, w in zip("ABCDEFG", [14, 18, 42, 14, 8, 12, 13]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- .eml
def _eml(to, subject, html_body, attach_bytes, attach_name):
    msg = EmailMessage()
    msg["To"] = to or ""
    msg["Subject"] = subject
    msg["X-Unsent"] = "1"        # Outlook opens it as an editable draft
    msg.set_content("Please view this email in HTML.")
    msg.add_alternative(html_body, subtype="html")
    if attach_bytes:
        maintype, subtype = XLSX_MIME.split("/", 1)
        msg.add_attachment(attach_bytes, maintype=maintype, subtype=subtype,
                           filename=attach_name)
    return bytes(msg)


def _body_table(po, lines):
    rows = "".join(
        f"<tr><td>{i}</td><td>{l.get('product_code') or '-'}</td>"
        f"<td>{l.get('description') or ''}</td>"
        f"<td align='center'>{l.get('qty')}</td>"
        f"<td align='right'>{(l.get('unit_price_thb') or 0):,.2f}</td>"
        f"<td align='right'>"
        f"{((l.get('unit_price_thb') or 0) * l.get('qty')):,.2f}</td></tr>"
        for i, l in enumerate(lines, 1))
    return f"""
    <table border="1" cellspacing="0" cellpadding="5"
     style="border-collapse:collapse;font-size:13px">
     <tr style="background:#715091;color:#fff"><th>#</th><th>รหัส / Item
     no.</th><th>รายการ / Detail</th><th>จำนวน / Qty</th>
     <th>ราคา/หน่วย ฿</th><th>รวม ฿</th></tr>{rows}
     <tr><td colspan="5" align="right"><b>รวมทั้งสิ้น / Grand total</b></td>
     <td align="right"><b>{(po.get('total_thb') or 0):,.2f}</b></td></tr>
    </table>"""


def request_eml(po, lines, approver_email, approver_name, level,
                requester_rec=None):
    """Approval-request draft to the current approver."""
    approvals = adb.rows_for("po", po["id"])
    xlsx = build_po_xlsx(po, lines, requester_rec, approvals)
    subject = (f"[ขออนุมัติสั่งซื้อ L{level}] {po['po_no']} — "
               f"{po.get('requester_name')} ({po.get('department')})")
    body = f"""<div style="font-family:Tahoma,'Sarabun';font-size:13.5px">
    <p>เรียนคุณ {approver_name},</p>
    <p>ขอความกรุณาพิจารณาอนุมัติคำขอสั่งซื้อต่อไปนี้ (รายละเอียดเต็มในไฟล์
    Excel แนบ) / Please review the purchase request below — the full detail
    is in the attached Excel:</p>
    <p><b>{po['po_no']}</b> • ผู้ขอ: {po.get('requester_name')}
    ({po.get('requester_emp_no')}) • แผนก: {po.get('department')}<br>
    วัตถุประสงค์: {po.get('purpose') or '-'}</p>
    {_body_table(po, lines)}
    <p>อนุมัติได้ในระบบ HR (เมนู Stationery ERP → Approvals) หรือตอบอีเมลนี้
    / Approve in the HR system (Stationery ERP → Approvals) or reply to
    this email — HR will attach your reply as evidence.</p>
    <p style="color:#888;font-size:11px">AMS HR System • PO workflow</p>
    </div>"""
    return _eml(approver_email, subject, body, xlsx, f"{po['po_no']}.xlsx")


def supplier_eml(po, lines, supplier):
    """RFQ / ใบเสนอราคา draft to the supplier, after approval."""
    xlsx = build_po_xlsx(po, lines)
    subject = (f"ขอใบเสนอราคา / Request for quotation — "
               f"ANCA AMS {po['po_no']}")
    body = f"""<div style="font-family:Tahoma,'Sarabun';font-size:13.5px">
    <p>เรียน {supplier.get('name')},</p>
    <p>บริษัท แอนคา แมนูแฟคเจอริ่ง โซลูชั่นส์ (ประเทศไทย) จำกัด
    มีความประสงค์สั่งซื้อสินค้าตามรายการด้านล่าง (ไฟล์ Excel แนบ)
    ขอความกรุณาส่งใบเสนอราคา / We would like to order the items below
    (Excel attached) — please send your quotation:</p>
    {_body_table(po, lines)}
    <p>จัดส่งที่ / Deliver to: ANCA Manufacturing Solutions (Thailand) Ltd.,
    109/14 M.4 T.Pluakdaeng A.Pluakdaeng Rayong 21140<br>
    อ้างอิง / Reference: {po['po_no']}</p>
    <p>ขอบคุณครับ/ค่ะ — ฝ่ายจัดซื้อ/ธุรการ AMS</p></div>"""
    return _eml(supplier.get("email"), subject, body, xlsx,
                f"{po['po_no']}.xlsx")
