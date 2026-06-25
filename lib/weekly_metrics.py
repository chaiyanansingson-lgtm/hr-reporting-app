# lib/weekly_metrics.py
# Standard HR weekly metrics (OT% & Absenteeism% by department) for the KPI
# Dashboard "Weekly" view. Two data sources:
#   - seeded history (FY tracker / Weekly Metric Report) up to a cutover week
#   - live computation from the 3 uploaded HR reports thereafter
# Formula (validated cell-by-cell against the FY tracker / Weekly Metric Report):
#   Working_dept = HC_dept * daily_hours * working_days - Leave_dept
#   OT%_dept     = OT / (OT + Working)         (department-relative)
#   Absent%_dept = Leave / (Working + Leave)
#   annual + off-site work + off-site training are excluded from Leave by default.
import io
import json
import re

from lib.db import get_conn, IS_POSTGRES, PH
from lib import employee_db as edb
from lib import attendance_db as att

SERIAL = "SERIAL PRIMARY KEY" if IS_POSTGRES else \
    "INTEGER PRIMARY KEY AUTOINCREMENT"

# The 11 reporting departments, in display order (from the FY tracker).
DEPT_ORDER = ["Laser", "Folding", "CNC Machine shop", "Weld", "Paint",
              "Assembly", "Misumi", "QC", "Packing", "Warehouse",
              "Office Staff"]

# Default cost-centre (MASTER column G, 3-digit code) -> reporting department.
DEFAULT_CC_MAP = {
    "220": "Laser", "221": "Folding", "210": "CNC Machine shop",
    "213": "Weld", "222": "Weld", "214": "Weld", "231": "Paint",
    "240": "Assembly", "223": "Misumi", "300": "QC", "280": "QC",
    "263": "Packing", "270": "Warehouse",
    "310": "Office Staff", "313": "Office Staff", "352": "Office Staff",
    "353": "Office Staff", "354": "Office Staff", "356": "Office Staff",
}

_MONTHS = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
           "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

DEFAULTS = {"daily_hours": "8", "working_days": "5", "ot_target": "0.28",
            "absent_target": "0.025",
            "exclude_leave": "annual,offsite_work,offsite_train"}


def _is_total(name):
    n = str(name or "").strip().lower()
    return (not n) or ("total" in n) or ("overall" in n) or ("รวม" in n)


# Different source files spell the same department/week differently; canonical-
# ise so the FY tracker and the Weekly Metric Report land on the same rows.
DEPT_ALIASES = {"mizumi": "Misumi", "misumi": "Misumi"}


def _norm_dept(name):
    n = str(name or "").strip()
    return DEPT_ALIASES.get(n.lower(), n)


def _norm_week(label):
    """Canonical 'MMM YY-Wn' (upper-case W, single space). 'Jun 26-w1' →
    'Jun 26-W1'."""
    m = re.match(r"\s*([A-Za-z]{3})\s*(\d{2})-[Ww]\s*(\d+)", str(label or ""))
    if not m:
        return str(label or "").strip()
    return f"{m.group(1).title()} {m.group(2)}-W{m.group(3)}"


_MON_SEQ = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _fy_real_week(label):
    """Repair the FY tracker's drag-filled month labels. A month has at most 5
    weeks, so 'Dec 25-W6' rolls into 'Jan 26-W1' and any week >= 7 is empty
    padding (returns None → drop it). Valid W1-W5 pass through unchanged."""
    m = re.match(r"([A-Za-z]{3}) (\d{2})-W(\d+)", label)
    if not m:
        return label
    mon, yy, wk = m.group(1), int(m.group(2)), int(m.group(3))
    if wk <= 5:
        return label
    if wk == 6 and mon in _MON_SEQ:
        i = _MON_SEQ.index(mon)
        return f"Jan {yy + 1}-W1" if i == 11 else f"{_MON_SEQ[i + 1]} {yy}-W1"
    return None                           # week >= 7 → drag-fill padding


def migrate():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""CREATE TABLE IF NOT EXISTS wm_metrics (
        id {SERIAL}, week_label TEXT, week_key REAL, dept TEXT,
        working_hrs REAL, ot_hrs REAL, leave_hrs REAL,
        ot_pct REAL, absent_pct REAL, source TEXT,
        UNIQUE(week_label, dept))""")
    cur.execute("""CREATE TABLE IF NOT EXISTS wm_settings (
        skey TEXT PRIMARY KEY, sval TEXT)""")
    conn.commit()


def week_key(label):
    """Sortable numeric key from a 'MMM YY-Wn' week label. Week is weighted
    below the month so multi-digit weeks (the FY tracker's Dec-W1..W43) still
    order correctly and stay within their month."""
    m = re.match(r"\s*([A-Za-z]{3})\s*(\d{2})-[Ww]\s*(\d+)", str(label or ""))
    if not m:
        return 0.0
    mon = _MONTHS.get(m.group(1).title(), 0)
    return (2000 + int(m.group(2))) * 10000 + mon * 100 + int(m.group(3))


# ------------------------------- settings ----------------------------------
def get_settings():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT skey, sval FROM wm_settings")
    out = dict(DEFAULTS)
    out.update({r[0]: r[1] for r in cur.fetchall()})
    return out


def set_setting(key, val):
    conn = get_conn(); cur = conn.cursor()
    if IS_POSTGRES:
        cur.execute("INSERT INTO wm_settings (skey,sval) VALUES (%s,%s) "
                    "ON CONFLICT (skey) DO UPDATE SET sval=EXCLUDED.sval",
                    (key, str(val)))
    else:
        cur.execute("INSERT OR REPLACE INTO wm_settings (skey,sval) "
                    "VALUES (?,?)", (key, str(val)))
    conn.commit()


def targets():
    s = get_settings()
    return float(s["ot_target"]), float(s["absent_target"])


def get_cc_map():
    """cost-centre 3-digit code -> reporting department, from settings
    (editable) falling back to DEFAULT_CC_MAP."""
    raw = get_settings().get("cc_map")
    if raw:
        try:
            m = json.loads(raw)
            if isinstance(m, dict) and m:
                return {str(k): v for k, v in m.items()}
        except Exception:
            pass
    return dict(DEFAULT_CC_MAP)


def set_cc_map(d):
    set_setting("cc_map", json.dumps(d, ensure_ascii=False))


def get_cc_groups():
    """The managed list of cost-centre group (department) names. Stored in
    settings; defaults to the standard departments plus any extra group that a
    mapping already points at."""
    raw = get_settings().get("cc_groups")
    if raw:
        try:
            g = json.loads(raw)
            if isinstance(g, list) and g:
                return [str(x) for x in g]
        except Exception:
            pass
    groups = list(DEPT_ORDER)
    for v in get_cc_map().values():
        if v and v not in groups:
            groups.append(v)
    return groups


def set_cc_groups(groups):
    seen, out = set(), []
    for g in groups:
        g = str(g).strip()
        if g and g not in seen:
            seen.add(g); out.append(g)
    set_setting("cc_groups", json.dumps(out, ensure_ascii=False))


def rename_cc_group(old, new):
    """Rename a group everywhere: the group list and every mapping using it."""
    old, new = str(old).strip(), str(new).strip()
    if not old or not new or old == new:
        return
    set_cc_groups([new if g == old else g for g in get_cc_groups()])
    set_cc_map({k: (new if v == old else v) for k, v in get_cc_map().items()})


def cost_centre_usage():
    """Scan the active employee master and count employees per 3-digit cost
    centre code → {code: count}. Powers 'discover from the employee list'."""
    out = {}
    try:
        for r in edb.list_records("active"):
            cc = str(r.get("cost_centre") or "").strip()
            if not cc:
                continue
            code = cc[:3]
            out[code] = out.get(code, 0) + 1
    except Exception:
        pass
    return out


# ------------------------------- storage -----------------------------------
def upsert(week_label, dept, working=None, ot=None, leave=None,
           ot_pct=None, absent_pct=None, source="seed", keep_raw=False):
    """Insert/replace one (week, dept) row. Percentages are computed from raw
    hours when those are supplied and the percentage isn't given explicitly.
    keep_raw=True: a percentage-only update will NOT overwrite an existing row
    that already carries raw hours (lets FY history sit under live weeks)."""
    week_label = _norm_week(week_label)
    dept = _norm_dept(dept)
    if keep_raw and working is None:
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT id, working_hrs FROM wm_metrics WHERE "
                    f"week_label={PH} AND dept={PH}", (week_label, dept))
        row = cur.fetchone()
        if row and row[1] is not None:
            # row already carries computed hours. Apply ONLY a real non-zero FY
            # percentage; an empty FY cell (0/None) must not wipe the computed %.
            sets, vals = [], []
            if ot_pct is not None and ot_pct > 0:
                sets.append(f"ot_pct={PH}"); vals.append(ot_pct)
            if absent_pct is not None and absent_pct > 0:
                sets.append(f"absent_pct={PH}"); vals.append(absent_pct)
            if sets:
                vals.append(row[0])
                cur.execute(f"UPDATE wm_metrics SET {', '.join(sets)} "
                            f"WHERE id={PH}", vals)
                conn.commit()
            return
    if ot_pct is None and working is not None and ot is not None \
            and (ot + working) > 0:
        ot_pct = ot / (ot + working)
    if absent_pct is None and working is not None and leave is not None \
            and (working + leave) > 0:
        absent_pct = leave / (working + leave)
    args = (week_label, week_key(week_label), dept, working, ot, leave,
            ot_pct, absent_pct, source)
    conn = get_conn(); cur = conn.cursor()
    if IS_POSTGRES:
        cur.execute(f"""INSERT INTO wm_metrics
            (week_label,week_key,dept,working_hrs,ot_hrs,leave_hrs,
             ot_pct,absent_pct,source) VALUES ({','.join([PH] * 9)})
            ON CONFLICT (week_label,dept) DO UPDATE SET
              working_hrs=EXCLUDED.working_hrs, ot_hrs=EXCLUDED.ot_hrs,
              leave_hrs=EXCLUDED.leave_hrs, ot_pct=EXCLUDED.ot_pct,
              absent_pct=EXCLUDED.absent_pct, source=EXCLUDED.source""", args)
    else:
        cur.execute(f"""INSERT OR REPLACE INTO wm_metrics
            (week_label,week_key,dept,working_hrs,ot_hrs,leave_hrs,
             ot_pct,absent_pct,source) VALUES ({','.join([PH] * 9)})""", args)
    conn.commit()


def clear_all():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM wm_metrics"); conn.commit()


def delete_week(week_label):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"DELETE FROM wm_metrics WHERE week_label={PH}", (week_label,))
    conn.commit()


# ------------------------------- queries -----------------------------------
def weeks():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT DISTINCT week_label, week_key FROM wm_metrics "
                "ORDER BY week_key")
    return [r[0] for r in cur.fetchall()]


def latest_week():
    w = weeks()
    return w[-1] if w else None


def prev_week_of(week_label):
    w = weeks()
    if week_label in w:
        i = w.index(week_label)
        return w[i - 1] if i > 0 else None
    return None


def week_data(week_label):
    """Per-department rows for a week, ordered by DEPT_ORDER."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT dept,working_hrs,ot_hrs,leave_hrs,ot_pct,absent_pct,
                    source FROM wm_metrics WHERE week_label={PH}""",
                (week_label,))
    by = {r[0]: dict(dept=r[0], working=r[1], ot=r[2], leave=r[3],
                     ot_pct=r[4], absent_pct=r[5], source=r[6])
          for r in cur.fetchall()}
    order = [d for d in DEPT_ORDER if d in by] + \
            [d for d in by if d not in DEPT_ORDER]
    return [by[d] for d in order]


def trend(n=4):
    """Last n weeks -> (week_list, {dept: {week: {ot_pct, absent_pct}}})."""
    ws = weeks()[-n:]
    conn = get_conn(); cur = conn.cursor()
    data = {}
    for wk in ws:
        cur.execute(f"SELECT dept,ot_pct,absent_pct FROM wm_metrics "
                    f"WHERE week_label={PH}", (wk,))
        for dept, op, ap in cur.fetchall():
            data.setdefault(dept, {})[wk] = dict(ot_pct=op, absent_pct=ap)
    return ws, data


def overall(week_label):
    """Org totals + overall OT%/Absent% for a week. The percentages are a
    working-hours-weighted average of the per-department values (which carry the
    FY-tracker numbers), falling back to hours-derived when no % is present —
    this keeps the summary cards consistent with the per-department chart lines.
    """
    rows = week_data(week_label)
    w = sum(r["working"] or 0 for r in rows)
    o = sum(r["ot"] or 0 for r in rows)
    lv = sum(r["leave"] or 0 for r in rows)

    def _wavg(key, fallback):
        num = den = 0.0
        for r in rows:
            v = r.get(key)
            wt = r["working"] or 0
            if v is not None and wt > 0:
                num += v * wt; den += wt
        return (num / den) if den else fallback

    return {"working": w, "ot": o, "leave": lv,
            "ot_pct": _wavg("ot_pct", (o / (o + w)) if (o + w) else None),
            "absent_pct": _wavg("absent_pct",
                                (lv / (w + lv)) if (w + lv) else None)}


# --------------------------- seed from report ------------------------------
def _read_trend(wb, sheet):
    """{dept: {week: pct}} from an 'OT/Absenteeism Trend' sheet."""
    if sheet not in wb.sheetnames:
        return {}
    rows = list(wb[sheet].iter_rows(values_only=True))
    hdr_i = next((i for i, r in enumerate(rows)
                  if r and str(r[0]).strip() == "Department"), None)
    if hdr_i is None:
        return {}
    wlabels = [str(c).strip() for c in rows[hdr_i][1:] if c]
    out = {}
    for r in rows[hdr_i + 1:]:
        if not r or _is_total(r[0]):
            continue
        dept = str(r[0]).strip()
        out[dept] = {}
        for j, wk in enumerate(wlabels):
            v = r[1 + j] if 1 + j < len(r) else None
            if v is not None:
                out[dept][wk] = float(v)
    return out


def _read_by_dept(wb, sheet, value_label):
    """(current_week, [(dept, working, value), ...]) from a by-dept sheet."""
    rows = list(wb[sheet].iter_rows(values_only=True))
    cur_week = None
    for r in rows[:6]:
        for c in r:
            m = re.search(r"([A-Za-z]{3}\s*\d{2}-[Ww]\d)", str(c or ""))
            if m:
                cur_week = m.group(1).replace("  ", " ")
                break
        if cur_week:
            break
    hdr_i = next(i for i, r in enumerate(rows)
                 if r and str(r[0]).strip() == "Department")
    hdr = [str(c).strip() if c else "" for c in rows[hdr_i]]
    wi = hdr.index("Working Hrs")
    vi = hdr.index(value_label)
    out = []
    for r in rows[hdr_i + 1:]:
        if not r or _is_total(r[0]):
            continue
        out.append((str(r[0]).strip(), float(r[wi] or 0), float(r[vi] or 0)))
    return cur_week, out


def seed_from_report(file_bytes):
    """Seed weekly history from a 'Weekly Metric Report' .xlsx (the format
    produced by this system). Returns (current_week, rows_written)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    n = 0
    # 1) trends (percentages for ~4 weeks) — combine OT + Absenteeism trends
    ot_tr = _read_trend(wb, "OT Trend")
    ab_tr = _read_trend(wb, "Absenteeism Trend")
    for dept in set(ot_tr) | set(ab_tr):
        wks = set(ot_tr.get(dept, {})) | set(ab_tr.get(dept, {}))
        for wk in wks:
            upsert(wk, dept,
                   ot_pct=ot_tr.get(dept, {}).get(wk),
                   absent_pct=ab_tr.get(dept, {}).get(wk), source="seed")
            n += 1
    # 2) current week raw hours (overrides the trend row for that week)
    if "OT by Dept" in wb.sheetnames:
        cur_week, ot_rows = _read_by_dept(wb, "OT by Dept", "OT Hrs")
        leave_by = {}
        if "Absenteeism by Dept" in wb.sheetnames:
            _, ab_rows = _read_by_dept(wb, "Absenteeism by Dept", "Leave Hrs")
            leave_by = {d: lv for d, _w, lv in ab_rows}
        for dept, working, ot in ot_rows:
            upsert(cur_week, dept, working=working, ot=ot,
                   leave=leave_by.get(dept, 0.0), source="seed")
            n += 1
    return latest_week(), n


def seed_from_fy(file_bytes, sheet="Data"):
    """Seed the full weekly OT%/Absent% history from the FY tracker.
      OT block:        header row 9,  dept rows 10-20, week columns J(10)…
      Absenteeism:     header row 26, dept rows 27-37, same week columns.
    Empty cells (a department that did not yet exist that week) are stored as 0
    so the trend line reads zero rather than breaking. Percentage-only, so it
    never clobbers a live week that already has raw hours.
    Returns (first_week, last_week, rows_written)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True,
                                data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    def cell(r, c):
        return rows[r - 1][c - 1] if r - 1 < len(rows) \
            and c - 1 < len(rows[r - 1]) else None

    week_cols = []
    for c in range(10, ws.max_column + 1):       # J onward
        h = cell(9, c)
        if h is None or not str(h).strip():
            continue
        lbl = _fy_real_week(_norm_week(h))
        if lbl is None:                          # empty drag-fill padding
            continue
        week_cols.append((c, lbl))

    ot_rows = list(range(10, 21))                # 10..20  (11 depts)
    ab_rows = list(range(27, 38))                # 27..37  (11 depts)
    n = 0
    for c, wk in week_cols:
        ot_by, ab_by = {}, {}
        for r in ot_rows:
            d = cell(r, 1)
            if _is_total(d):
                continue
            v = cell(r, c)
            ot_by[_norm_dept(d)] = float(v) if v is not None else 0.0
        for r in ab_rows:
            d = cell(r, 1)
            if _is_total(d):
                continue
            v = cell(r, c)
            ab_by[_norm_dept(d)] = float(v) if v is not None else 0.0
        for d in (set(ot_by) | set(ab_by)):
            upsert(wk, d, ot_pct=ot_by.get(d, 0.0),
                   absent_pct=ab_by.get(d, 0.0), source="fy", keep_raw=True)
            n += 1
    first = week_cols[0][1] if week_cols else None
    last = week_cols[-1][1] if week_cols else None
    return first, last, n


def seed_file(file_bytes):
    """Auto-detect a Weekly-Metric-Report vs an FY tracker and seed it.
    Returns (kind, ...seed-return)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True,
                                data_only=True)
    names = wb.sheetnames
    if "OT by Dept" in names or "OT Trend" in names:
        return ("report",) + tuple(seed_from_report(file_bytes))
    if "Data" in names:
        return ("fy",) + tuple(seed_from_fy(file_bytes))
    # last resort: try report, then FY
    try:
        return ("report",) + tuple(seed_from_report(file_bytes))
    except Exception:
        return ("fy",) + tuple(seed_from_fy(file_bytes))
def compute_week(week_label, date_from, date_to, source="computed"):
    """Compute weekly metrics per department from the active timesheet for the
    date window [date_from, date_to] (ISO 'YYYY-MM-DD').

    Working / OT / Leave HOURS are summed from the actual report rows (the 3 HR
    reports). OT% / Absent% that already exist for the week (e.g. seeded from the
    FY tracker) are PRESERVED — they are only derived from the hours as a
    fallback when the week has no percentage yet. dept comes from the MASTER
    cost-centre (column G). Returns the number of departments written."""
    s = get_settings()
    daily = float(s["daily_hours"])
    excl = set(x.strip() for x in s["exclude_leave"].split(","))
    cc_map = get_cc_map()

    emp_dept = {}
    for r in edb.list_records("active"):
        cc = str(r.get("cost_centre") or "").strip()
        code = cc[:3] if len(cc) >= 3 else cc
        d = cc_map.get(code)
        if d:
            emp_dept[str(r.get("emp_no"))] = d

    work_by, ot_by, leave_by = {}, {}, {}
    for row in att.timesheet_rows():
        wd = row["work_date"]
        if wd < date_from or wd > date_to:
            continue
        d = emp_dept.get(row["emp_no"])
        if not d:
            continue
        work_by[d] = work_by.get(d, 0.0) + (row["normal_hours"] or 0)
        ot_by[d] = ot_by.get(d, 0.0) + sum(
            row[k] or 0 for k in ("ot1", "ot15", "ot2", "ot3"))
        lv = (row["sick"] or 0) + (row["personal"] or 0)
        if "annual" not in excl:
            lv += (row["annual"] or 0)
        leave_by[d] = leave_by.get(d, 0.0) + lv * daily

    depts = set(work_by) | set(ot_by) | set(leave_by)
    nlabel = _norm_week(week_label)
    conn = get_conn(); cur = conn.cursor()
    n = 0
    for dept in depts:
        ndept = _norm_dept(dept)
        w = work_by.get(dept, 0.0)
        o = ot_by.get(dept, 0.0)
        lv = leave_by.get(dept, 0.0)
        cur.execute(f"SELECT id, ot_pct, absent_pct FROM wm_metrics "
                    f"WHERE week_label={PH} AND dept={PH}", (nlabel, ndept))
        ex = cur.fetchone()
        if ex:
            op = ex[1] if (ex[1] is not None and ex[1] > 0) \
                else ((o / (o + w)) if (o + w) else None)
            ap = ex[2] if (ex[2] is not None and ex[2] > 0) \
                else ((lv / (w + lv)) if (w + lv) else None)
            cur.execute(f"UPDATE wm_metrics SET working_hrs={PH}, ot_hrs={PH}, "
                        f"leave_hrs={PH}, ot_pct={PH}, absent_pct={PH}, "
                        f"source={PH} WHERE id={PH}",
                        (w, o, lv, op, ap, source, ex[0]))
        else:
            op = (o / (o + w)) if (o + w) else None
            ap = (lv / (w + lv)) if (w + lv) else None
            cur.execute(f"INSERT INTO wm_metrics (week_label,week_key,dept,"
                        f"working_hrs,ot_hrs,leave_hrs,ot_pct,absent_pct,source)"
                        f" VALUES ({','.join([PH] * 9)})",
                        (nlabel, week_key(nlabel), ndept, w, o, lv, op, ap,
                         source))
        n += 1
    conn.commit()
    return n
