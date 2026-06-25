# lib/doc_templates.py
# ============================================================================
# DOCUMENTS v2 (§8)
# 1) TEMPLATE VERSIONING + OVERLAY-ON-ORIGINAL
#    Each printable form can carry uploaded background page images (the real
#    government/company form scans) + a field-position map. The printout then
#    lays the data ON the original at calibrated positions — files identically
#    to a hand-filled original (ISO 14001 paper records). When the standard
#    changes: upload a new version (effective date), old versions retained,
#    NO code edit. If no template is active, rendering falls back to the
#    built-in reconstruction in lib/print_docs.py.
# 2) COMPLETION CHECKLIST
#    Required-field sets per document -> % complete per employee, missing
#    drill-down, Excel export.
# ============================================================================
import base64
import datetime as dt
import html as H
import json

from lib.db import get_conn, IS_POSTGRES, PH
from lib import employee_db as edb

SERIAL = "SERIAL PRIMARY KEY" if IS_POSTGRES else \
         "INTEGER PRIMARY KEY AUTOINCREMENT"
BLOB = "BYTEA" if IS_POSTGRES else "BLOB"

DOC_KEYS = {
    "fm_hr_003": "FM-HR-003 Application for Employment",
    "addendum": "เอกสารแนบท้าย (Additional disclosures)",
    "sso103": "สปส.1-03 แบบขึ้นทะเบียนผู้ประกันตน",
    "ly01": "ล.ย.01 แบบแจ้งรายการเพื่อการหักลดหย่อน",
}


def migrate():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""CREATE TABLE IF NOT EXISTS doc_templates (
        id {SERIAL},
        doc_key TEXT NOT NULL,
        version TEXT NOT NULL,
        effective_date TEXT,
        field_map TEXT,                  -- JSON list (see below)
        active INTEGER NOT NULL DEFAULT 0,
        created_by TEXT, created_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS doc_template_pages (
        id {SERIAL},
        template_id INTEGER NOT NULL,
        page INTEGER NOT NULL,
        image {BLOB}, mime TEXT)""")
    conn.commit()


def _rows(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def _ts():
    return dt.datetime.now().isoformat(timespec="seconds")


# ---------------------------------------------------------------- templates
def save_template(doc_key, version, effective_date, field_map_json, pages,
                  actor, activate=True):
    """pages: [(image_bytes, mime), ...] in page order.
    field_map JSON format (list):
      [{"key":"emp_name_th","page":1,"x":22.5,"y":31.0,"size":13,
        "w":40,"align":"left"}, ...]
    x/y/w are PERCENT of the page; key is an employee-schema key or a
    computed key: th_day/th_month/th_year (Buddhist date parts of
    birth_day), today_th, id_d1..id_d13 (ID-card digit boxes)."""
    try:
        json.loads(field_map_json or "[]")
    except Exception as e:
        return None, f"field map JSON ไม่ถูกต้อง: {e}"
    conn = get_conn(); cur = conn.cursor()
    if activate:
        cur.execute(f"UPDATE doc_templates SET active=0 WHERE doc_key={PH}",
                    (doc_key,))
    cur.execute(f"""INSERT INTO doc_templates (doc_key, version,
                    effective_date, field_map, active, created_by,
                    created_at) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
                (doc_key, version, str(effective_date),
                 field_map_json or "[]", 1 if activate else 0, actor, _ts()))
    conn.commit()
    if IS_POSTGRES:
        cur.execute("SELECT MAX(id) FROM doc_templates")
        tid = cur.fetchone()[0]
    else:
        tid = cur.lastrowid
    for i, (img, mime) in enumerate(pages, 1):
        cur.execute(f"""INSERT INTO doc_template_pages (template_id, page,
                        image, mime) VALUES ({PH},{PH},{PH},{PH})""",
                    (tid, i, img, mime))
    conn.commit()
    edb._audit(conn, actor, "doc_template_save",
               detail={"doc": doc_key, "version": version})
    conn.commit()
    return tid, None


def list_templates(doc_key=None):
    conn = get_conn(); cur = conn.cursor()
    if doc_key:
        cur.execute(f"""SELECT id, doc_key, version, effective_date, active,
                        created_by, created_at FROM doc_templates
                        WHERE doc_key={PH} ORDER BY id DESC""", (doc_key,))
    else:
        cur.execute("""SELECT id, doc_key, version, effective_date, active,
                       created_by, created_at FROM doc_templates
                       ORDER BY doc_key, id DESC""")
    return _rows(cur)


def active_template(doc_key):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT * FROM doc_templates WHERE doc_key={PH}
                    AND active=1 ORDER BY id DESC LIMIT 1""", (doc_key,))
    r = cur.fetchone()
    if not r:
        return None
    cols = [d[0] for d in cur.description]
    t = dict(zip(cols, r)) if IS_POSTGRES else dict(r)
    cur.execute(f"""SELECT page, image, mime FROM doc_template_pages
                    WHERE template_id={PH} ORDER BY page""", (t["id"],))
    t["pages"] = [(p[1], p[2]) for p in cur.fetchall()]
    return t


def set_active(template_id, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT doc_key FROM doc_templates WHERE id={PH}",
                (template_id,))
    dk = cur.fetchone()[0]
    cur.execute(f"UPDATE doc_templates SET active=0 WHERE doc_key={PH}",
                (dk,))
    cur.execute(f"UPDATE doc_templates SET active=1 WHERE id={PH}",
                (template_id,))
    conn.commit()
    edb._audit(conn, actor, "doc_template_activate",
               detail={"id": template_id, "doc": dk})
    conn.commit()


# ---------------------------------------------------------------- values
def _th_parts(iso):
    try:
        d = dt.date.fromisoformat(str(iso)[:10])
        months = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม",
                  "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม",
                  "พฤศจิกายน", "ธันวาคม"]
        return str(d.day), months[d.month], str(d.year + 543)
    except Exception:
        return "", "", ""


def field_value(rec, key):
    """Schema key OR computed key."""
    if key in ("th_day", "th_month", "th_year"):
        d, m, y = _th_parts(rec.get("birth_day"))
        return {"th_day": d, "th_month": m, "th_year": y}[key]
    if key == "today_th":
        d, m, y = _th_parts(dt.date.today().isoformat())
        return f"{d} {m} {y}"
    if key.startswith("id_d"):
        digits = "".join(ch for ch in str(rec.get("id_card") or "")
                         if ch.isdigit())
        i = int(key[4:]) - 1
        return digits[i] if i < len(digits) else ""
    v = rec.get(key)
    return "" if v is None else str(v)


# ---------------------------------------------------------------- overlay
def render_overlay(rec, template, calibrate=False):
    """HTML: each page = the original image as background, values laid on
    top at % coordinates. calibrate=True draws a 5% grid for positioning."""
    fmap = json.loads(template.get("field_map") or "[]")
    pages_html = []
    for pno, (img, mime) in enumerate(template["pages"], 1):
        b = bytes(img) if not isinstance(img, bytes) else img
        uri = f"data:{mime};base64,{base64.b64encode(b).decode()}"
        spans = ""
        for f in fmap:
            if int(f.get("page", 1)) != pno:
                continue
            val = field_value(rec, f["key"])
            if not val and not calibrate:
                continue
            spans += (
                f'<div class="fv" style="left:{f.get("x",0)}%;'
                f'top:{f.get("y",0)}%;width:{f.get("w",30)}%;'
                f'font-size:{f.get("size",13)}px;'
                f'text-align:{f.get("align","left")}">'
                f'{H.escape(val) if val else ("["+f["key"]+"]" if calibrate else "")}'
                f'</div>')
        grid = ""
        if calibrate:
            for i in range(1, 20):
                grid += (f'<div class="gl" style="left:{i*5}%"></div>'
                         f'<div class="gh" style="top:{i*5}%"></div>'
                         f'<div class="gt" style="left:{i*5}%;top:0">'
                         f'{i*5}</div>'
                         f'<div class="gt" style="left:0;top:{i*5}%">'
                         f'{i*5}</div>')
        pages_html.append(
            f'<div class="page"><img src="{uri}">{spans}{grid}</div>')
    return f"""<!doctype html><html lang="th"><head><meta charset="utf-8">
<title>{H.escape(template["doc_key"])} {H.escape(template["version"])}</title>
<style>
@page{{size:A4 portrait;margin:0}}
body{{margin:0;font-family:'Sarabun','Leelawadee UI',Tahoma,sans-serif}}
.page{{position:relative;width:210mm;height:297mm;page-break-after:always;
 overflow:hidden}}
.page img{{position:absolute;inset:0;width:100%;height:100%;
 object-fit:fill}}
.fv{{position:absolute;color:#0b2a6b;font-weight:600;line-height:1.15;
 white-space:pre-wrap}}
.gl{{position:absolute;top:0;bottom:0;width:0;border-left:1px dashed
 rgba(220,38,38,.45)}}
.gh{{position:absolute;left:0;right:0;height:0;border-top:1px dashed
 rgba(220,38,38,.45)}}
.gt{{position:absolute;font-size:8px;color:#dc2626}}
.noprint{{position:fixed;top:8px;right:10px;z-index:9}}
.noprint button{{padding:8px 16px;border:0;border-radius:8px;color:#fff;
 font-weight:700;cursor:pointer;
 background:linear-gradient(135deg,#009ADE,#715091)}}
@media print{{.noprint{{display:none}}}}
</style></head><body onload="window.print()">
<div class="noprint"><button onclick="window.print()">🖨️ Print</button></div>
{''.join(pages_html)}
</body></html>"""


DEFAULT_FIELD_MAP_EXAMPLE = json.dumps([
    {"key": "emp_name_th", "page": 1, "x": 22, "y": 18, "size": 14, "w": 40},
    {"key": "id_d1", "page": 1, "x": 60.0, "y": 18, "size": 14, "w": 3,
     "align": "center"},
    {"key": "th_day", "page": 1, "x": 30, "y": 24, "size": 13, "w": 6},
    {"key": "th_month", "page": 1, "x": 38, "y": 24, "size": 13, "w": 14},
    {"key": "th_year", "page": 1, "x": 54, "y": 24, "size": 13, "w": 8},
], ensure_ascii=False, indent=1)


# ---------------------------------------------------------------- checklist
REQUIRED = {
    "ทะเบียนพนักงาน / Master": ["emp_no", "emp_name_th", "emp_name_en",
                                "title", "dept_location", "joined_date",
                                "mobile"],
    "FM-HR-003 ใบสมัคร": ["birth_day", "id_card", "marital_status",
                          "education", "cur_addr_no", "cur_addr_province",
                          "emergency_name", "emergency_phone"],
    "สปส.1-03": ["id_card", "birth_day", "sex", "nationality",
                 "hospital_choice_1"],
    "ล.ย.01": ["id_card", "emp_name_th"],
    "รูปถ่าย / Photo": ["photo"],
}


def checklist():
    """[{Emp, Name, Dept, %, missing:{group:[fields]}}] for active staff."""
    out = []
    total_fields = sum(len(v) for v in REQUIRED.values())
    from lib import employee_schema as schema
    for r in edb.list_records("active"):
        missing = {}
        filled = 0
        for grp, keys in REQUIRED.items():
            miss = []
            for k in keys:
                v = r.get(k)
                if v in (None, "", 0) and not (k == "photo" and r.get(k)):
                    f = schema.BY_KEY.get(k)
                    miss.append(f.th if f else k)
                else:
                    filled += 1
            if miss:
                missing[grp] = miss
        out.append({"Emp": r.get("emp_no"),
                    "Name": r.get("emp_name_en"),
                    "Dept": r.get("dept_location"),
                    "pct": round(100 * filled / total_fields),
                    "missing": missing})
    return sorted(out, key=lambda x: x["pct"])


def checklist_xlsx(rows):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "Doc completeness"
    ws.append(["Emp No.", "Name", "Department", "% complete",
               "Missing items"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="715091")
    for r in rows:
        miss = "; ".join(f"{g}: {', '.join(v)}"
                         for g, v in r["missing"].items())
        ws.append([r["Emp"], r["Name"], r["Dept"], r["pct"], miss])
    ws.column_dimensions["E"].width = 80
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()
