# lib/attendance_db.py
# ============================================================================
# TIME-ATTENDANCE REPORTS (§5) — built around your THREE standard files
# (layouts verified against the 12/06/2026 uploads):
#   1. รายงานผลการคำนวณบันทึกเวลาแสดงตามพนักงาน (timesheet calc)
#        emp id col 0 (forward-fill) · name col 2 · date col 7 (พ.ศ.) ·
#        shift col 9 · scans col 10 · normal hrs col 11 · late col 13 ·
#        early col 14 · OTx1/x1.5/x2/x3 cols 17-20 · absent col 21 ·
#        sick col 22 · personal col 24 · annual col 25
#   2. รายงานรายละเอียดขออนุมัติใบลา (leave detail)
#        dept header (แผนก:) · emp header (พนักงาน: id col 9, name col 13) ·
#        rows: doc col 8 · type col 14 · dates col 19 · time col 23 ·
#        days col 29 · status col 35
#   3. รายงานรายละเอียดขออนุมัติใบโอที (OT detail — the SHIFTED layout)
#        rows: doc col 8 · type col 14 · dates col 19 · time col 23 ·
#        hours col 29 · minutes col 31 · status col 34
# Each upload replaces the active snapshot per kind (history kept).
# Abnormality engine = the agreed May logic: late/early, no-show without
# leave, pending leave, OT-done-not-submitted (gap >= 0.5 h), pending OT,
# duplicate approved OT, org summary. Managers see ONLY their reporting
# subtree (Mgr-column chain).
# ============================================================================
import datetime as dt

from lib.db import get_conn, IS_POSTGRES, PH
from lib import employee_db as edb

SERIAL = "SERIAL PRIMARY KEY" if IS_POSTGRES else \
         "INTEGER PRIMARY KEY AUTOINCREMENT"


def _ts():
    return dt.datetime.now().isoformat(timespec="seconds")


def migrate():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""CREATE TABLE IF NOT EXISTS att_uploads (
        id {SERIAL},
        kind TEXT NOT NULL,            -- timesheet | leave | ot
        filename TEXT, period_from TEXT, period_to TEXT,
        n_rows INTEGER, active INTEGER NOT NULL DEFAULT 1,
        uploaded_by TEXT, uploaded_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS att_timesheet (
        id {SERIAL}, upload_id INTEGER NOT NULL,
        emp_no TEXT, emp_name TEXT, work_date TEXT, shift_code TEXT,
        scans TEXT, normal_hours REAL, late_min REAL, early_min REAL,
        ot1 REAL, ot15 REAL, ot2 REAL, ot3 REAL,
        absent REAL, sick REAL, personal REAL, annual REAL)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS att_requests (
        id {SERIAL}, upload_id INTEGER NOT NULL,
        kind TEXT NOT NULL,            -- leave | ot
        emp_no TEXT, emp_name TEXT, dept_code TEXT, dept_name TEXT,
        doc_no TEXT, req_type TEXT,
        date_start TEXT, date_end TEXT, time_range TEXT,
        days REAL, hours REAL, status TEXT, remark TEXT)""")
    cur.execute("""CREATE INDEX IF NOT EXISTS ix_ts_emp
                   ON att_timesheet (upload_id, emp_no, work_date)""")
    cur.execute("""CREATE INDEX IF NOT EXISTS ix_rq_emp
                   ON att_requests (upload_id, emp_no, date_start)""")
    conn.commit()


# ---------------------------------------------------------------- helpers
def _th_date_iso(s):
    """'03/01/2569' (พ.ศ.) -> '2026-01-03'."""
    try:
        d, m, y = str(s).strip().split("/")
        return f"{int(y) - 543:04d}-{int(m):02d}-{int(d):02d}"
    except Exception:
        return None


def _f(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def _rows(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def active_upload(kind):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT * FROM att_uploads WHERE kind={PH} AND active=1
                    ORDER BY id DESC LIMIT 1""", (kind,))
    r = cur.fetchone()
    if not r:
        return None
    cols = [d[0] for d in cur.description]
    return dict(zip(cols, r)) if IS_POSTGRES else dict(r)


def list_uploads(kind):
    """Full upload history for a report kind (newest first) — powers the
    admin file-source chooser (req. 10)."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM att_uploads WHERE kind={PH} ORDER BY id DESC",
                (kind,))
    return _rows(cur)


def set_active(kind, upload_id):
    """Make a specific previously-uploaded file the active source for a kind
    (so admin can revert to an earlier good file)."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"UPDATE att_uploads SET active=0 WHERE kind={PH}", (kind,))
    cur.execute(f"UPDATE att_uploads SET active=1 WHERE id={PH} AND kind={PH}",
                (int(upload_id), kind))
    conn.commit()
    return cur.rowcount > 0


def _new_upload(kind, filename, n, pfrom, pto, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"UPDATE att_uploads SET active=0 WHERE kind={PH}", (kind,))
    cur.execute(f"""INSERT INTO att_uploads (kind, filename, period_from,
                    period_to, n_rows, active, uploaded_by, uploaded_at)
                    VALUES ({PH},{PH},{PH},{PH},{PH},1,{PH},{PH})""",
                (kind, filename, pfrom, pto, n, actor, _ts()))
    conn.commit()
    if IS_POSTGRES:
        cur.execute("SELECT MAX(id) FROM att_uploads WHERE kind=%s", (kind,))
        return cur.fetchone()[0]
    return cur.lastrowid


# ---------------------------------------------------------------- parsers
def import_timesheet(file_bytes, filename, actor):
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sh = wb.sheet_by_index(0)
    rows, emp_no, emp_name = [], None, None
    dates = []
    for r in range(8, sh.nrows):
        v0 = str(sh.cell_value(r, 0)).strip()
        if v0 and v0.replace(".0", "").isdigit():
            emp_no = v0.replace(".0", "")
            emp_name = str(sh.cell_value(r, 2)).strip()
        d = _th_date_iso(sh.cell_value(r, 7))
        if not d or not emp_no:
            continue
        dates.append(d)
        rows.append((emp_no, emp_name, d,
                     str(sh.cell_value(r, 9)).strip(),
                     str(sh.cell_value(r, 10)).strip(),
                     _f(sh.cell_value(r, 11)), _f(sh.cell_value(r, 13)),
                     _f(sh.cell_value(r, 14)), _f(sh.cell_value(r, 17)),
                     _f(sh.cell_value(r, 18)), _f(sh.cell_value(r, 19)),
                     _f(sh.cell_value(r, 20)), _f(sh.cell_value(r, 21)),
                     _f(sh.cell_value(r, 22)), _f(sh.cell_value(r, 24)),
                     _f(sh.cell_value(r, 25))))
    uid = _new_upload("timesheet", filename, len(rows),
                      min(dates) if dates else None,
                      max(dates) if dates else None, actor)
    conn = get_conn(); cur = conn.cursor()
    cur.executemany(
        f"""INSERT INTO att_timesheet (upload_id, emp_no, emp_name,
            work_date, shift_code, scans, normal_hours, late_min, early_min,
            ot1, ot15, ot2, ot3, absent, sick, personal, annual)
            VALUES ({uid},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},
            {PH},{PH},{PH},{PH},{PH},{PH})""", rows)
    conn.commit()
    return uid, len(rows)


def _import_requests(kind, file_bytes, filename, actor, status_col,
                     days_col, hours_col, minutes_col=None):
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sh = wb.sheet_by_index(0)
    rows = []
    dept_code = dept_name = emp_no = emp_name = None
    dates = []
    for r in range(8, sh.nrows):
        c1 = str(sh.cell_value(r, 1)).strip()
        # The "พนักงาน :" (employee) label sits in a column that SHIFTS between
        # report layouts — col 6 in the OT report but col 5 in the leave report.
        # Detect it in either position (the employee code stays in col 9). The
        # dept header label stays in col 1.
        c5 = str(sh.cell_value(r, 5)).strip()
        c6 = str(sh.cell_value(r, 6)).strip()
        if c1.startswith("แผนก"):
            dept_code = str(sh.cell_value(r, 5)).strip()
            dept_name = str(sh.cell_value(r, 10)).strip()
            continue
        if c5.startswith("พนักงาน") or c6.startswith("พนักงาน"):
            emp_no = str(sh.cell_value(r, 9)).strip().replace(".0", "")
            emp_name = str(sh.cell_value(r, 13)).strip()
            continue
        doc = str(sh.cell_value(r, 8)).strip()
        if not doc or not emp_no:
            continue
        dr = str(sh.cell_value(r, 19)).strip()
        ds = de = None
        if " - " in dr:
            a, b = dr.split(" - ", 1)
            ds, de = _th_date_iso(a), _th_date_iso(b)
        if ds:
            dates.append(ds)
        # Hours = whole-hours column (+ optional minutes column / 60). The OT
        # report carries hours and minutes in SEPARATE columns; reading only the
        # hours column silently drops the minutes portion of every OT record.
        hours = _f(sh.cell_value(r, hours_col)) if hours_col is not None else 0.0
        if minutes_col is not None:
            hours += _f(sh.cell_value(r, minutes_col)) / 60.0
        rows.append((kind, emp_no, emp_name, dept_code, dept_name, doc,
                     str(sh.cell_value(r, 14)).strip(), ds, de,
                     str(sh.cell_value(r, 23)).strip(),
                     _f(sh.cell_value(r, days_col))
                     if days_col is not None else 0.0,
                     hours,
                     str(sh.cell_value(r, status_col)).strip(),
                     str(sh.cell_value(r, status_col + 1)).strip()))
    uid = _new_upload(kind, filename, len(rows),
                      min(dates) if dates else None,
                      max(dates) if dates else None, actor)
    conn = get_conn(); cur = conn.cursor()
    cur.executemany(
        f"""INSERT INTO att_requests (upload_id, kind, emp_no, emp_name,
            dept_code, dept_name, doc_no, req_type, date_start, date_end,
            time_range, days, hours, status, remark)
            VALUES ({uid},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},
            {PH},{PH},{PH},{PH})""", rows)
    conn.commit()
    return uid, len(rows)


def import_leave(file_bytes, filename, actor):
    # leave: quantity(days) col 27, unit col 29, status col 31
    return _import_requests("leave", file_bytes, filename, actor,
                            status_col=31, days_col=27, hours_col=None)


def import_ot(file_bytes, filename, actor):
    # OT: hours col 29 + minutes col 31, status col 34
    return _import_requests("ot", file_bytes, filename, actor,
                            status_col=34, days_col=None, hours_col=29,
                            minutes_col=31)


# ---------------------------------------------------------------- data access
def timesheet_rows(emp_nos=None):
    up = active_upload("timesheet")
    if not up:
        return []
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM att_timesheet WHERE upload_id={PH}",
                (up["id"],))
    rows = _rows(cur)
    if emp_nos is not None:
        rows = [r for r in rows if r["emp_no"] in emp_nos]
    return rows


def request_rows(kind, emp_nos=None):
    up = active_upload(kind)
    if not up:
        return []
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM att_requests WHERE upload_id={PH} "
                f"AND kind={PH}", (up["id"], kind))
    rows = _rows(cur)
    if emp_nos is not None:
        rows = [r for r in rows if r["emp_no"] in emp_nos]
    return rows


# ---------------------------------------------------------------- team scope
def _clean(n):
    s = str(n or "")
    for t in ("Mr.", "Ms.", "Mrs.", "Miss"):
        s = s.replace(t, "")
    return " ".join(s.split())


def subordinate_emp_nos(mgr_rec, include_self=True):
    """All emp_nos in the reporting subtree of this manager (Mgr column),
    matching the org-chart logic."""
    act = edb.list_records("active")
    by_mgr = {}
    for r in act:
        by_mgr.setdefault(_clean(r.get("mgr_name")).lower(), []).append(r)
    root = _clean(mgr_rec.get("emp_name_en")).lower()
    out, seen = set(), set()
    if include_self:
        out.add(str(mgr_rec.get("emp_no")))
    stack = [root]
    while stack:
        nm = stack.pop()
        if nm in seen:
            continue
        seen.add(nm)
        for r in by_mgr.get(nm, []):
            out.add(str(r.get("emp_no")))
            stack.append(_clean(r.get("emp_name_en")).lower())
    return out


# ---------------------------------------------------------------- engine
def _approved_leave_dates(emp_nos=None):
    """{(emp_no, iso_date)} covered by APPROVED leave."""
    cov = set()
    for r in request_rows("leave", emp_nos):
        if "อนุมัติ" not in (r["status"] or "") or \
                "ไม่อนุมัติ" in (r["status"] or ""):
            continue
        try:
            d0 = dt.date.fromisoformat(r["date_start"])
            d1 = dt.date.fromisoformat(r["date_end"] or r["date_start"])
        except Exception:
            continue
        d = d0
        while d <= d1:
            cov.add((r["emp_no"], d.isoformat()))
            d += dt.timedelta(days=1)
    return cov


def _any_leave_dates(emp_nos=None):
    cov = set()
    for r in request_rows("leave", emp_nos):
        if "ยกเลิก" in (r["status"] or ""):
            continue
        try:
            d0 = dt.date.fromisoformat(r["date_start"])
            d1 = dt.date.fromisoformat(r["date_end"] or r["date_start"])
        except Exception:
            continue
        d = d0
        while d <= d1:
            cov.add((r["emp_no"], d.isoformat()))
            d += dt.timedelta(days=1)
    return cov


def abnormalities(emp_nos=None, run_date=None):
    """All checks for the active snapshots, optionally scoped to a team.
    Returns dict of lists. run_date suppresses same-day partial-scan
    false positives (data pulled mid-shift)."""
    ts = timesheet_rows(emp_nos)
    ots = request_rows("ot", emp_nos)
    lvs = request_rows("leave", emp_nos)
    appr_leave = _approved_leave_dates(emp_nos)
    any_leave = _any_leave_dates(emp_nos)
    run_date = str(run_date or "")

    out = {"late_early": [], "noshow": [], "leave_pending": [],
           "ot_not_submitted": [], "ot_pending": [], "ot_duplicate": []}

    # R1A late / leave-early (the file's own computed minutes)
    for r in ts:
        if (r["late_min"] or 0) > 0 or (r["early_min"] or 0) > 0:
            if (r["emp_no"], r["work_date"]) in appr_leave:
                continue
            out["late_early"].append({
                "Emp": r["emp_no"], "Name": r["emp_name"],
                "Date": r["work_date"], "Shift": r["shift_code"],
                "Late (min)": r["late_min"], "Early (min)": r["early_min"],
                "Scans": r["scans"]})

    # R2 no-show: absent flag & no leave request at all (= ขาดงานไม่มีใบลา)
    for r in ts:
        if (r["absent"] or 0) > 0 and r["work_date"] != run_date:
            has_any = (r["emp_no"], r["work_date"]) in any_leave
            if (r["emp_no"], r["work_date"]) in appr_leave:
                continue
            out["noshow"].append({
                "Emp": r["emp_no"], "Name": r["emp_name"],
                "Date": r["work_date"], "Shift": r["shift_code"],
                "ใบลา": "รออนุมัติ" if has_any else "ไม่มีใบลา"})

    # R3 leave pending
    for r in lvs:
        if "รอ" in (r["status"] or ""):
            out["leave_pending"].append({
                "Emp": r["emp_no"], "Name": r["emp_name"],
                "Doc": r["doc_no"], "Type": r["req_type"],
                "From": r["date_start"], "To": r["date_end"],
                "Days": r["days"], "Status": r["status"]})

    # R4 OT done but not submitted (gap >= 0.5 h per emp/date)
    sub = {}
    for r in ots:
        if "ยกเลิก" in (r["status"] or "") or "ไม่อนุมัติ" in (r["status"]
                                                               or ""):
            continue
        key = (r["emp_no"], r["date_start"])
        sub[key] = sub.get(key, 0.0) + (r["hours"] or 0)
    for r in ts:
        done = sum(r[k] or 0 for k in ("ot1", "ot15", "ot2", "ot3"))
        if done <= 0:
            continue
        got = sub.get((r["emp_no"], r["work_date"]), 0.0)
        if done - got >= 0.5:
            out["ot_not_submitted"].append({
                "Emp": r["emp_no"], "Name": r["emp_name"],
                "Date": r["work_date"], "TS OT (h)": round(done, 2),
                "Submitted (h)": round(got, 2),
                "Gap (h)": round(done - got, 2)})

    # R5 OT pending
    for r in ots:
        if "รอ" in (r["status"] or ""):
            out["ot_pending"].append({
                "Emp": r["emp_no"], "Name": r["emp_name"],
                "Doc": r["doc_no"], "Type": r["req_type"],
                "Date": r["date_start"], "Time": r["time_range"],
                "Hours": r["hours"], "Status": r["status"]})

    # R6 duplicate APPROVED OT: same emp+date, overlapping time ranges
    def _rng(t):
        try:
            a, b = t.replace(" น.", "").split(" - ")
            f = lambda x: int(x[:2]) * 60 + int(x[3:5])
            s_, e_ = f(a), f(b)
            return (s_, e_ + (1440 if e_ <= s_ else 0))
        except Exception:
            return None
    appr_ot = [r for r in ots if "อนุมัติ" in (r["status"] or "")
               and "ไม่อนุมัติ" not in (r["status"] or "")]
    by_day = {}
    for r in appr_ot:
        by_day.setdefault((r["emp_no"], r["date_start"]), []).append(r)
    for (emp, day), grp in by_day.items():
        if len(grp) < 2:
            continue
        for i in range(len(grp)):
            for j in range(i + 1, len(grp)):
                a, b = _rng(grp[i]["time_range"]), _rng(grp[j]["time_range"])
                if a and b and a[0] < b[1] and b[0] < a[1]:
                    out["ot_duplicate"].append({
                        "Emp": emp, "Name": grp[i]["emp_name"], "Date": day,
                        "Doc A": grp[i]["doc_no"],
                        "Time A": grp[i]["time_range"],
                        "Doc B": grp[j]["doc_no"],
                        "Time B": grp[j]["time_range"]})
    return out


def team_summary(emp_nos=None, date_from=None, date_to=None):
    """KPI roll-up per employee for the dashboard. Optional date_from/date_to
    (ISO 'YYYY-MM-DD') restrict the rows to a weekly or monthly window."""
    ts = timesheet_rows(emp_nos)
    agg = {}
    for r in ts:
        wd = r["work_date"]
        if date_from and wd < date_from:
            continue
        if date_to and wd > date_to:
            continue
        a = agg.setdefault(r["emp_no"], {
            "Emp": r["emp_no"], "Name": r["emp_name"], "Days": 0,
            "Hours": 0.0, "OT h": 0.0, "Late (min)": 0.0,
            "Absent": 0.0, "Sick": 0.0, "Personal": 0.0, "Annual": 0.0})
        if (r["normal_hours"] or 0) > 0:
            a["Days"] += 1
        a["Hours"] += r["normal_hours"] or 0
        a["OT h"] += sum(r[k] or 0 for k in ("ot1", "ot15", "ot2", "ot3"))
        a["Late (min)"] += r["late_min"] or 0
        a["Absent"] += r["absent"] or 0
        a["Sick"] += r["sick"] or 0
        a["Personal"] += r["personal"] or 0
        a["Annual"] += r["annual"] or 0
    rows = list(agg.values())
    for r in rows:
        for k in ("Hours", "OT h", "Late (min)"):
            r[k] = round(r[k], 1)
    return sorted(rows, key=lambda r: r["Emp"])


def export_xlsx(abn, summary):
    """One Excel: summary + a sheet per abnormality list."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    H = Font(bold=True, color="FFFFFF")
    F = PatternFill("solid", fgColor="715091")

    def sheet(name, rows):
        ws = wb.create_sheet(name[:28])
        if not rows:
            ws.append(["(none)"])
            return
        cols = list(rows[0].keys())
        ws.append(cols)
        for c in ws[1]:
            c.font = H; c.fill = F
        for r in rows:
            ws.append([r.get(c) for c in cols])

    wb.remove(wb.active)
    sheet("Summary", summary)
    titles = {"late_early": "R1 Late-Early", "noshow": "R2 NoShow",
              "leave_pending": "R3 LeavePending",
              "ot_not_submitted": "R4 OT_NotSubmitted",
              "ot_pending": "R5 OT_Pending",
              "ot_duplicate": "R6 OT_Duplicate"}
    for k, t in titles.items():
        sheet(t, abn.get(k, []))
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def noshow_runs(emp_nos=None, min_days=2):
    """Consecutive missed WORKING days per employee from the active
    timesheet snapshot (LPA s.119(5) watchdog basis).
    Day classes: missed = absent flag > 0 · worked = hours/scans ·
    on_leave = sick/personal/annual > 0 · off = none of those (scheduled
    day off — skipped WITHOUT breaking the run, weekends/holidays).
    Cross-checks the leave snapshot: a date covered by ANY non-cancelled
    leave request does not count as missed.
    Returns runs with: emp, name, start, end, days, open (run reaches the
    employee's last snapshot date => still ongoing)."""
    ts = timesheet_rows(emp_nos)
    any_leave = _any_leave_dates(emp_nos)
    by_emp = {}
    for r in ts:
        by_emp.setdefault(r["emp_no"], []).append(r)
    runs = []
    for emp, rows in by_emp.items():
        rows.sort(key=lambda r: r["work_date"])
        last_date = rows[-1]["work_date"]
        cur_run = []
        def _close(open_flag_date=None):
            nonlocal cur_run
            if len(cur_run) >= min_days:
                runs.append({
                    "emp_no": emp, "emp_name": cur_run[0]["emp_name"],
                    "start": cur_run[0]["work_date"],
                    "end": cur_run[-1]["work_date"],
                    "days": len(cur_run),
                    "open": cur_run[-1]["work_date"] == last_date})
            cur_run = []
        for r in rows:
            missed = (r["absent"] or 0) > 0 and \
                (emp, r["work_date"]) not in any_leave
            worked = (r["normal_hours"] or 0) > 0 or bool(r["scans"])
            on_leave = ((r["sick"] or 0) + (r["personal"] or 0)
                        + (r["annual"] or 0)) > 0 or \
                (emp, r["work_date"]) in any_leave
            if missed:
                cur_run.append(r)
            elif worked or on_leave:
                _close()
            # else: scheduled off day — neither extends nor breaks
        _close()
    return sorted(runs, key=lambda x: (-x["open"], -x["days"]))


# ---------------------------------------------------------------- dashboards
def monthly_trends(emp_nos=None):
    """{month: {hours, ot, absent_days, leave_days, late_min}} for charts."""
    out = {}
    for r in timesheet_rows(emp_nos):
        m = r["work_date"][:7]
        a = out.setdefault(m, {"ชม.งาน": 0.0, "ชม.OT": 0.0,
                               "ขาดงาน(วัน)": 0.0, "ลา(วัน)": 0.0,
                               "สาย(นาที)": 0.0})
        a["ชม.งาน"] += r["normal_hours"] or 0
        a["ชม.OT"] += sum(r[k] or 0 for k in ("ot1", "ot15", "ot2", "ot3"))
        a["ขาดงาน(วัน)"] += r["absent"] or 0
        a["ลา(วัน)"] += (r["sick"] or 0) + (r["personal"] or 0) + \
            (r["annual"] or 0)
        a["สาย(นาที)"] += r["late_min"] or 0
    return dict(sorted(out.items()))


def dept_comparison(emp_nos=None):
    """Per-department roll-up using the employee master mapping."""
    by_emp = {str(r.get("emp_no")): (r.get("dept_location") or "—")
              for r in edb.list_records("active")}
    out = {}
    for r in timesheet_rows(emp_nos):
        d = by_emp.get(r["emp_no"], "—")
        a = out.setdefault(d, {"แผนก": d, "ชม.งาน": 0.0, "ชม.OT": 0.0,
                               "ขาด(วัน)": 0.0, "สาย(นาที)": 0.0,
                               "คน": set()})
        a["ชม.งาน"] += r["normal_hours"] or 0
        a["ชม.OT"] += sum(r[k] or 0 for k in ("ot1", "ot15", "ot2", "ot3"))
        a["ขาด(วัน)"] += r["absent"] or 0
        a["สาย(นาที)"] += r["late_min"] or 0
        a["คน"].add(r["emp_no"])
    rows = []
    for a in out.values():
        n = max(len(a["คน"]), 1)
        rows.append({"แผนก": a["แผนก"], "คน": len(a["คน"]),
                     "ชม.งาน": round(a["ชม.งาน"], 0),
                     "ชม.OT": round(a["ชม.OT"], 0),
                     "OT/คน": round(a["ชม.OT"] / n, 1),
                     "ขาด(วัน)": a["ขาด(วัน)"],
                     "สาย(นาที)": round(a["สาย(นาที)"], 0)})
    return sorted(rows, key=lambda r: -r["ชม.OT"])


def rankings(emp_nos=None, top=10, date_from=None, date_to=None):
    """Top-N people by OT hours, late minutes, absence days. Optional
    date_from/date_to restrict to a weekly or monthly window."""
    agg = {}
    for r in timesheet_rows(emp_nos):
        wd = r["work_date"]
        if date_from and wd < date_from:
            continue
        if date_to and wd > date_to:
            continue
        a = agg.setdefault(r["emp_no"], {"name": r["emp_name"], "ot": 0.0,
                                         "late": 0.0, "absent": 0.0})
        a["ot"] += sum(r[k] or 0 for k in ("ot1", "ot15", "ot2", "ot3"))
        a["late"] += r["late_min"] or 0
        a["absent"] += r["absent"] or 0

    def pick(key):
        rows = sorted(agg.items(), key=lambda kv: -kv[1][key])[:top]
        return {f"{v['name'][:18]}": round(v[key], 1) for k, v in rows
                if v[key] > 0}
    return {"ot": pick("ot"), "late": pick("late"), "absent": pick("absent")}


def employee_daily(emp_no, date_from=None, date_to=None):
    """Daily series + requests for one person (drill-down). Optional
    date_from/date_to restrict to a weekly or monthly window."""
    ts = sorted(timesheet_rows({str(emp_no)}),
                key=lambda r: r["work_date"])
    if date_from:
        ts = [r for r in ts if r["work_date"] >= date_from]
    if date_to:
        ts = [r for r in ts if r["work_date"] <= date_to]
    daily = {r["work_date"]: {"ชม.งาน": r["normal_hours"] or 0,
                              "ชม.OT": sum(r[k] or 0 for k in
                                           ("ot1", "ot15", "ot2", "ot3"))}
             for r in ts}
    leaves = request_rows("leave", {str(emp_no)})
    ots = request_rows("ot", {str(emp_no)})
    if date_from or date_to:
        def _in(d):
            d = d or ""
            return ((not date_from or d >= date_from)
                    and (not date_to or d <= date_to))
        leaves = [l for l in leaves if _in(l.get("date_start"))]
        ots = [o for o in ots if _in(o.get("date_start"))]
    return ts, daily, leaves, ots
