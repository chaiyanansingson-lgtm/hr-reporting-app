# lib/employee_excel.py
# ============================================================================
# Requirement 3:
#   - Download the employee-list TEMPLATE (mirrors "Copy of Employee List
#     MASTER V.6", sheet "Headcount Updated": header rows 4-5, data row 6+)
#   - BULK UPLOAD updates from a filled template (also accepts the original
#     MASTER V.6 file directly — same header detection)
#   - Download INTERNAL export (all fields except the salary tier)
#   - Download EXTERNAL export (PDPA-protected fields excluded)
# ============================================================================

import io
import datetime as dt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from lib.employee_schema import (FIELDS, BY_KEY, INTERNAL_KEYS,
                                 EXTERNAL_KEYS, SALARY_KEYS)
from lib import employee_db as edb

ANCA_BLUE = "009ADE"
ANCA_PURPLE = "715091"

def _nh(s):
    """Normalize a header: strip + collapse whitespace."""
    return " ".join(str(s).split())


# Map: MASTER V.6 header text (row4 or "row4|row5") -> field key.
# Keys are normalized; ALIASES cover merged-cell quirks in the real file
# (e.g. the emergency-contact 'Name' sub-header sits one column before the
# 'Contact Emergency' top header anchor).
MASTER_MAP = {_nh(f.master_col): f.key for f in FIELDS if f.master_col}
MASTER_MAP.update({
    "Current Address|Name": "emergency_name",   # merged-cell offset in V.6
    "SSO|Inform": "sso_inform",
    "Personal|Email": "personal_email",
    "Transfer from sub to sub company": "transfer_sub",
    "Educational": "education",
})


# ---------------------------------------------------------------- template
def build_template(include_salary=False) -> bytes:
    """One header row, one column per field, EN + TH label, colour-coded:
    blue = identity, purple = PDPA, dark = salary tier (super admin)."""
    wb = Workbook(); ws = wb.active; ws.title = "Employee Upload"
    keys = INTERNAL_KEYS + (SALARY_KEYS if include_salary else [])
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for c, k in enumerate(keys, 1):
        f = BY_KEY[k]
        cell = ws.cell(row=1, column=c, value=f"{f.en}\n{f.th}")
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        fill = ANCA_BLUE
        if f.salary:
            fill = "2B2F3A"
        elif f.pdpa:
            fill = ANCA_PURPLE
        cell.fill = PatternFill("solid", start_color=fill)
        ws.cell(row=2, column=c, value=k).font = Font(
            italic=True, size=8, color="808080", name="Arial")
        ws.column_dimensions[cell.column_letter].width = max(
            14, min(len(f.en) + 2, 30))
    ws.freeze_panes = "B3"
    info = wb.create_sheet("READ ME")
    info["A1"] = "AMS HRM — Employee bulk upload template"
    info["A1"].font = Font(bold=True, size=12, name="Arial")
    notes = [
        "Row 1 = labels (EN/TH). Row 2 = field keys — DO NOT edit rows 1-2.",
        "Data starts on row 3. 'emp_no' identifies the employee:",
        "  - existing emp_no  -> that employee's record is UPDATED",
        "  - new emp_no       -> a NEW active employee is CREATED",
        "Dates: YYYY-MM-DD.",
        "Purple columns are PDPA-protected (excluded from External export).",
        "Dark columns (salary) appear only in the Super Admin template.",
        "You can also upload the original 'Employee List MASTER' file "
        "directly — the system detects its 'Emp. No.' header row.",
        "Every change is written to the audit log with your username.",
    ]
    for i, n in enumerate(notes, 3):
        info.cell(row=i, column=1, value=n).font = Font(name="Arial", size=10)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- bulk upload
def _detect_format(ws):
    """Return ('flat', header_row) for our template (field keys on row 2),
    or ('master', header_row) when an 'Emp. No.' header row is found
    (works for MASTER V.6 'Headcount Updated' where it's row 4)."""
    for r in range(1, 8):
        vals = [str(c.value).strip() if c.value is not None else ""
                for c in ws[r]]
        if "emp_no" in vals:
            return "flat", r
        if "Emp. No." in vals:
            return "master", r
    raise ValueError("Could not find a header row ('emp_no' or 'Emp. No.') "
                     "in the first 7 rows of the sheet.")


def _norm(v, f):
    if v is None:
        return None
    if f.typ == "date":
        if isinstance(v, dt.datetime):
            return v.date().isoformat()
        if isinstance(v, dt.date):
            return v.isoformat()
        return str(v).strip()[:10] or None
    if f.typ in ("int", "float"):
        try:
            return (int(v) if f.typ == "int" else float(v))
        except (TypeError, ValueError):
            return None
    s = str(v).strip()
    return s or None


def parse_upload(file_bytes: bytes, sheet=None, allow_salary=False):
    """Returns (rows, fmt, colmap_used). rows = list of {field: value} dicts
    keyed by emp_no. Salary columns are silently dropped unless
    allow_salary=True (super admin) — so a normal admin can NEVER smuggle
    salary changes through bulk upload."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet] if sheet else wb.active
    # MASTER V.6 lands on "Headcount Updated" if present and no sheet given
    if sheet is None and "Headcount Updated" in wb.sheetnames:
        ws = wb["Headcount Updated"]
    fmt, hr = _detect_format(ws)

    colmap = {}  # column index (0-based) -> field key
    if fmt == "flat":
        for c, cell in enumerate(ws[hr]):
            k = str(cell.value).strip() if cell.value else ""
            if k in BY_KEY:
                colmap[c] = k
    else:  # master layout: combine header row + sub-header row (hr+1)
        top = [_nh(c.value) if c.value is not None else "" for c in ws[hr]]
        sub = [_nh(c.value) if c.value is not None else ""
               for c in ws[hr + 1]]
        # carry top header rightward across merged cells
        carry = ""
        for i in range(len(top)):
            if top[i]:
                carry = top[i]
            combined = f"{carry}|{sub[i]}" if i < len(sub) and sub[i] else \
                (top[i] or carry if top[i] else "")
            for probe in ([top[i], combined] if top[i] else [combined]):
                if probe in MASTER_MAP:
                    colmap[i] = MASTER_MAP[probe]
        hr += 1  # data starts after the sub-header row

    rows = []
    for r in ws.iter_rows(min_row=hr + 1, values_only=True):
        if fmt == "flat" and r and str(r[0] or "").strip() in BY_KEY:
            continue  # skip the field-key row itself
        rec = {}
        for c, k in colmap.items():
            if c < len(r):
                f = BY_KEY[k]
                if f.salary and not allow_salary:
                    continue
                rec[k] = _norm(r[c], f)
        if rec.get("emp_no"):
            rows.append(rec)
    return rows, fmt, colmap


def apply_upload(rows, actor):
    """Upsert by emp_no. Returns summary dict; everything audited."""
    created, updated, unchanged = 0, 0, 0
    for rec in rows:
        existing = edb.get_record(emp_no=str(rec["emp_no"]))
        rec = {k: v for k, v in rec.items() if v is not None}
        if existing:
            diff = edb.update_record(existing["id"], rec, actor)
            updated += 1 if diff else 0
            unchanged += 0 if diff else 1
        else:
            edb.create_record(rec, actor, record_status="active")
            created += 1
    conn = edb.get_conn()
    edb._audit(conn, actor, "bulk_upload",
               detail={"created": created, "updated": updated,
                       "unchanged": unchanged})
    conn.commit()
    return {"created": created, "updated": updated, "unchanged": unchanged}


# ---------------------------------------------------------------- exports
def _export(keys, records, title) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = title[:31]
    ws.append([f"ANCA Manufacturing Solutions (Thailand) — {title}"])
    ws["A1"].font = Font(bold=True, size=12, name="Arial",
                         color=ANCA_PURPLE)
    ws.append([f"Generated {dt.datetime.now():%Y-%m-%d %H:%M}  •  "
               f"{len(records)} records"])
    ws.append([])
    hdr = [BY_KEY[k].en for k in keys]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=4, column=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=ANCA_BLUE)
    for rec in records:
        ws.append([rec.get(k) for k in keys])
    ws.freeze_panes = "A5"
    for c, k in enumerate(keys, 1):
        ws.column_dimensions[ws.cell(row=4, column=c).column_letter].width = \
            max(12, min(len(BY_KEY[k].en) + 3, 32))
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def export_internal(record_status="active", actor="system",
                    include_salary=False) -> bytes:
    """Internal HR export. include_salary=True only reachable behind the
    employee.view_salary capability (Super Admin)."""
    keys = (INTERNAL_KEYS + SALARY_KEYS) if include_salary else INTERNAL_KEYS
    recs = edb.list_records(record_status)
    conn = edb.get_conn()
    edb._audit(conn, actor, "export_internal",
               detail={"records": len(recs), "salary": include_salary})
    conn.commit()
    return _export(keys, recs, "Employee List (Internal)")


def export_external(record_status="active", actor="system") -> bytes:
    """External export — every PDPA-flagged field and the whole salary tier
    are excluded by construction (driven by the schema flags, requirement 3)."""
    recs = edb.list_records(record_status)
    conn = edb.get_conn()
    edb._audit(conn, actor, "export_external", detail={"records": len(recs)})
    conn.commit()
    return _export(EXTERNAL_KEYS, recs, "Employee List (External - PDPA safe)")
