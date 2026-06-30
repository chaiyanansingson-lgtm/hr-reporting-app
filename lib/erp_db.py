# lib/erp_db.py
# ============================================================================
# Mini ERP — Stationery (requirement 3, 12 Jun)
#   1. CATALOG: the OfficeMate products extracted from the 2026 catalogues
#      (same schema as the extraction project: product_code, item_no_model,
#      category_type, description, brand, capacity, color, min_order_qty,
#      price_thb, currency, image_file, source_catalogue, source_page,
#      catalogue_note). Import the master CSV/XLSX as the extraction
#      sessions complete. Retail item photos need no download — OfficeMate's
#      CDN URL is deterministic:
#          https://pim-cdn0.ofm.co.th/products/large/<code-without-dash>.jpg
#   2. PURCHASE ORDERS to OFFICEMATE: staff build a cart -> submit ->
#      approval (erp.approve) -> Admin places the real OFFICEMATE order and
#      records the order no. -> goods received.
#      Status: draft -> submitted -> approved -> ordered -> received
#              (or rejected / cancelled)
#   3. REIMBURSEMENTS: staff bought stationery with their own money ->
#      claim with amount + receipt photo -> approval -> Finance marks paid.
#      Status: submitted -> approved -> paid (or rejected)
# All actions audited through employee_audit_log (same trail as HR data).
# ============================================================================
import io
import datetime as dt

from lib.db import get_conn, IS_POSTGRES, PH
from lib import employee_db as edb

SERIAL = "SERIAL PRIMARY KEY" if IS_POSTGRES else \
         "INTEGER PRIMARY KEY AUTOINCREMENT"
BLOB = "BYTEA" if IS_POSTGRES else "BLOB"

PRODUCT_COLS = ["product_code", "item_no_model", "category_type",
                "description", "brand", "capacity", "color",
                "min_order_qty", "price_thb", "currency", "image_file",
                "source_catalogue", "source_page", "catalogue_note",
                "catalogue_group"]


def _ts():
    return dt.datetime.now().isoformat(timespec="seconds")


def ofm_image_url(product_code):
    code = str(product_code or "").replace("-", "").strip()
    return (f"https://pim-cdn0.ofm.co.th/products/large/{code}.jpg"
            if code else "")


def migrate():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""CREATE TABLE IF NOT EXISTS erp_products (
        id {SERIAL},
        product_code TEXT UNIQUE NOT NULL,
        item_no_model TEXT, category_type TEXT, description TEXT,
        brand TEXT, capacity TEXT, color TEXT,
        min_order_qty TEXT, price_thb REAL, currency TEXT DEFAULT 'THB',
        image_file TEXT, source_catalogue TEXT, source_page TEXT,
        catalogue_note TEXT,
        active INTEGER NOT NULL DEFAULT 1, updated_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS erp_purchase_orders (
        id {SERIAL},
        po_no TEXT UNIQUE,
        requester_emp_no TEXT, requester_name TEXT, department TEXT,
        purpose TEXT,
        status TEXT NOT NULL DEFAULT 'submitted',
        total_thb REAL,
        created_by TEXT, created_at TEXT,
        approved_by TEXT, approved_at TEXT, approve_note TEXT,
        ofm_order_no TEXT, ordered_by TEXT, ordered_at TEXT,
        received_by TEXT, received_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS erp_po_lines (
        id {SERIAL},
        po_id INTEGER NOT NULL,
        product_code TEXT, description TEXT,
        qty REAL NOT NULL, unit_price_thb REAL, line_total_thb REAL)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS erp_suppliers (
        id {SERIAL},
        supplier_key TEXT UNIQUE NOT NULL,        -- e.g. officemate, makro
        name TEXT NOT NULL, email TEXT, phone TEXT,
        lead_time_days INTEGER DEFAULT 7, payment_terms TEXT,
        active INTEGER NOT NULL DEFAULT 1)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS po_approval_lines (
        emp_no TEXT PRIMARY KEY,
        l1_emp_no TEXT, l2_emp_no TEXT, l3_emp_no TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS erp_budgets (
        id {SERIAL}, year INTEGER NOT NULL, department TEXT NOT NULL,
        budget_thb REAL NOT NULL DEFAULT 0)""")
    # Persist the CREATEs before the ALTER loop — on Postgres a failing ALTER's
    # rollback would otherwise wipe the uncommitted tables above (the lms_db
    # UndefinedTable failure mode).
    conn.commit()
    # additive columns (idempotent)
    for ddl in (
        "ALTER TABLE erp_products ADD COLUMN supplier_key TEXT",
        "ALTER TABLE erp_products ADD COLUMN catalogue_group TEXT",
        "ALTER TABLE erp_purchase_orders ADD COLUMN supplier_key TEXT",
        "ALTER TABLE erp_purchase_orders ADD COLUMN approver TEXT",
        "ALTER TABLE erp_purchase_orders ADD COLUMN sent_to_supplier_at TEXT",
        f"ALTER TABLE erp_purchase_orders ADD COLUMN approve_evidence {BLOB}",
        "ALTER TABLE erp_purchase_orders ADD COLUMN approve_evidence_mime TEXT",
        f"ALTER TABLE erp_purchase_orders ADD COLUMN quotation {BLOB}",
        "ALTER TABLE erp_purchase_orders ADD COLUMN quotation_mime TEXT",
        "ALTER TABLE erp_po_lines ADD COLUMN qty_received REAL DEFAULT 0",
        "ALTER TABLE erp_po_lines ADD COLUMN line_status TEXT DEFAULT 'open'",
        "ALTER TABLE erp_po_lines ADD COLUMN cancel_reason TEXT",
    ):
        try:
            cur.execute(ddl)
            conn.commit()
        except Exception:
            conn.rollback() if IS_POSTGRES else None
    # default supplier
    if IS_POSTGRES:
        cur.execute("""INSERT INTO erp_suppliers (supplier_key, name, email)
                       VALUES ('officemate','OfficeMate (OFM)',
                       'cs@officemate.co.th')
                       ON CONFLICT (supplier_key) DO NOTHING""")
    else:
        cur.execute("""INSERT OR IGNORE INTO erp_suppliers
                       (supplier_key, name, email) VALUES
                       ('officemate','OfficeMate (OFM)',
                       'cs@officemate.co.th')""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS erp_reimbursements (
        id {SERIAL},
        claim_no TEXT UNIQUE,
        requester_emp_no TEXT, requester_name TEXT, department TEXT,
        expense_date TEXT, vendor TEXT, items_desc TEXT,
        amount_thb REAL NOT NULL,
        receipt {BLOB}, receipt_mime TEXT,
        status TEXT NOT NULL DEFAULT 'submitted',
        created_by TEXT, created_at TEXT,
        approved_by TEXT, approved_at TEXT, approve_note TEXT,
        paid_by TEXT, paid_at TEXT, pay_ref TEXT)""")
    conn.commit()


# ---------------------------------------------------------------- catalog
def import_catalog(file_bytes, filename, actor):
    """Import the extraction master (CSV or XLSX with the agreed columns).
    Upserts on product_code. Returns summary."""
    import csv
    rows = []
    if filename.lower().endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        rows = list(csv.DictReader(io.StringIO(text)))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(hdr, r)))
    conn = get_conn(); cur = conn.cursor()
    created = updated = 0
    for r in rows:
        code = str(r.get("product_code") or "").strip()
        if not code:
            continue
        vals = {c: r.get(c) for c in PRODUCT_COLS}
        try:
            vals["price_thb"] = float(str(vals.get("price_thb") or "")
                                      .replace(",", "")) or None
        except ValueError:
            vals["price_thb"] = None
        cur.execute(f"SELECT id FROM erp_products WHERE product_code={PH}",
                    (code,))
        ex = cur.fetchone()
        if ex:
            sets = ", ".join(f"{c}={PH}" for c in PRODUCT_COLS[1:])
            cur.execute(f"UPDATE erp_products SET {sets}, updated_at={PH} "
                        f"WHERE product_code={PH}",
                        [vals[c] for c in PRODUCT_COLS[1:]] + [_ts(), code])
            updated += 1
        else:
            cols = ",".join(PRODUCT_COLS)
            cur.execute(f"INSERT INTO erp_products ({cols}, updated_at) "
                        f"VALUES ({','.join([PH]*len(PRODUCT_COLS))},{PH})",
                        [vals[c] for c in PRODUCT_COLS] + [_ts()])
            created += 1
    conn.commit()
    edb._audit(conn, actor, "erp_catalog_import",
               detail={"file": filename, "created": created,
                       "updated": updated})
    conn.commit()
    return {"created": created, "updated": updated}


def search_products(q="", category="", brand="", group="", limit=60):
    conn = get_conn(); cur = conn.cursor()
    sql = "SELECT * FROM erp_products WHERE active=1"
    args = []
    if q:
        like = f"%{q}%"
        sql += (f" AND (product_code LIKE {PH} OR description LIKE {PH} "
                f"OR brand LIKE {PH} OR category_type LIKE {PH})")
        args += [like, like, like, like]
    if category:
        sql += f" AND category_type={PH}"; args.append(category)
    if brand:
        sql += f" AND brand={PH}"; args.append(brand)
    if group:
        sql += f" AND catalogue_group={PH}"; args.append(group)
    sql += f" ORDER BY category_type, description LIMIT {int(limit)}"
    cur.execute(sql, args)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def categories():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT DISTINCT category_type FROM erp_products
                   WHERE active=1 AND category_type IS NOT NULL
                   ORDER BY category_type""")
    return [r[0] for r in cur.fetchall()]


def brands():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT DISTINCT brand FROM erp_products
                   WHERE active=1 AND brand IS NOT NULL AND brand<>''
                   ORDER BY brand""")
    return [r[0] for r in cur.fetchall()]


def catalogue_groups():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT DISTINCT catalogue_group FROM erp_products
                   WHERE active=1 AND catalogue_group IS NOT NULL
                   AND catalogue_group<>'' ORDER BY catalogue_group""")
    return [r[0] for r in cur.fetchall()]


# ---------------------------------------------------------------- PO
def _next_no(prefix, table, col):
    conn = get_conn(); cur = conn.cursor()
    ym = dt.date.today().strftime("%y%m")
    like = f"{prefix}-{ym}-%"
    cur.execute(f"SELECT COUNT(*) FROM {table} WHERE {col} LIKE {PH}",
                (like,))
    n = cur.fetchone()[0] + 1
    return f"{prefix}-{ym}-{n:03d}"


def create_po(requester, lines, purpose, actor, supplier_key=None):
    """lines: [{product_code, description, qty, unit_price_thb}]"""
    conn = get_conn(); cur = conn.cursor()
    po_no = _next_no("PO", "erp_purchase_orders", "po_no")
    total = sum((l.get("unit_price_thb") or 0) * l["qty"] for l in lines)
    cur.execute(
        f"""INSERT INTO erp_purchase_orders (po_no, requester_emp_no,
            requester_name, department, purpose, status, total_thb,
            created_by, created_at, supplier_key)
            VALUES ({PH},{PH},{PH},{PH},{PH},'submitted',{PH},{PH},{PH},
            {PH})""",
        (po_no, requester.get("emp_no"), requester.get("emp_name_en"),
         requester.get("dept_location"), purpose, total, actor, _ts(),
         supplier_key))
    if IS_POSTGRES:
        cur.execute("SELECT id FROM erp_purchase_orders WHERE po_no=%s",
                    (po_no,))
        po_id = cur.fetchone()[0]
    else:
        po_id = cur.lastrowid
    for l in lines:
        cur.execute(
            f"""INSERT INTO erp_po_lines (po_id, product_code, description,
                qty, unit_price_thb, line_total_thb)
                VALUES ({PH},{PH},{PH},{PH},{PH},{PH})""",
            (po_id, l.get("product_code"), l.get("description"), l["qty"],
             l.get("unit_price_thb"),
             (l.get("unit_price_thb") or 0) * l["qty"]))
    conn.commit()
    edb._audit(conn, actor, "erp_po_submit",
               detail={"po_no": po_no, "lines": len(lines),
                       "total_thb": total})
    conn.commit()
    return po_no


def list_pos(status=None, requester_emp_no=None):
    conn = get_conn(); cur = conn.cursor()
    sql = "SELECT * FROM erp_purchase_orders WHERE 1=1"
    args = []
    if status:
        sql += f" AND status={PH}"; args.append(status)
    if requester_emp_no:
        sql += f" AND requester_emp_no={PH}"
        args.append(str(requester_emp_no))
    sql += " ORDER BY id DESC"
    cur.execute(sql, args)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def po_lines(po_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM erp_po_lines WHERE po_id={PH}", (po_id,))
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def po_action(po_id, action, actor, note="", ofm_order_no="", pay_ref=""):
    """action: approve | reject | order | receive | cancel"""
    conn = get_conn(); cur = conn.cursor()
    now = _ts()
    if action == "approve":
        cur.execute(f"""UPDATE erp_purchase_orders SET status='approved',
                        approved_by={PH}, approved_at={PH}, approve_note={PH}
                        WHERE id={PH}""", (actor, now, note, po_id))
    elif action == "reject":
        cur.execute(f"""UPDATE erp_purchase_orders SET status='rejected',
                        approved_by={PH}, approved_at={PH}, approve_note={PH}
                        WHERE id={PH}""", (actor, now, note, po_id))
    elif action == "order":
        cur.execute(f"""UPDATE erp_purchase_orders SET status='ordered',
                        ofm_order_no={PH}, ordered_by={PH}, ordered_at={PH}
                        WHERE id={PH}""", (ofm_order_no, actor, now, po_id))
    elif action == "receive":
        cur.execute(f"""UPDATE erp_purchase_orders SET status='received',
                        received_by={PH}, received_at={PH} WHERE id={PH}""",
                    (actor, now, po_id))
    elif action == "cancel":
        cur.execute(f"""UPDATE erp_purchase_orders SET status='cancelled'
                        WHERE id={PH}""", (po_id,))
    conn.commit()
    edb._audit(conn, actor, f"erp_po_{action}",
               detail={"po_id": po_id, "note": note,
                       "ofm_order_no": ofm_order_no})
    conn.commit()


# ---------------------------------------------------------------- reimburse
def create_reimbursement(requester, expense_date, vendor, items_desc,
                         amount_thb, receipt_bytes, receipt_mime, actor):
    conn = get_conn(); cur = conn.cursor()
    claim_no = _next_no("RB", "erp_reimbursements", "claim_no")
    cur.execute(
        f"""INSERT INTO erp_reimbursements (claim_no, requester_emp_no,
            requester_name, department, expense_date, vendor, items_desc,
            amount_thb, receipt, receipt_mime, status, created_by,
            created_at) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},
            {PH},{PH},'submitted',{PH},{PH})""",
        (claim_no, requester.get("emp_no"), requester.get("emp_name_en"),
         requester.get("dept_location"), str(expense_date), vendor,
         items_desc, amount_thb, receipt_bytes, receipt_mime, actor, _ts()))
    conn.commit()
    edb._audit(conn, actor, "erp_reimburse_submit",
               detail={"claim_no": claim_no, "amount_thb": amount_thb})
    conn.commit()
    return claim_no


def list_reimbursements(status=None, requester_emp_no=None):
    conn = get_conn(); cur = conn.cursor()
    sql = """SELECT id, claim_no, requester_emp_no, requester_name,
             department, expense_date, vendor, items_desc, amount_thb,
             status, created_at, approved_by, approve_note, paid_by,
             paid_at, pay_ref, receipt_mime FROM erp_reimbursements
             WHERE 1=1"""
    args = []
    if status:
        sql += f" AND status={PH}"; args.append(status)
    if requester_emp_no:
        sql += f" AND requester_emp_no={PH}"
        args.append(str(requester_emp_no))
    sql += " ORDER BY id DESC"
    cur.execute(sql, args)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def get_receipt(claim_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT receipt, receipt_mime FROM erp_reimbursements "
                f"WHERE id={PH}", (claim_id,))
    r = cur.fetchone()
    return (bytes(r[0]), r[1]) if r and r[0] else (None, None)


def reimburse_action(claim_id, action, actor, note="", pay_ref=""):
    """action: approve | reject | pay"""
    conn = get_conn(); cur = conn.cursor()
    now = _ts()
    if action == "approve":
        cur.execute(f"""UPDATE erp_reimbursements SET status='approved',
                        approved_by={PH}, approved_at={PH}, approve_note={PH}
                        WHERE id={PH}""", (actor, now, note, claim_id))
    elif action == "reject":
        cur.execute(f"""UPDATE erp_reimbursements SET status='rejected',
                        approved_by={PH}, approved_at={PH}, approve_note={PH}
                        WHERE id={PH}""", (actor, now, note, claim_id))
    elif action == "pay":
        cur.execute(f"""UPDATE erp_reimbursements SET status='paid',
                        paid_by={PH}, paid_at={PH}, pay_ref={PH}
                        WHERE id={PH}""", (actor, now, pay_ref, claim_id))
    conn.commit()
    edb._audit(conn, actor, f"erp_reimburse_{action}",
               detail={"claim_id": claim_id, "note": note,
                       "pay_ref": pay_ref})
    conn.commit()


# ---------------------------------------------------------------- suppliers
def list_suppliers(active_only=True):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM erp_suppliers" +
                (" WHERE active=1" if active_only else "") + " ORDER BY name")
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def upsert_supplier(supplier_key, name, email, phone="", lead_time_days=7,
                    payment_terms="", active=1):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT id FROM erp_suppliers WHERE supplier_key={PH}",
                (supplier_key,))
    if cur.fetchone():
        cur.execute(f"""UPDATE erp_suppliers SET name={PH}, email={PH},
                        phone={PH}, lead_time_days={PH}, payment_terms={PH},
                        active={PH} WHERE supplier_key={PH}""",
                    (name, email, phone, lead_time_days, payment_terms,
                     active, supplier_key))
    else:
        cur.execute(f"""INSERT INTO erp_suppliers (supplier_key, name, email,
                        phone, lead_time_days, payment_terms, active)
                        VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
                    (supplier_key, name, email, phone, lead_time_days,
                     payment_terms, active))
    conn.commit()


# ------------------------------------------------------- PO approval lines
def get_approval_line(emp_no):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM po_approval_lines WHERE emp_no={PH}",
                (str(emp_no),))
    r = cur.fetchone()
    if not r:
        return None
    cols = [d[0] for d in cur.description]
    return dict(zip(cols, r)) if IS_POSTGRES else dict(r)


def set_approval_line(emp_no, l1, l2, l3):
    conn = get_conn(); cur = conn.cursor()
    if IS_POSTGRES:
        cur.execute("""INSERT INTO po_approval_lines (emp_no, l1_emp_no,
                       l2_emp_no, l3_emp_no) VALUES (%s,%s,%s,%s)
                       ON CONFLICT (emp_no) DO UPDATE SET
                       l1_emp_no=EXCLUDED.l1_emp_no,
                       l2_emp_no=EXCLUDED.l2_emp_no,
                       l3_emp_no=EXCLUDED.l3_emp_no""",
                    (str(emp_no), l1 or None, l2 or None, l3 or None))
    else:
        cur.execute("INSERT OR REPLACE INTO po_approval_lines VALUES "
                    "(?,?,?,?)", (str(emp_no), l1 or None, l2 or None,
                                  l3 or None))
    conn.commit()


def import_approval_lines(file_bytes, actor):
    """Bulk upload: Excel/CSV with columns Emp No., L1, L2, L3 (emp nos)."""
    import io, csv
    rows = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        hdr = [str(c.value or "").strip().lower() for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(hdr, r)))
    except Exception:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        rows = [{k.lower().strip(): v for k, v in r.items()}
                for r in csv.DictReader(io.StringIO(text))]

    def pick(d, *names):
        for n in names:
            for k, v in d.items():
                if n in k:
                    return str(v).strip() if v not in (None, "") else ""
        return ""
    n = 0
    for r in rows:
        emp = pick(r, "emp")
        if not emp:
            continue
        set_approval_line(emp, pick(r, "l1", "level1", "level 1"),
                          pick(r, "l2", "level2", "level 2"),
                          pick(r, "l3", "level3", "level 3"))
        n += 1
    conn = get_conn()
    edb._audit(conn, actor, "erp_approval_lines_import", detail={"rows": n})
    conn.commit()
    return n


def resolve_po_chain(requester_rec):
    """Custom PO line if configured, else the Mgr-column chain.
    Returns [(level, emp_record), ...] for approval_db.open_approvals."""
    from lib import approval_db as adb
    line = get_approval_line(requester_rec.get("emp_no"))
    if line:
        chain = []
        lvl = 0
        for key in ("l1_emp_no", "l2_emp_no", "l3_emp_no"):
            no = (line.get(key) or "").strip()
            if not no:
                continue
            rec = edb.get_record(emp_no=no)
            if rec and str(rec.get("emp_no")) !=                     str(requester_rec.get("emp_no")):
                lvl += 1
                chain.append((lvl, rec))
        if chain:
            return chain
    return adb.resolve_chain(requester_rec)


# ------------------------------------------------------- pipeline v2
def po_set_status(po_id, status, actor, **kw):
    conn = get_conn(); cur = conn.cursor()
    sets = [f"status={PH}"]; args = [status]
    if status == "sent_to_supplier":
        sets.append(f"sent_to_supplier_at={PH}")
        args.append(dt.datetime.now().isoformat(timespec="seconds"))
    if kw.get("ofm_order_no") is not None:
        sets.append(f"ofm_order_no={PH}"); args.append(kw["ofm_order_no"])
    args.append(po_id)
    cur.execute(f"UPDATE erp_purchase_orders SET {', '.join(sets)} "
                f"WHERE id={PH}", args)
    conn.commit()
    edb._audit(conn, actor, f"erp_po_{status}", detail={"po_id": po_id})
    conn.commit()


def po_attach(po_id, field, data, mime, actor):
    """field: approve_evidence | quotation"""
    assert field in ("approve_evidence", "quotation")
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"UPDATE erp_purchase_orders SET {field}={PH}, "
                f"{field}_mime={PH} WHERE id={PH}", (data, mime, po_id))
    conn.commit()
    edb._audit(conn, actor, f"erp_po_attach_{field}", detail={"po_id": po_id})
    conn.commit()


def receive_line(line_id, qty, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT po_id, qty, qty_received FROM erp_po_lines "
                f"WHERE id={PH}", (line_id,))
    r = cur.fetchone()
    po_id, qty_ord, qty_rec = r[0], float(r[1]), float(r[2] or 0)
    new_rec = min(qty_rec + float(qty), qty_ord)
    status = "received" if new_rec >= qty_ord else "partial"
    cur.execute(f"""UPDATE erp_po_lines SET qty_received={PH},
                    line_status={PH} WHERE id={PH}""",
                (new_rec, status, line_id))
    conn.commit()
    _refresh_po_receive_status(po_id, actor)


def cancel_line(line_id, reason, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT po_id FROM erp_po_lines WHERE id={PH}", (line_id,))
    po_id = cur.fetchone()[0]
    cur.execute(f"""UPDATE erp_po_lines SET line_status='cancelled',
                    cancel_reason={PH} WHERE id={PH}""", (reason, line_id))
    conn.commit()
    edb._audit(conn, actor, "erp_line_cancel",
               detail={"line_id": line_id, "reason": reason})
    conn.commit()
    _refresh_po_receive_status(po_id, actor)


def _refresh_po_receive_status(po_id, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT line_status FROM erp_po_lines WHERE po_id={PH}",
                (po_id,))
    sts = [r[0] for r in cur.fetchall()]
    live = [x for x in sts if x != "cancelled"]
    if live and all(x == "received" for x in live):
        new = "received"
    elif any(x in ("received", "partial") for x in live):
        new = "partially_received"
    else:
        return
    cur.execute(f"UPDATE erp_purchase_orders SET status={PH} WHERE id={PH} "
                f"AND status IN ('ordered','partially_received')",
                (new, po_id))
    conn.commit()


# ------------------------------------------------------- budgets & reports
def set_budget(year, department, budget_thb):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"DELETE FROM erp_budgets WHERE year={PH} AND "
                f"department={PH}", (year, department))
    cur.execute(f"INSERT INTO erp_budgets (year, department, budget_thb) "
                f"VALUES ({PH},{PH},{PH})", (year, department, budget_thb))
    conn.commit()


def budget_vs_actual(year):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT department, SUM(total_thb) FROM
                    erp_purchase_orders
                    WHERE status IN ('approved','sent_to_supplier','ordered',
                    'partially_received','received')
                    AND created_at LIKE {PH} GROUP BY department""",
                (f"{year}%",))
    actual = {r[0] or "—": float(r[1] or 0) for r in cur.fetchall()}
    cur.execute(f"SELECT department, budget_thb FROM erp_budgets "
                f"WHERE year={PH}", (year,))
    budget = {r[0]: float(r[1] or 0) for r in cur.fetchall()}
    depts = sorted(set(actual) | set(budget))
    return [{"department": d, "budget_thb": budget.get(d, 0.0),
             "actual_thb": actual.get(d, 0.0),
             "remaining_thb": budget.get(d, 0.0) - actual.get(d, 0.0)}
            for d in depts]


def spend_summary():
    """dicts for charts: by month, by department, by supplier."""
    conn = get_conn(); cur = conn.cursor()
    out = {}
    cur.execute("""SELECT substr(created_at,1,7), SUM(total_thb)
                   FROM erp_purchase_orders WHERE status NOT IN
                   ('rejected','cancelled','draft','submitted',
                   'pending_l1','pending_l2','pending_l3')
                   GROUP BY substr(created_at,1,7) ORDER BY 1""")
    out["month"] = {r[0]: float(r[1] or 0) for r in cur.fetchall()}
    cur.execute("""SELECT department, SUM(total_thb)
                   FROM erp_purchase_orders WHERE status NOT IN
                   ('rejected','cancelled','draft','submitted',
                   'pending_l1','pending_l2','pending_l3')
                   GROUP BY department""")
    out["department"] = {r[0] or "—": float(r[1] or 0)
                         for r in cur.fetchall()}
    cur.execute("""SELECT supplier_key, SUM(total_thb)
                   FROM erp_purchase_orders WHERE status NOT IN
                   ('rejected','cancelled','draft','submitted',
                   'pending_l1','pending_l2','pending_l3')
                   GROUP BY supplier_key""")
    out["supplier"] = {r[0] or "—": float(r[1] or 0) for r in cur.fetchall()}
    return out


def ageing():
    """Open POs with days in current status (sent>7d / ordered>14d = red)."""
    now = dt.datetime.now()
    rows = []
    for po in list_pos():
        if po["status"] not in ("sent_to_supplier", "ordered",
                                "partially_received"):
            continue
        ref = po.get("sent_to_supplier_at") or po.get("ordered_at")             or po.get("created_at")
        try:
            days = (now - dt.datetime.fromisoformat(ref)).days
        except Exception:
            days = 0
        limit = 7 if po["status"] == "sent_to_supplier" else 14
        rows.append({"po_no": po["po_no"], "status": po["status"],
                     "supplier": po.get("supplier_key") or "—",
                     "days": days, "overdue": days > limit,
                     "total_thb": po.get("total_thb") or 0})
    return sorted(rows, key=lambda r: -r["days"])
