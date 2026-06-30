# lib/lms_db.py
# ============================================================================
# TRAINING / LMS (§7)
# Course = ordered LESSONS (YouTube video / slide pages / test), assignable
# to person, department, position, or role (incl. sealed visitors).
# Anti-skip video: seek blocked, speed locked 1x, pauses on tab blur; a
# COMPLETION CODE appears only at >=95% genuine watch-time and is verified
# server-side (HMAC). Slides require paging to the last page. Tests unlock
# only when every lesson before them is done.
# Test engine: mcq1 (auto) · mcqN (auto, partial credit) · short (auto by
# accepted answers, else manual) · long (manual grading queue — for
# engineering working / formulas).
# Completion -> bilingual certificate AMS-TRN-yymm-NNN.
# ============================================================================
import datetime as dt
import hashlib
import hmac
import json

from lib.db import get_conn, IS_POSTGRES, PH
from lib import employee_db as edb

SERIAL = "SERIAL PRIMARY KEY" if IS_POSTGRES else \
         "INTEGER PRIMARY KEY AUTOINCREMENT"
_SECRET = b"ams-lms-2026"   # for watch-codes only (not security-critical)


def _ts():
    return dt.datetime.now().isoformat(timespec="seconds")


def migrate():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_courses (
        id {SERIAL}, code TEXT UNIQUE, title_th TEXT, title_en TEXT,
        purpose TEXT, pass_pct REAL DEFAULT 70,
        active INTEGER NOT NULL DEFAULT 1,
        created_by TEXT, created_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_lessons (
        id {SERIAL}, course_id INTEGER NOT NULL, seq INTEGER NOT NULL,
        kind TEXT NOT NULL,              -- video | slides | test | interactive
        title TEXT,
        youtube_id TEXT,                 -- video
        duration_min REAL DEFAULT 0,     -- video/interactive: server time-floor
        pages TEXT,                      -- slides: JSON [text,...]
        asset_path TEXT,                 -- interactive: repo-relative .html path
        test_id INTEGER)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_tests (
        id {SERIAL}, course_id INTEGER NOT NULL, title TEXT,
        pass_pct REAL DEFAULT 70, attempts_allowed INTEGER DEFAULT 3,
        show_answers INTEGER DEFAULT 0)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_questions (
        id {SERIAL}, test_id INTEGER NOT NULL, seq INTEGER,
        kind TEXT NOT NULL,              -- mcq1 | mcqN | short | long
        text TEXT, options TEXT,         -- JSON list
        answer_key TEXT,                 -- JSON (idx | [idx] | [accepted])
        points REAL DEFAULT 1)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_assignments (
        id {SERIAL}, course_id INTEGER NOT NULL,
        target_kind TEXT NOT NULL,       -- person|department|position|role
        target_value TEXT NOT NULL,
        start_date TEXT, due_date TEXT,
        remind_days TEXT DEFAULT '[14,7,1]',
        created_by TEXT, created_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_enrollments (
        id {SERIAL}, course_id INTEGER NOT NULL,
        emp_key TEXT NOT NULL,           -- emp_no or 'user:<username>'
        display_name TEXT, department TEXT,
        assignment_id INTEGER, due_date TEXT,
        status TEXT NOT NULL DEFAULT 'assigned',
        completed_at TEXT, cert_no TEXT)""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS ux_enroll
                   ON lms_enrollments (course_id, emp_key)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_progress (
        id {SERIAL}, enrollment_id INTEGER NOT NULL,
        lesson_id INTEGER NOT NULL,
        page_reached INTEGER DEFAULT 0, watch_seconds INTEGER DEFAULT 0,
        first_opened_at TEXT,            -- server clock: anti-fast-forward
        flags INTEGER DEFAULT 0,         -- failed/too-early attempts
        done INTEGER NOT NULL DEFAULT 0, done_at TEXT)""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS ux_prog
                   ON lms_progress (enrollment_id, lesson_id)""")
    for ddl in ("ALTER TABLE lms_lessons ADD COLUMN duration_min REAL DEFAULT 0",
                "ALTER TABLE lms_lessons ADD COLUMN asset_path TEXT",
                "ALTER TABLE lms_progress ADD COLUMN first_opened_at TEXT",
                "ALTER TABLE lms_progress ADD COLUMN flags INTEGER DEFAULT 0",
                # DSD (กรมพัฒนาฝีมือแรงงาน) course metadata
                "ALTER TABLE lms_courses ADD COLUMN dsd_code TEXT",
                "ALTER TABLE lms_courses ADD COLUMN level TEXT",
                "ALTER TABLE lms_courses ADD COLUMN course_group TEXT",
                "ALTER TABLE lms_courses ADD COLUMN course_type TEXT",
                "ALTER TABLE lms_courses ADD COLUMN occupation_branch TEXT",
                "ALTER TABLE lms_courses ADD COLUMN instructor TEXT",
                "ALTER TABLE lms_courses ADD COLUMN objectives TEXT",
                "ALTER TABLE lms_tests ADD COLUMN role TEXT DEFAULT 'quiz'"):
        try:
            cur.execute(ddl); conn.commit()
        except Exception:
            conn.rollback() if IS_POSTGRES else None
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_attempts (
        id {SERIAL}, enrollment_id INTEGER NOT NULL,
        test_id INTEGER NOT NULL, started_at TEXT, submitted_at TEXT,
        score REAL DEFAULT 0, max_score REAL DEFAULT 0,
        passed INTEGER DEFAULT 0,
        grading TEXT NOT NULL DEFAULT 'auto')""")  # auto|pending|done
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_view_log (
        id {SERIAL}, enrollment_id INTEGER NOT NULL,
        lesson_id INTEGER NOT NULL, viewed_at TEXT)""")  # DSD 2-yr log
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_surveys (
        id {SERIAL}, enrollment_id INTEGER UNIQUE NOT NULL,
        q1 INTEGER, q2 INTEGER, q3 INTEGER, q4 INTEGER, q5 INTEGER,
        comment TEXT, created_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_skill_evals (
        id {SERIAL}, enrollment_id INTEGER UNIQUE NOT NULL,
        c1 REAL, c2 REAL, c3 REAL, c4 REAL, c5 REAL,   -- Generic9 criteria
        income_range TEXT, productivity INTEGER DEFAULT 0,
        evaluator TEXT, eval_date TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS lms_answers (
        id {SERIAL}, attempt_id INTEGER NOT NULL,
        question_id INTEGER NOT NULL, answer TEXT,
        score REAL, graded_by TEXT, comment TEXT)""")
    conn.commit()


def _rows(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def _one(cur):
    r = cur.fetchone()
    if not r:
        return None
    cols = [d[0] for d in cur.description]
    return dict(zip(cols, r)) if IS_POSTGRES else dict(r)


# ---------------------------------------------------------------- courses
def create_course(code, title_th, title_en, purpose, pass_pct, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""INSERT INTO lms_courses (code, title_th, title_en,
                    purpose, pass_pct, created_by, created_at)
                    VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
                (code, title_th, title_en, purpose, pass_pct, actor, _ts()))
    conn.commit()
    cur.execute(f"SELECT id FROM lms_courses WHERE code={PH}", (code,))
    cid = cur.fetchone()[0]
    edb._audit(conn, actor, "lms_course_create", detail={"code": code})
    conn.commit()
    return cid


def list_courses(active_only=True):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM lms_courses" +
                (" WHERE active=1" if active_only else "") + " ORDER BY id")
    return _rows(cur)


def get_course(cid):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_courses WHERE id={PH}", (cid,))
    return _one(cur)


def add_lesson(course_id, kind, title, youtube_id=None, pages=None,
               test_id=None, duration_min=0, asset_path=None):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT COALESCE(MAX(seq),0)+1 FROM lms_lessons "
                f"WHERE course_id={PH}", (course_id,))
    seq = cur.fetchone()[0]
    cur.execute(f"""INSERT INTO lms_lessons (course_id, seq, kind, title,
                    youtube_id, duration_min, pages, asset_path, test_id)
                    VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
                (course_id, seq, kind, title, youtube_id, duration_min,
                 json.dumps(pages or [], ensure_ascii=False),
                 asset_path, test_id))
    conn.commit()
    if IS_POSTGRES:
        cur.execute("SELECT MAX(id) FROM lms_lessons"); return cur.fetchone()[0]
    return cur.lastrowid


def lessons(course_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_lessons WHERE course_id={PH} "
                f"ORDER BY seq", (course_id,))
    return _rows(cur)


def create_test(course_id, title, pass_pct, attempts, show_answers,
                role="quiz"):
    """role: pre | post | quiz. DSD: post-test pass >= 60%; the pre-test
    records a baseline and never blocks completion."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""INSERT INTO lms_tests (course_id, title, pass_pct,
                    attempts_allowed, show_answers, role)
                    VALUES ({PH},{PH},{PH},{PH},{PH},{PH})""",
                (course_id, title, pass_pct, attempts,
                 1 if show_answers else 0, role))
    conn.commit()
    if IS_POSTGRES:
        cur.execute("SELECT MAX(id) FROM lms_tests"); return cur.fetchone()[0]
    return cur.lastrowid


def add_question(test_id, kind, text, options, answer_key, points=1):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT COALESCE(MAX(seq),0)+1 FROM lms_questions "
                f"WHERE test_id={PH}", (test_id,))
    seq = cur.fetchone()[0]
    cur.execute(f"""INSERT INTO lms_questions (test_id, seq, kind, text,
                    options, answer_key, points)
                    VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
                (test_id, seq, kind, text,
                 json.dumps(options or [], ensure_ascii=False),
                 json.dumps(answer_key, ensure_ascii=False), points))
    conn.commit()


def questions(test_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_questions WHERE test_id={PH} "
                f"ORDER BY seq", (test_id,))
    return _rows(cur)


def get_test(tid):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_tests WHERE id={PH}", (tid,))
    return _one(cur)


def update_course_meta(cid, **kw):
    cols = ["dsd_code", "level", "course_group", "course_type",
            "occupation_branch", "instructor", "objectives", "pass_pct",
            "title_th", "title_en", "purpose"]
    sets, args = [], []
    for k, v in kw.items():
        if k in cols:
            sets.append(f"{k}={PH}"); args.append(v)
    if not sets:
        return
    args.append(cid)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"UPDATE lms_courses SET {', '.join(sets)} WHERE id={PH}",
                args)
    conn.commit()


def course_video_hours(cid):
    return round(sum((l.get("duration_min") or 0) for l in lessons(cid)
                     if l["kind"] == "video") / 60, 2)


def log_view(enrollment_id, lesson_id):
    """DSD evidence log: a row per lesson open (kept indefinitely > 2 yrs)."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""INSERT INTO lms_view_log (enrollment_id, lesson_id,
                    viewed_at) VALUES ({PH},{PH},{PH})""",
                (enrollment_id, lesson_id, _ts()))
    conn.commit()


def save_survey(enrollment_id, q, comment):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"DELETE FROM lms_surveys WHERE enrollment_id={PH}",
                (enrollment_id,))
    cur.execute(f"""INSERT INTO lms_surveys (enrollment_id, q1, q2, q3, q4,
                    q5, comment, created_at)
                    VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
                (enrollment_id, q[0], q[1], q[2], q[3], q[4], comment,
                 _ts()))
    conn.commit()


def has_survey(enrollment_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT 1 FROM lms_surveys WHERE enrollment_id={PH}",
                (enrollment_id,))
    return bool(cur.fetchone())


def pre_post(enrollment_id):
    """Best pre-test % and best post-test % for the enrollment ->
    (pre_pct, post_pct, improvement) — the DSD before/after comparison."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT t.role, MAX(CASE WHEN a.max_score>0 THEN
                    100.0*a.score/a.max_score END) AS pct
                    FROM lms_attempts a JOIN lms_tests t ON t.id=a.test_id
                    WHERE a.enrollment_id={PH} AND a.grading='done'
                    GROUP BY t.role""", (enrollment_id,))
    d = {r[0]: (round(float(r[1]), 1) if r[1] is not None else None)
         for r in cur.fetchall()}
    pre, post = d.get("pre"), d.get("post") or d.get("quiz")
    imp = round(post - pre, 1) if (pre is not None and post is not None)         else None
    return pre, post, imp


# ---------------------------------------------------------------- assign
def assign(course_id, target_kind, target_value, due_date, remind_days,
           actor):
    """Expand target -> enrollments. Returns count enrolled.
    person = emp_no | department = dept_location | position = title |
    role = role_key (e.g. visitor -> all users of that role)."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""INSERT INTO lms_assignments (course_id, target_kind,
                    target_value, due_date, remind_days, created_by,
                    created_at) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
                (course_id, target_kind, target_value, str(due_date),
                 json.dumps(remind_days), actor, _ts()))
    conn.commit()
    if IS_POSTGRES:
        cur.execute("SELECT MAX(id) FROM lms_assignments")
        aid = cur.fetchone()[0]
    else:
        aid = cur.lastrowid

    targets = []   # (emp_key, name, dept)
    if target_kind == "person":
        r = edb.get_record(emp_no=str(target_value))
        if r:
            targets = [(str(r["emp_no"]), r.get("emp_name_en"),
                        r.get("dept_location"))]
    elif target_kind in ("department", "position"):
        fld = "dept_location" if target_kind == "department" else "title"
        targets = [(str(r["emp_no"]), r.get("emp_name_en"),
                    r.get("dept_location"))
                   for r in edb.list_records("active")
                   if (r.get(fld) or "") == target_value]
    elif target_kind == "role":
        from lib.auth import list_users
        for u in list_users():
            if u.get("role_key") != target_value or not u.get("active"):
                continue
            if u.get("emp_no"):
                r = edb.get_record(emp_no=str(u["emp_no"]))
                targets.append((str(u["emp_no"]),
                                (r or {}).get("emp_name_en") or u["username"],
                                (r or {}).get("dept_location")))
            else:
                targets.append((f"user:{u['username']}", u["username"],
                                "(external)"))
    n = 0
    for key, name, dept in targets:
        try:
            cur.execute(f"""INSERT INTO lms_enrollments (course_id, emp_key,
                            display_name, department, assignment_id,
                            due_date) VALUES ({PH},{PH},{PH},{PH},{PH},{PH})""",
                        (course_id, key, name, dept, aid, str(due_date)))
            n += 1
        except Exception:
            conn.rollback() if IS_POSTGRES else None   # already enrolled
    conn.commit()
    edb._audit(conn, actor, "lms_assign",
               detail={"course": course_id, "kind": target_kind,
                       "value": target_value, "enrolled": n})
    conn.commit()
    return n


def my_enrollments(emp_key):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT e.*, c.code, c.title_th, c.title_en, c.purpose,
                    c.pass_pct FROM lms_enrollments e
                    JOIN lms_courses c ON c.id = e.course_id
                    WHERE e.emp_key={PH} AND c.active=1
                    ORDER BY e.due_date""", (str(emp_key),))
    return _rows(cur)


def course_enrollments(course_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_enrollments WHERE course_id={PH}",
                (course_id,))
    return _rows(cur)


# ---------------------------------------------------------------- progress
def watch_code(enrollment_id, lesson_id):
    """6-char completion code revealed by the player at >=95% watch."""
    raw = f"{enrollment_id}:{lesson_id}".encode()
    return hmac.new(_SECRET, raw, hashlib.sha256).hexdigest()[:6].upper()


def _prog_row(enrollment_id, lesson_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT * FROM lms_progress WHERE enrollment_id={PH}
                    AND lesson_id={PH}""", (enrollment_id, lesson_id))
    return _one(cur)


def set_progress(enrollment_id, lesson_id, page_reached=None,
                 watch_seconds=None, done=None):
    conn = get_conn(); cur = conn.cursor()
    if not _prog_row(enrollment_id, lesson_id):
        cur.execute(f"""INSERT INTO lms_progress (enrollment_id, lesson_id)
                        VALUES ({PH},{PH})""", (enrollment_id, lesson_id))
    sets, args = [], []
    if page_reached is not None:
        sets.append(f"page_reached=MAX(page_reached,{PH})"
                    if not IS_POSTGRES else
                    f"page_reached=GREATEST(page_reached,{PH})")
        args.append(int(page_reached))
    if watch_seconds is not None:
        sets.append(f"watch_seconds={PH}"); args.append(int(watch_seconds))
    if done:
        sets.append(f"done=1, done_at={PH}"); args.append(_ts())
    if sets:
        args += [enrollment_id, lesson_id]
        cur.execute(f"""UPDATE lms_progress SET {', '.join(sets)}
                        WHERE enrollment_id={PH} AND lesson_id={PH}""", args)
    conn.commit()
    _mark_in_progress(enrollment_id)


def mark_opened(enrollment_id, lesson_id):
    """Server clock starts the moment the player first renders."""
    conn = get_conn(); cur = conn.cursor()
    if not _prog_row(enrollment_id, lesson_id):
        cur.execute(f"""INSERT INTO lms_progress (enrollment_id, lesson_id)
                        VALUES ({PH},{PH})""", (enrollment_id, lesson_id))
    cur.execute(f"""UPDATE lms_progress SET first_opened_at={PH}
                    WHERE enrollment_id={PH} AND lesson_id={PH}
                    AND first_opened_at IS NULL""",
                (_ts(), enrollment_id, lesson_id))
    conn.commit()


def _flag(enrollment_id, lesson_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""UPDATE lms_progress SET flags=COALESCE(flags,0)+1
                    WHERE enrollment_id={PH} AND lesson_id={PH}""",
                (enrollment_id, lesson_id))
    conn.commit()


def verify_watch_code(enrollment_id, lesson_id, code):
    """(ok, msg). SERVER-SIDE time floor: even with the correct code, the
    lesson cannot complete before >=90% of the video's real length has
    elapsed since first open (server clock) — fast-forward/skip/source-
    reading cannot beat wall-clock time. Failures increment the cheat
    flag visible to admins."""
    if (code or "").strip().upper() != watch_code(enrollment_id, lesson_id):
        _flag(enrollment_id, lesson_id)
        return False, "โค้ดไม่ถูกต้อง — ต้องดูครบ 95% ก่อน / wrong code"
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT duration_min FROM lms_lessons WHERE id={PH}",
                (lesson_id,))
    row = cur.fetchone()
    dur_min = float(row[0] or 0) if row else 0
    if dur_min > 0:
        p = _prog_row(enrollment_id, lesson_id)
        opened = (p or {}).get("first_opened_at")
        if not opened:
            mark_opened(enrollment_id, lesson_id)
            _flag(enrollment_id, lesson_id)
            return False, ("ระบบเพิ่งเริ่มจับเวลา — ต้องเปิดดูจนครบความยาว"
                           "วิดีโอ / timer just started; watch in full")
        elapsed = (dt.datetime.now()
                   - dt.datetime.fromisoformat(opened)).total_seconds()
        need = dur_min * 60 * 0.95
        if elapsed < need:
            _flag(enrollment_id, lesson_id)
            left = int((need - elapsed) / 60) + 1
            return False, (f"เร็วเกินความยาววิดีโอจริง — ระบบบังคับให้ดูต่อ"
                           f"อีกอย่างน้อย ~{left} นาที (ตัวจับเวลาฝั่ง"
                           f"เซิร์ฟเวอร์) / faster than the video's real "
                           f"length; ~{left} min of genuine watch time "
                           f"still required")
    set_progress(enrollment_id, lesson_id, done=True)
    return True, "บันทึกแล้ว ✅"


def _mark_in_progress(enrollment_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""UPDATE lms_enrollments SET status='in_progress'
                    WHERE id={PH} AND status='assigned'""", (enrollment_id,))
    conn.commit()


def lesson_state(enrollment_id, course_id):
    """[(lesson, done, unlocked)] — lesson N unlocks when 1..N-1 done."""
    ls = lessons(course_id)
    out, prev_done = [], True
    for l in ls:
        p = _prog_row(enrollment_id, l["id"])
        done = bool(p and p["done"])
        out.append((l, done, prev_done))
        prev_done = prev_done and done
    return out


def enrollment_pct(enrollment_id, course_id):
    st_ = lesson_state(enrollment_id, course_id)
    if not st_:
        return 0
    return round(100 * sum(1 for _, d, _ in st_ if d) / len(st_))


# ---------------------------------------------------------------- tests
def start_attempt(enrollment_id, test_id):
    t = get_test(test_id)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT COUNT(*) FROM lms_attempts WHERE
                    enrollment_id={PH} AND test_id={PH}""",
                (enrollment_id, test_id))
    used = cur.fetchone()[0]
    if used >= (t["attempts_allowed"] or 99):
        return None, "ใช้สิทธิ์สอบครบแล้ว / attempts exhausted"
    cur.execute(f"""INSERT INTO lms_attempts (enrollment_id, test_id,
                    started_at) VALUES ({PH},{PH},{PH})""",
                (enrollment_id, test_id, _ts()))
    conn.commit()
    if IS_POSTGRES:
        cur.execute("SELECT MAX(id) FROM lms_attempts")
        return cur.fetchone()[0], None
    return cur.lastrowid, None


def submit_attempt(attempt_id, answers_by_qid):
    """Auto-grades mcq1/mcqN/short; 'long' -> manual queue.
    Returns the attempt dict after grading."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_attempts WHERE id={PH}", (attempt_id,))
    at = _one(cur)
    qs = questions(at["test_id"])
    score = 0.0; max_score = 0.0; has_long = False
    for q in qs:
        max_score += q["points"] or 1
        ans = answers_by_qid.get(q["id"])
        key = json.loads(q["answer_key"] or "null")
        pts = None
        if q["kind"] == "mcq1":
            pts = (q["points"] or 1) if ans == key else 0.0
        elif q["kind"] == "mcqN":
            ks, asel = set(key or []), set(ans or [])
            if not ks:
                pts = 0.0
            else:   # partial credit: correct picks minus wrong picks
                good = len(ks & asel); bad = len(asel - ks)
                pts = max(0.0, (good - bad) / len(ks)) * (q["points"] or 1)
        elif q["kind"] == "short":
            accepted = [str(a).strip().lower() for a in (key or [])]
            if accepted:
                pts = (q["points"] or 1) if str(ans or "").strip().lower() \
                    in accepted else 0.0
            else:
                has_long = True            # no key -> manual
        else:                              # long
            has_long = True
        cur.execute(f"""INSERT INTO lms_answers (attempt_id, question_id,
                        answer, score) VALUES ({PH},{PH},{PH},{PH})""",
                    (attempt_id, q["id"],
                     json.dumps(ans, ensure_ascii=False), pts))
        score += pts or 0.0
    grading = "pending" if has_long else "done"
    t = get_test(at["test_id"])
    passed = 0
    if grading == "done":
        passed = 1 if max_score and (score / max_score * 100) >= \
            (t["pass_pct"] or 70) else 0
    if (t.get("role") or "quiz") == "pre":
        passed = 1          # baseline only — never blocks completion
    cur.execute(f"""UPDATE lms_attempts SET submitted_at={PH}, score={PH},
                    max_score={PH}, passed={PH}, grading={PH}
                    WHERE id={PH}""",
                (_ts(), score, max_score, passed, grading, attempt_id))
    conn.commit()
    if grading == "done" and passed:
        _check_complete(at["enrollment_id"])
    cur.execute(f"SELECT * FROM lms_attempts WHERE id={PH}", (attempt_id,))
    return _one(cur)


def grading_queue():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT a.*, e.display_name, e.emp_key, t.title AS
                   test_title, t.pass_pct, c.code AS course_code
                   FROM lms_attempts a
                   JOIN lms_enrollments e ON e.id=a.enrollment_id
                   JOIN lms_tests t ON t.id=a.test_id
                   JOIN lms_courses c ON c.id=t.course_id
                   WHERE a.grading='pending' ORDER BY a.id""")
    return _rows(cur)


def attempt_answers(attempt_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT an.*, q.kind, q.text, q.points, q.answer_key
                    FROM lms_answers an JOIN lms_questions q
                    ON q.id=an.question_id
                    WHERE an.attempt_id={PH} ORDER BY q.seq""", (attempt_id,))
    return _rows(cur)


def grade_answer(answer_id, score, comment, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""UPDATE lms_answers SET score={PH}, comment={PH},
                    graded_by={PH} WHERE id={PH}""",
                (score, comment, actor, answer_id))
    conn.commit()


def finalize_grading(attempt_id, actor):
    conn = get_conn(); cur = conn.cursor()
    ans = attempt_answers(attempt_id)
    if any(a["score"] is None for a in ans):
        return False, "ยังตรวจไม่ครบทุกข้อ / some answers ungraded"
    score = sum(a["score"] or 0 for a in ans)
    max_score = sum(a["points"] or 1 for a in ans)
    cur.execute(f"SELECT * FROM lms_attempts WHERE id={PH}", (attempt_id,))
    at = _one(cur)
    t = get_test(at["test_id"])
    passed = 1 if max_score and (score / max_score * 100) >= \
        (t["pass_pct"] or 70) else 0
    cur.execute(f"""UPDATE lms_attempts SET score={PH}, max_score={PH},
                    passed={PH}, grading='done' WHERE id={PH}""",
                (score, max_score, passed, attempt_id))
    conn.commit()
    if passed:
        _check_complete(at["enrollment_id"])
    return True, f"คะแนน {score:g}/{max_score:g} — " + \
        ("ผ่าน ✅" if passed else "ไม่ผ่าน ❌")


def _check_complete(enrollment_id):
    """All lessons done + every test lesson has a passed attempt
    -> completed + certificate number."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_enrollments WHERE id={PH}",
                (enrollment_id,))
    e = _one(cur)
    if not e or e["status"] == "completed":
        return
    states = lesson_state(enrollment_id, e["course_id"])
    if not states or not all(d for _, d, _ in states):
        # test lessons get 'done' via passed attempt below; check those
        for l, d, _ in states:
            if d or l["kind"] != "test":
                continue
            cur.execute(f"""SELECT 1 FROM lms_attempts WHERE
                            enrollment_id={PH} AND test_id={PH} AND
                            passed=1""", (enrollment_id, l["test_id"]))
            if cur.fetchone():
                set_progress(enrollment_id, l["id"], done=True)
        states = lesson_state(enrollment_id, e["course_id"])
        if not all(d for _, d, _ in states):
            return
    ym = dt.date.today().strftime("%y%m")
    cur.execute(f"SELECT COUNT(*) FROM lms_enrollments WHERE cert_no "
                f"LIKE {PH}", (f"AMS-TRN-{ym}-%",))
    cert = f"AMS-TRN-{ym}-{cur.fetchone()[0] + 1:03d}"
    cur.execute(f"""UPDATE lms_enrollments SET status='completed',
                    completed_at={PH}, cert_no={PH} WHERE id={PH}""",
                (_ts(), cert, enrollment_id))
    conn.commit()


def record_test_pass(enrollment_id, lesson_id):
    set_progress(enrollment_id, lesson_id, done=True)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_enrollments WHERE id={PH}",
                (enrollment_id,))
    e = _one(cur)
    _check_complete(enrollment_id)


# ---------------------------------------------------------------- analytics
def analytics():
    out = {"courses": []}
    for c in list_courses(active_only=False):
        ens = course_enrollments(c["id"])
        if not ens:
            continue
        done = [e for e in ens if e["status"] == "completed"]
        today = dt.date.today().isoformat()
        overdue = [e for e in ens if e["status"] != "completed"
                   and (e["due_date"] or "9999") < today]
        out["courses"].append({
            "Course": f"{c['code']} {c['title_th'] or c['title_en']}",
            "Assigned": len(ens), "Completed": len(done),
            "Completion %": round(100 * len(done) / len(ens)),
            "Overdue": len(overdue)})
    return out


def cheat_flags():
    """Enrollments with failed/too-early completion attempts — the
    anti-cheat report for admins."""
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT e.emp_key, e.display_name, e.department,
                   c.code AS course, l.title AS lesson,
                   p.flags, p.done
                   FROM lms_progress p
                   JOIN lms_enrollments e ON e.id=p.enrollment_id
                   JOIN lms_lessons l ON l.id=p.lesson_id
                   JOIN lms_courses c ON c.id=e.course_id
                   WHERE COALESCE(p.flags,0) > 0
                   ORDER BY p.flags DESC""")
    return _rows(cur)


def overdue_list():
    today = dt.date.today().isoformat()
    rows = []
    for c in list_courses():
        for e in course_enrollments(c["id"]):
            if e["status"] != "completed" and (e["due_date"] or "9999") \
                    < today:
                rows.append({"Course": c["code"], "Emp": e["emp_key"],
                             "Name": e["display_name"],
                             "Dept": e["department"],
                             "Due": e["due_date"], "Status": e["status"]})
    return rows


# ---------------------------------------------------------------- certificate
def certificate_html(enr, course):
    import html as H
    d = str(enr.get("completed_at") or "")[:10]
    return f"""<!doctype html><html lang="th"><head><meta charset="utf-8">
<title>{H.escape(enr.get('cert_no') or '')}</title><style>
@page{{size:A4 landscape;margin:0}}
body{{font-family:'Sarabun','Leelawadee UI',Tahoma,sans-serif;margin:0;
 height:100vh;display:grid;place-items:center;
 background:linear-gradient(120deg,#f6f9ff,#ffffff 40%,#fdf3fa)}}
.frame{{width:88%;border:3px solid #715091;border-radius:18px;
 padding:46px 60px;text-align:center;background:#fff;position:relative;
 box-shadow:0 8px 40px rgba(15,23,42,.10)}}
.frame::before{{content:'';position:absolute;inset:8px;border:1.5px solid
 #d9c9ec;border-radius:12px}}
.band{{height:8px;border-radius:6px;margin:-46px -60px 30px;
 background:linear-gradient(90deg,#009ADE,#715091 55%,#E31D93)}}
.logo{{width:54px;height:54px;border-radius:13px;display:inline-grid;
 place-items:center;color:#fff;font-weight:800;font-size:26px;
 background:linear-gradient(135deg,#009ADE,#715091 55%,#E31D93)}}
h1{{font-size:30px;margin:14px 0 2px;color:#26303E}}
h2{{font-size:16px;margin:0 0 22px;color:#715091;font-weight:600}}
.name{{font-size:34px;font-weight:800;color:#715091;margin:8px 0}}
.course{{font-size:20px;font-weight:700;color:#26303E}}
.sub{{color:#5a6478;font-size:13px}}
.sig{{display:flex;justify-content:space-around;margin-top:42px}}
.sig div{{width:34%}}
.line{{border-bottom:1.5px dotted #8a93a6;height:36px;margin-bottom:6px}}
.cert{{position:absolute;bottom:14px;right:24px;font-size:11px;color:#8a93a6}}
.noprint{{position:fixed;top:8px;right:10px}}
.noprint button{{padding:8px 16px;border:0;border-radius:8px;color:#fff;
 font-weight:700;cursor:pointer;
 background:linear-gradient(135deg,#009ADE,#715091)}}
@media print{{.noprint{{display:none}}}}
</style></head><body onload="window.print()">
<div class="noprint"><button onclick="window.print()">🖨️ Print / PDF</button>
</div>
<div class="frame"><div class="band"></div>
<span class="logo">A</span>
<h1>ประกาศนียบัตร · Certificate of Completion</h1>
<h2>ANCA Manufacturing Solutions (Thailand) Ltd.</h2>
<div class="sub">ขอมอบประกาศนียบัตรฉบับนี้เพื่อรับรองว่า /
This is to certify that</div>
<div class="name">{H.escape(enr.get('display_name') or '')}</div>
<div class="sub">ได้ผ่านการอบรมหลักสูตร / has successfully completed</div>
<div class="course">{H.escape(course.get('title_th') or '')}<br>
<span style="font-size:15px;color:#5a6478">
{H.escape(course.get('title_en') or '')}</span></div>
<div class="sub" style="margin-top:14px">
รหัสหลักสูตร / Course code: <b>{H.escape(course.get('dsd_code') or
course.get('code') or '')}</b> &nbsp;•&nbsp; ระยะเวลา / Duration:
<b>{course.get('hours_display') or ''} ชั่วโมง</b> &nbsp;•&nbsp;
วันที่สำเร็จการอบรม / Date of completion: <b>{H.escape(d)}</b></div>
<div class="sig">
<div><div class="line"></div>ฝ่ายทรัพยากรบุคคล / HR Department</div>
<div><div class="line"></div>ผู้จัดการทั่วไป / General Manager</div>
</div>
<div class="cert">เลขที่ / Certificate No. {H.escape(enr.get('cert_no')
or '')}</div>
</div></body></html>"""


# ============================================================ DSD REPORTS
CRITERIA_TH = ["1. ความรู้จากการฝึกอบรม", "2. ทักษะในการปฏิบัติงาน",
               "3. ทัศนคติที่มีต่อการปฏิบัติงาน",
               "4. การแก้ปัญหาในการทำงาน",
               "5. ความตระหนักในด้านความปลอดภัย"]
INCOME_RANGES = ["less than 10,000 (น้อยกว่า 10,000)", "10,001-20,000",
                 "20,001-30,000", "30,001-40,000", "40,001-50,000",
                 "more than 50,001 (มากกว่า 50,001)", "N/A (ไม่ระบุ)"]


def save_skill_eval(enrollment_id, c, income_range, productivity,
                    evaluator):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"DELETE FROM lms_skill_evals WHERE enrollment_id={PH}",
                (enrollment_id,))
    cur.execute(f"""INSERT INTO lms_skill_evals (enrollment_id, c1, c2, c3,
                    c4, c5, income_range, productivity, evaluator,
                    eval_date) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},
                    {PH},{PH},{PH})""",
                (enrollment_id, c[0], c[1], c[2], c[3], c[4], income_range,
                 1 if productivity else 0, evaluator,
                 dt.date.today().isoformat()))
    conn.commit()


def get_skill_eval(enrollment_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM lms_skill_evals WHERE enrollment_id={PH}",
                (enrollment_id,))
    return _one(cur)


def suggest_c1(enrollment_id):
    """Auto-suggest criterion 1 (ความรู้จากการฝึกอบรม) from the pre/post
    comparison: +20 points or more -> 2, improved -> 1, else 0."""
    pre, post, imp = pre_post(enrollment_id)
    if imp is None:
        return 1 if (post or 0) >= 60 else 0
    return 2 if imp >= 20 else (1 if imp > 0 else 0)


def _th_name_parts(display_name, master_rec=None):
    """'นายชัยนันท์ สิงห์สน' -> (นาย, ชัยนันท์, สิงห์สน)."""
    nm = (master_rec or {}).get("emp_name_th") or display_name or ""
    nm = " ".join(str(nm).split())
    title = ""
    for t in ("นางสาว", "นาง", "นาย", "น.ส.", "Mr.", "Mrs.", "Ms."):
        if nm.startswith(t):
            title = t.rstrip(".") if t.startswith("น.ส") else t
            nm = nm[len(t):].strip()
            break
    parts = nm.split(" ", 1)
    first = parts[0] if parts else ""
    last = parts[1].strip() if len(parts) > 1 else ""
    return title, first, last


def dsd_course_rows(course_id, completed_only=False, date_from=None,
                    date_to=None):
    """Per-enrollee compliance rows: pre/post/Δ, watch hours, eval scores,
    potential result — everything the Generic9 needs, pulled live.
    Filters: completed_only; completed_at between date_from..date_to."""
    out = []
    ens = course_enrollments(course_id)
    if completed_only:
        ens = [e for e in ens if e["status"] == "completed"]
    if date_from:
        ens = [e for e in ens if (str(e.get("completed_at") or "")[:10]
                                  or "9999") >= str(date_from)]
    if date_to:
        ens = [e for e in ens if (str(e.get("completed_at") or "")[:10]
                                  or "0000") <= str(date_to)]
    for e in ens:
        rec = edb.get_record(emp_no=e["emp_key"]) \
            if not str(e["emp_key"]).startswith("user:") else None
        pre, post, imp = pre_post(e["id"])
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"""SELECT COALESCE(SUM(watch_seconds),0) FROM
                        lms_progress WHERE enrollment_id={PH}""", (e["id"],))
        wh = round(float(cur.fetchone()[0] or 0) / 3600, 2)
        ev = get_skill_eval(e["id"]) or {}
        c = [ev.get(f"c{i}") for i in range(1, 6)]
        if c[0] is None:
            c[0] = suggest_c1(e["id"])
        total = sum(x or 0 for x in c)
        title, first, last = _th_name_parts(e["display_name"], rec)
        out.append({
            "enrollment_id": e["id"], "emp_key": e["emp_key"],
            "id_card": (rec or {}).get("id_card") or "",
            "title": title, "first": first, "last": last,
            "position": (rec or {}).get("title") or "",
            "gender": (rec or {}).get("sex") or "",
            "dept": e.get("department"),
            "status": e["status"], "pre": pre, "post": post, "imp": imp,
            "watch_hours": wh,
            "c": c, "total": total,
            "potential": "เพิ่มขึ้น" if total > 0 else "คงเดิม",
            "productivity": bool(ev.get("productivity")),
            "income": ev.get("income_range") or "N/A (ไม่ระบุ)",
            "cert_no": e.get("cert_no")})
    return out


def thai_id_check(idc):
    """Replaces the template's macro: 13-digit + checksum validation.
    Returns 'ถูกต้อง' | 'ไม่ครบ13หลัก' | 'ผิด'."""
    d = "".join(ch for ch in str(idc or "") if ch.isdigit())
    if len(d) != 13:
        return "ไม่ครบ13หลัก"
    chk = (11 - sum(int(d[i]) * (13 - i) for i in range(12)) % 11) % 10
    return "ถูกต้อง" if chk == int(d[12]) else "ผิด"


def generic9_xlsx(course_ids, completed_only=True, date_from=None,
                  date_to=None):
    """FULL-FIDELITY replica of the official Generic9 (rev1): same sheets,
    cell positions, colours (Angsana New, yellow headers, FFFF99 score
    area, FFCC99 assessment cols, green income col), verbatim instruction
    texts, 53 pre-formatted rows, the blue spotlight row + radar chart,
    the income list, and the Data sheet with the red officer title and
    coloured gender totals. The template's macro buttons cannot be
    embedded — instead their FUNCTION runs here: Thai-ID checksum,
    13-digit and duplicate checks are computed and the results pre-filled
    in the same cells the macros would write."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import RadarChart, Reference

    if isinstance(course_ids, int):
        course_ids = [course_ids]
    rows = []
    for cid in course_ids:
        rows += dsd_course_rows(cid, completed_only, date_from, date_to)

    INCOME_DIGIT = {v: i + 1 for i, v in enumerate(INCOME_RANGES)}
    AN = "Angsana New"
    F = lambda **kw: Font(name=AN, size=kw.pop("size", 14), **kw)
    CT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    TH = Border(*[Side(style="thin")] * 4)
    YEL = PatternFill("solid", fgColor="FFFF00")
    SCORE = PatternFill("solid", fgColor="FFFF99")
    PEACH = PatternFill("solid", fgColor="FFCC99")
    GREEN = PatternFill("solid", fgColor="339966")
    BLUE = PatternFill("solid", fgColor="99CCFF")
    CREAM = PatternFill("solid", fgColor="FFFF99")
    GRAY = PatternFill("solid", fgColor="C0C0C0")

    wb = Workbook(); ws = wb.active; ws.title = "Summary"

    # ---------- instruction texts (verbatim from the original) ----------
    ws.merge_cells("A1:E4")
    c = ws["A1"]
    c.value = ("***อย่าลืมเปิด Macro เพื่อใช้งาน\n"
               "ปุ่มตรวจสอบเลขบัตร/ชื่อ ซ้ำ")
    c.font = F(bold=True, color="FF0000", size=16)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells("O1:T4")
    c = ws["O1"]
    c.value = ("กรณีเลขบัตรเป็นคนต่างด้าว ซึ่งมีเลข 0 นำหน้า ระบบจะฟ้องว่า "
               "ไม่ครบ 13 หลัก  แต่ระบบยังคงนำเข้าฐานข้อมูลได้ตามปกติ")
    c.font = F(color="008000", underline="single")
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells("A5:N5")
    c = ws["A5"]
    c.value = ("โปรดกรอกผลการประเมินคะแนนพร้อมบันทึกช่วงรายได้ของพนักงาน"
               "ในช่องสีเขียว (หากไม่ยินยอม กรุณาเลือกไม่ยินยอมบันทึก"
               "ช่วงรายได้)  แล้เลือกตัวเลือกศักยภาพในช่องว่างในช่อง"
               "สีฟ้าเท่านั้น")
    c.font = F(color="FF0000", size=12)

    # ---------- main table headers (rows 6-7) ----------
    heads = {1: "ที่", 2: "เลขบัตรประจำตัวประชาชน\n(1.ตรวจสอบเลขบัตรฯ ซ้ำ "
             "— ระบบตรวจให้แล้ว)", 3: "คำนำหน้า", 4: "ชื่อ", 5: "สกุล",
             6: "ตำแหน่ง", 12: "การประเมิน\nศักยภาพ",
             13: "การประเมิน\nผลิตภาพ", 14: "ช่วงรายได้",
             15: "ตรวจสอบความถูกต้องของเลขบัตรประชาชน",
             16: "ชื่อ-สกุล\n(2.ตรวจสอบชื่อ-สกุลซ้ำ — ระบบตรวจให้แล้ว)",
             17: "STATUS", 18: "เช็ค ü หาก ผ่าน"}
    for col, txt in heads.items():
        ws.merge_cells(start_row=6, start_column=col, end_row=7,
                       end_column=col)
        cc = ws.cell(row=6, column=col, value=txt)
        cc.font = F(bold=True, size=12); cc.alignment = CT
        cc.fill = YEL; cc.border = TH
        ws.cell(row=7, column=col).border = TH
    ws.merge_cells(start_row=6, start_column=7, end_row=6, end_column=11)
    cc = ws.cell(row=6, column=7,
                 value="ผลการประเมินศักยภาพผู้ผ่านการอบรม")
    cc.font = F(bold=True, size=12); cc.alignment = CT
    cc.fill = YEL; cc.border = TH
    for i, name in enumerate(CRITERIA_TH):
        cc = ws.cell(row=7, column=7 + i, value=name)
        cc.font = F(bold=True, size=11); cc.alignment = CT
        cc.fill = YEL; cc.border = TH
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 78

    # ---------- 53 pre-formatted rows (8..60), fill data on top ----------
    ids_seen, names_seen = {}, {}
    for x in rows:
        ids_seen[str(x["id_card"])] = ids_seen.get(str(x["id_card"]), 0) + 1
        nm = f"{x['first']} {x['last']}"
        names_seen[nm] = names_seen.get(nm, 0) + 1
    for r in range(8, 61):
        for col in range(1, 19):
            cc = ws.cell(row=r, column=col)
            cc.border = TH; cc.font = F(size=12)
            if 7 <= col <= 11:
                cc.fill = SCORE
            elif col in (12, 13, 15):
                cc.fill = PEACH
            elif col == 14:
                cc.fill = GREEN
    r = 8
    for i, x in enumerate(rows, 1):
        nm = f"{x['first']} {x['last']}"
        idchk = thai_id_check(x["id_card"])
        dup = (ids_seen.get(str(x["id_card"]), 0) > 1
               or names_seen.get(nm, 0) > 1)
        if dup:
            idchk += " /ซ้ำ"
        prod_lbl = "เพิ่มขึ้น" if x["productivity"] else "คงเดิม"
        vals = [i, str(x["id_card"]), x["title"], x["first"], x["last"],
                x["position"]] + [int(v or 0) for v in x["c"]] + \
               [x["potential"], prod_lbl, x["income"], idchk, nm, "",
                "ü" if (idchk == "ถูกต้อง") else ""]
        for j, v in enumerate(vals, 1):
            cc = ws.cell(row=r, column=j, value=v)
            if j == 2:
                cc.number_format = "@"
            if j in (1, 7, 8, 9, 10, 11):
                cc.alignment = Alignment(horizontal="center")
            if j == 15 and idchk != "ถูกต้อง":
                cc.font = F(size=12, color="FF0000")
            if j == 18:
                cc.font = Font(name="Wingdings", size=12)
                cc.alignment = Alignment(horizontal="center")
        r += 1

    # ---------- right block: spotlight + radar + summary + income list ---
    n = len(rows)
    inc_pot = sum(1 for x in rows if x["potential"] == "เพิ่มขึ้น")
    inc_prod = sum(1 for x in rows if x["productivity"])
    ws.merge_cells("U1:AD1")
    cc = ws["U1"]; cc.value = "ตารางแสดงผลศักยภาพของพนักงานที่ผ่านการฝึกอบรม"
    cc.font = F(bold=True, size=14); cc.alignment = CT
    ws.merge_cells("U3:V4")
    cc = ws["U3"]; cc.value = "ชื่อ-สกุล"
    cc.font = F(bold=True, size=12); cc.fill = YEL
    cc.alignment = CT; cc.border = TH
    ws.merge_cells("W3:AA3")
    cc = ws["W3"]; cc.value = "ผลการประเมินศักยภาพผู้ผ่านการอบรม"
    cc.font = F(bold=True, size=12); cc.fill = YEL
    cc.alignment = CT; cc.border = TH
    for i in range(5):
        cc = ws.cell(row=4, column=23 + i, value=i + 1)
        cc.font = F(bold=True, size=12); cc.fill = YEL
        cc.alignment = CT; cc.border = TH
    for col, txt in ((28, "รวมคะแนน"), (29, "การประเมิน\nศักยภาพ"),
                     (30, "การประเมิน\nผลิตภาพ")):
        ws.merge_cells(start_row=3, start_column=col, end_row=4,
                       end_column=col)
        cc = ws.cell(row=3, column=col, value=txt)
        cc.font = F(bold=True, size=12); cc.fill = YEL
        cc.alignment = CT; cc.border = TH
    ws.merge_cells("U5:V5")
    if rows:
        x = rows[0]
        cc = ws["U5"]; cc.value = f"{x['first']} {x['last']}"
        cc.fill = BLUE; cc.font = F(bold=True, size=12)
        cc.alignment = CT; cc.border = TH
        for i in range(5):
            cc = ws.cell(row=5, column=23 + i, value=int(x["c"][i] or 0))
            cc.fill = PEACH; cc.font = F(bold=True, size=12)
            cc.alignment = CT; cc.border = TH
        for col, v in ((28, x["potential"]),
                       (29, x["potential"]),
                       (30, "เพิ่มขึ้น" if x["productivity"] else "คงเดิม")):
            cc = ws.cell(row=5, column=col, value=v)
            cc.fill = PEACH; cc.font = F(bold=True, size=12)
            cc.alignment = CT; cc.border = TH
        # radar data (helper area, mirrors the original chart)
        labels = ["ความรู้", "การปฏิบัติงาน", "ทัศนคติ", "การจัดการ",
                  "ความปลอดภัย"]
        for i, (lb, vv) in enumerate(zip(labels, x["c"])):
            ws.cell(row=29 + i, column=32, value=lb)
            ws.cell(row=29 + i, column=33, value=int(vv or 0))
            ws.cell(row=29 + i, column=34, value=2)
        ch = RadarChart(); ch.type = "standard"
        ch.title = None; ch.height = 8.5; ch.width = 12
        data = Reference(ws, min_col=33, max_col=34, min_row=29, max_row=33)
        cats = Reference(ws, min_col=32, min_row=29, max_row=33)
        ch.add_data(data, titles_from_data=False)
        ch.set_categories(cats)
        ch.legend = None
        ws.add_chart(ch, "W7")
    ws.cell(row=16, column=21, value="สรุปภาพรวมการฝึกอบรม").font = \
        F(bold=True, size=13)
    summ = [("มีพนักงานที่เข้าร่วมอบรมทั้งหมด", n, "คน", 17),
            ("จำนวนผู้ผ่านการอบรมที่มีผลการพัฒนาศักยภาพเพิ่มขึ้น",
             inc_pot, "คน", 18),
            ("จำนวนผู้ผ่านการอบรมที่มีผลิตภาพแรงงานเพิ่มขึ้น",
             inc_prod, "คน", 19)]
    for lbl, v, unit, rr in summ:
        ws.cell(row=rr, column=21, value=lbl).font = F(size=12)
        ws.cell(row=rr, column=21).fill = CREAM
        ws.cell(row=rr, column=28, value="จำนวน").font = F(size=12)
        ws.cell(row=rr, column=29, value=v).font = F(bold=True, size=12)
        ws.cell(row=rr, column=30, value=unit).font = F(size=12)
    for i, cname in enumerate(CRITERIA_TH):
        avg = (sum((x["c"][i] or 0) for x in rows) / n) if n else 0
        ws.cell(row=21 + i, column=21, value=cname).font = F(size=12)
        ws.cell(row=21 + i, column=21).fill = CREAM
        ws.cell(row=21 + i, column=28, value="เฉลี่ย").font = F(size=12)
        ws.cell(row=21 + i, column=29,
                value=round(avg, 2)).font = F(bold=True, size=12)
        ws.cell(row=21 + i, column=30, value="คะแนน").font = F(size=12)
    allavg = (sum(x["total"] for x in rows) / (5 * n)) if n else 0
    ws.cell(row=26, column=21,
            value="คะแนนทุกเกณฑ์การประเมิน").font = F(bold=True, size=12)
    ws.cell(row=26, column=21).fill = CREAM
    ws.cell(row=26, column=28, value="เฉลี่ย").font = F(size=12)
    ws.cell(row=26, column=29, value=round(allavg, 2)).font = \
        F(bold=True, size=12)
    ws.cell(row=26, column=30, value="คะแนน").font = F(size=12)
    for i, inc in enumerate(INCOME_RANGES):
        ws.cell(row=28 + i, column=23, value=inc).font = F(size=12)

    widths = {1: 5, 2: 19, 3: 9, 4: 12, 5: 13, 6: 13, 12: 10, 13: 10,
              14: 24, 15: 16, 16: 16, 17: 8, 18: 9, 21: 36, 29: 9, 30: 9}
    for col in range(7, 12):
        widths[col] = 8
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ---------- Data sheet (officer layout) ----------
    ws2 = wb.create_sheet("Data")
    for j, h in enumerate(["PERSON_ID", "TITLE", "FIRSTNAME", "LASTNAME",
                           "GENDER", "POSITION", "STATUS", "REMARK",
                           "ตรวจสอบ\nPersonal_ID"], 1):
        cc = ws2.cell(row=1, column=j, value=h)
        cc.fill = YEL; cc.border = TH; cc.alignment = CT
        cc.font = Font(size=11, bold=(j == 9))
    ws2.merge_cells("J1:Q1")
    cc = ws2["J1"]
    cc.value = "ชีทนี้ สำหรับเจ้าหน้าที่\nกรมพัฒนาฝีมือแรงงาน"
    cc.font = Font(size=26, bold=True, color="FF0000")
    cc.fill = GRAY; cc.alignment = CT
    ws2.row_dimensions[1].height = 64
    males = females = 0
    for r0, x in enumerate(rows, 2):
        g = ("ชาย" if str(x["gender"]).upper().startswith(("M", "ช"))
             else "หญิง" if str(x["gender"]).upper().startswith(("F", "ห"))
             else "")
        males += g == "ชาย"; females += g == "หญิง"
        remark = "".join(str(int(v or 0)) for v in x["c"]) + \
            str(INCOME_DIGIT.get(x["income"], 7))
        idchk = thai_id_check(x["id_card"])
        vals = [str(x["id_card"]), x["title"], x["first"], x["last"], g,
                x["position"], "", remark, idchk]
        for j, v in enumerate(vals, 1):
            cc = ws2.cell(row=r0, column=j, value=v)
            cc.border = TH
            if j == 1:
                cc.number_format = "@"
            if j == 9 and idchk != "ถูกต้อง":
                cc.font = Font(size=11, color="FF0000")
        ws2.cell(row=r0, column=12, value=x["income"])
        ws2.cell(row=r0, column=14,
                 value=INCOME_DIGIT.get(x["income"], 7))
    for r0 in range(len(rows) + 2, 9):
        for j in range(1, 10):
            ws2.cell(row=r0, column=j).border = TH
    gpos = [("รวมชาย", males, YEL, 9), ("รวมหญิง", females,
             PatternFill("solid", fgColor="33CCCC"), 11),
            ("รวมทั้งสิ้น", males + females,
             PatternFill("solid", fgColor="339966"), 13)]
    for lbl, v, fill, rr in gpos:
        ws2.cell(row=rr, column=10, value=lbl)
        cc = ws2.cell(row=rr + 1, column=10, value=v)
        cc.fill = fill; cc.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDEFGHI", [16, 7, 12, 12, 8, 14, 8, 9, 13]):
        ws2.column_dimensions[col].width = w

    # ---------- เกณฑ์ผลิตภาพ (verbatim reference sheet) ----------
    ws3 = wb.create_sheet("เกณฑ์ผลิตภาพ")
    for row in PRODUCTIVITY_SHEET:
        ws3.append(list(row))
    ws3.column_dimensions["C"].width = 70

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def view_log_xlsx(course_id=None):
    """DSD evidence: per-person log (course, lesson, every open timestamp,
    watch seconds, completion) — keep >= 2 years; export any time."""
    import io
    from openpyxl import Workbook
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT e.emp_key, e.display_name, c.code, c.title_th,
                    l.title AS lesson, v.viewed_at,
                    p.watch_seconds, p.done, p.done_at
                    FROM lms_view_log v
                    JOIN lms_enrollments e ON e.id=v.enrollment_id
                    JOIN lms_courses c ON c.id=e.course_id
                    JOIN lms_lessons l ON l.id=v.lesson_id
                    LEFT JOIN lms_progress p ON p.enrollment_id=
                    v.enrollment_id AND p.lesson_id=v.lesson_id
                    {"WHERE c.id=" + str(int(course_id)) if course_id
                     else ""}
                    ORDER BY e.emp_key, v.viewed_at""")
    wb = Workbook(); ws = wb.active; ws.title = "DSD log"
    ws.append(["Emp", "Name", "Course", "Course name", "Lesson",
               "Opened at", "Watch seconds", "Done", "Done at"])
    for r in cur.fetchall():
        ws.append(list(r))
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_paper_html(test_id):
    """Printable exam paper (DSD requires submitting the test as PDF with
    the ยป.2 application)."""
    import html as H
    t = get_test(test_id)
    qs = questions(test_id)
    items = ""
    for q in qs:
        opts = json.loads(q["options"] or "[]")
        oo = "".join(f"<div style='margin-left:22px'>☐ {H.escape(o)}</div>"
                     for o in opts)
        items += (f"<div style='margin:10px 0'><b>ข้อ {q['seq']}</b> "
                  f"({q['points']:g} คะแนน) {H.escape(q['text'])}{oo}"
                  + ("<div style='border-bottom:1px dotted #777;"
                     "height:46px'></div>" if q['kind'] in ('short', 'long')
                     else "") + "</div>")
    return (f"<!doctype html><html lang='th'><head><meta charset='utf-8'>"
            f"<style>@page{{size:A4;margin:18mm}}body{{font-family:"
            f"'Sarabun',Tahoma;font-size:14px}}</style></head>"
            f"<body onload='window.print()'><h2>{H.escape(t['title'])} "
            f"({(t.get('role') or 'quiz').upper()})</h2>"
            f"<p>เกณฑ์ผ่าน {t['pass_pct']:g}% • จำนวน {len(qs)} ข้อ</p>"
            f"{items}</body></html>")
# auto-extracted from the official Generic9 template (sheet เกณฑ์ผลิตภาพ)
PRODUCTIVITY_SHEET = [
    ('หมวดใหญ่', 'หมวดย่อย', 'กลุ่มอุตสหกรรม', 'กลุ่ม 1', 'กลุ่ม 2', 'กลุ่ม 3', 'กลุ่ม 4', 'กลุ่ม 4'),
    ('A', 1, 'A 1 การเพาะปลูกและการเลี้ยงสัตว์ การล่าสัตว์และกิจกรรมบริการที่เกี่ยวข้อง', 0, 1, 0, 1, 0),
    ('A', 2, 'A 2 ป่าไม้และการทำไม้', 0, 1, 0, 0, 1),
    ('A', 3, 'A 3 การประมงและการเพาะเลี้ยงสัตว์น้ำ', 0, 1, 0, 1, 1),
    ('B', 5, 'B 5 การทำเหมืองถ่านหินและลิกไนต์', 0, 1, 0, 1, 0),
    ('B', 6, 'B 6 การผลิตน้ำมันดิบและก๊าซธรรมชาติ', 0, 1, 0, 0, 1),
    ('B', 7, 'B 7 การทำเหมืองสินแร่โลหะ', 0, 1, 1, 1, 0),
    ('B', 8, 'B 8 การทำเหมืองแร่และเหมืองหินอื่น ๆ', 0, 1, 0, 1, 0),
    ('B', 9, 'B 9 กิจกรรมบริการที่สนับสนุนการทำเหมืองแร่ และสนับสนุนการผลิตปิโตรเลียม', 0, 1, 1, 1, 0),
    ('C', 10, 'C 10 การผลิตผลิตภัณฑ์อาหาร', 0, 1, 0, 1, 0),
    ('C', 11, 'C 11 การผลิตเครื่องดื่ม', 0, 1, 0, 1, 0),
    ('C', 12, 'C 12 การผลิตสิ่งทอ', 0, 1, 0, 1, 0),
    ('C', 13, 'C 13 การผลิตเสื้อผ้าเครื่องแต่งกาย', 0, 1, 0, 1, 0),
    ('C', 14, 'C 14 การผลิตเครื่องหนัง และผลิตภัณฑ์ที่เกี่ยวข้อง', 0, 1, 1, 1, 0),
    ('C', 15, 'C 15 การผลิตไม้และผลิตภัณฑ์จากไม้ และไม้ก๊อก ยกเว้น เฟอร์นิเจอร์ การผลิตสิ่งของจากฟางและวัสดุถักสานอื่น ๆ', 0, 1, 0, 1, 0),
    ('C', 16, 'C 16 การผลิตกระดาษและ ผลิตภัณฑ์ที่แปรรูปจากกระดาษ', 0, 1, 0, 1, 0),
    ('C', 17, 'C 17 การพิมพ์ และการผลิตซ้ำสื่อบันทึก', 0, 1, 0, 1, 0),
    ('C', 18, 'C 18 การผลิตถ่านโค้ก และผลิตภัณฑ์จากการกลั่นปิโตรเลียม', 0, 1, 0, 1, 0),
    ('C', 19, 'C 19 การผลิตเคมีภัณฑ์ และผลิตภัณฑ์เคมี', 0, 1, 0, 1, 0),
    ('C', 20, 'C 20 การผลิตเภสัชภัณฑ์พื้นฐานและการผลิตสูตรตำรับทางเภสัชกรรม', 0, 1, 0, 1, 0),
    ('C', 21, 'C 21 การผลิตผลิตภัณฑ์ยางและพลาสติก', 0, 1, 0, 1, 0),
    ('C', 22, 'C 22 การผลิตผลิตภัณฑ์อื่น ๆ ที่ทำจากแร่อโลหะ', 0, 1, 0, 1, 0),
    ('C', 23, 'C 23 การผลิตโลหะขั้นมูลฐาน', 0, 1, 0, 1, 1),
    ('C', 24, 'C 24 การผลิตผลิตภัณฑ์ที่ทำจากโลหะประดิษฐ์ ยกเว้น เครื่องจักรและอุปกรณ์', 0, 1, 0, 1, 1),
    ('C', 25, 'C 25 การผลิตผลิตภัณฑ์คอมพิวเตอร์ อิเล็กทรอนิกส์ และอุปกรณ์ทางทัศนศาสตร์', 0, 1, 0, 1, 0),
    ('C', 26, 'C 26 การผลิตอุปกรณ์ไฟฟ้า', 0, 1, 0, 1, 1),
    ('C', 27, 'C 27 การผลิตเครื่องจักรและเครื่องมือ ซึ่งมิได้จัดประเภทไว้ในที่อื่น', 0, 1, 0, 1, 0),
    ('C', 28, 'C 28 การผลิตยานยนต์ รถพ่วงและรถกึ่งพ่วง', 0, 1, 0, 1, 0),
    ('C', 29, 'C 29 การผลิตอุปกรณ์ขนส่งอื่น ๆ', 0, 1, 0, 1, 0),
    ('C', 30, 'C 30 การผลิตเฟอร์นิเจอร์', 0, 1, 0, 1, 0),
    ('C', 31, 'C 31 การผลิตผลิตภัณฑ์ประเภทอื่น ๆ', 0, 1, 0, 1, 0),
    ('C', 32, 'C 32 การซ่อม และการติดตั้งเครื่องจักรและอุปกรณ์', 0, 1, 0, 0, 1),
    ('C', 33, 'C 33 การผลิตผลิตภัณฑ์จากยาสูบ', 0, 1, 0, 1, 0),
    ('D', 35, 'D 35 ไฟฟ้า ก๊าซ ไอน้ำและระบบการปรับอากาศ', 0, 1, 0, 1, 1),
    ('E', 36, 'E 36 การเก็บกักน้ำ การจัดหาน้ำ และการจ่ายน้ำ', 0, 1, 1, 1, 0),
    ('E', 37, 'E 37 การจัดการน้ำเสีย', 0, 1, 0, 1, 1),
    ('E', 38, 'E 38 การเก็บรวบรวมของเสีย  การบำบัดและการกำจัดของเสีย การนำของเสียกลับมาใช้ใหม่', 0, 1, 0, 1, 0),
    ('E', 39, 'E 39 กิจกรรมการบำบัด และการจัดบริการเก็บของเสียอื่น ๆ', 0, 1, 1, 1, 0),
    ('F', 41, 'F 41 การก่อสร้างอาคาร', 0, 1, 0, 0, 1),
    ('F', 42, 'F 42 งานวิศวกรรมโยธา', 0, 1, 0, 1, 1),
    ('F', 43, 'F 43 งานก่อสร้างเฉพาะทาง', 0, 1, 0, 1, 1),
    ('G', 45, 'G 45 การขายส่ง การขายปลีก การซ่อมยานยนต์และจักรยานยนต์', 0, 1, 1, 1, 0),
    ('G', 46, 'G 46 การขายส่ง ยกเว้น ยานยนต์และจักรยานยนต์', 0, 1, 1, 1, 0),
    ('G', 47, 'G 47 การขายปลีก ยกเว้น ยานยนต์และจักรยานยนต์', 0, 1, 1, 0, 0),
    ('H', 49, 'H 49 การขนส่งทางบกและการขนส่งทางท่อลำเลียง', 0, 1, 1, 1, 0),
    ('H', 50, 'H 50 การขนส่งทางน้ำ', 0, 1, 1, 0, 1),
    ('H', 51, 'H 51 การขนส่งทางอากาศ', 0, 1, 0, 0, 1),
    ('H', 52, 'H 52 กิจกรรมที่เกี่ยวกับคลังสินค้าและกิจกรรมสนับสนุนการขนส่ง', 0, 1, 0, 0, 1),
    ('H', 53, 'H 53 กิจกรรมไปรษณีย์และการรับส่งเอกสาร/สิ่งของ', 0, 1, 1, 0, 0),
    ('I', 55, 'I 55 ที่พักแรม', 0, 1, 1, 0, 0),
    ('I', 56, 'I 56 การบริการด้านอาหารและเครื่องดื่ม', 0, 1, 1, 0, 0),
    ('J', 58, 'J 58 การจัดพิมพ์จำหน่ายหรือเผยแพร่', 0, 1, 0, 1, 0),
    ('J', 59, 'J 59 กิจกรรมการผลิตภาพยนตร์ วีดิทัศน์และรายการโทรทัศน์ การบันทึกเสียงและการจัดพิมพ์จำหน่ายหรือเผยแพร่ดนตรี', 0, 1, 0, 0, 0),
    ('J', 60, 'J 60 กิจกรรมการจัดผังรายการโทรทัศน์ และกิจกรรมการแพร่ภาพกระจายเสียง', 0, 1, 0, 0, 1),
    ('J', 61, 'J 61 การโทรคมนาคม', 0, 1, 0, 1, 0),
    ('J', 62, 'J 62 กิจกรรมการจัดทำโปรแกรมคอมพิวเตอร์ การให้คำปรึกษาเกี่ยวกับคอมพิวเตอร์ และกิจกรรมที่เกี่ยวข้อง', 0, 1, 1, 1, 0),
    ('J', 63, 'J 63 กิจกรรมการบริการสารสนเทศ', 0, 1, 1, 0, 0),
    ('K', 64, 'K 64 กิจกรรมบริการทางการเงิน ยกเว้น การประกันภัยและกองทุนบำเหน็จบำนาญ', 0, 1, 0, 1, 0),
    ('K', 65, 'K 65 การประกันภัย การประกันภัยต่อและกองทุนบำเหน็จบำนาญ ยกเว้น การประกันสังคมภาคบังคับ', 0, 1, 1, 1, 0),
    ('K', 66, 'K 66 กิจกรรมสนับสนุนบริการทางการเงินและกิจกรรมการประกันภัย', 0, 1, 0, 1, 0),
    ('L', 68, 'L 68 กิจกรรมเกี่ยวกับอสังหาริมทรัพย์', 0, 1, 1, 1, 0),
    ('M', 69, 'M 69 กิจกรรมทางกฎหมายและบัญชี', 0, 1, 0, 1, 0),
    ('M', 70, 'M 70 กิจกรรมของสำนักงานใหญ่ กิจกรรมการให้คำปรึกษาด้านการบริหารจัดการ', 0, 1, 1, 0, 0),
    ('M', 71, 'M 71 กิจกรรมงานสถาปัตยกรรมและวิศวกรรม การทดสอบและวิเคราะห์ทางเทคนิค', 0, 1, 0, 1, 0),
    ('M', 72, 'M 72 การวิจัยและพัฒนาเชิงวิทยาศาสตร์', 0, 1, 0, 1, 0),
    ('M', 73, 'M 73 การโฆษณาและการวิจัยตลาด', 0, 1, 1, 1, 0),
    ('M', 74, 'M 74 กิจกรรมทางวิชาชีพ วิทยาศาสตร์ และเทคนิคอื่นๆ', 0, 1, 1, 1, 0),
    ('M', 75, 'M 75 กิจกรรมเกี่ยวกับสัตวแพทย์', 0, 1, 1, 1, 0),
    ('N', 77, 'N 77 กิจกรรมการให้เช่าและให้เช่าแบบลีสซิ่ง', 0, 1, 0, 1, 0),
    ('N', 78, 'N 78 กิจกรรมการจัดหางาน', 0, 1, 1, 1, 0),
    ('N', 79, 'N 79 ตัวแทนธุรกิจการเดินทาง  ธุรกิจจัดนำเที่ยว และบริการสำรอง  และกิจกรรมที่เกี่ยวข้อง', 1, 1, 1, 0, 0),
    ('N', 80, 'N 80 กิจกรรมการบริการรักษาความปลอดภัยและการสืบสวน', 0, 1, 0, 0, 1),
    ('N', 81, 'N 81 กิจกรรมบริการสำหรับอาคารและภูมิทัศน์', 0, 1, 1, 0, 0),
    ('N', 82, 'N 82 การบริหารสำนักงาน บริการสนับสนุนสำนักงานและบริการสนับสนุนทางธุรกิจอื่น ๆ', 0, 1, 1, 1, 0),
    ('O', 84, 'O 84 การบริหารราชการและการป้องกันประเทศ รวมถึงการประกันสังคมภาคบังคับ', 0, 1, 1, 1, 0),
    ('P', 85, 'P 85 การศึกษา', 0, 1, 1, 1, 0),
    ('Q', 86, 'Q 86 กิจกรรมด้านสุขภาพของมนุษย์', 0, 1, 1, 0, 0),
    ('Q', 87, 'Q 87 กิจกรรมการดูแลรักษาในสถานที่ที่มีที่พักและมีคนดูแลประจำ', 0, 1, 1, 0, 0),
    ('Q', 88, 'Q 88 กิจกรรมสังคมสงเคราะห์โดยไม่มีที่พักอาศัย', 0, 1, 1, 0, 0),
    ('R', 90, 'R 90 กิจกรรมการสร้างสรรค์ศิลปะและความบันเทิง', 0, 1, 1, 0, 0),
    ('R', 91, 'R 91 ห้องสมุด หอจดหมายเหตุ พิพิธภัณฑ์และกิจกรรมทางด้านวัฒนธรรมอื่น ๆ', 1, 1, 1, 0, 0),
    ('R', 92, 'R 92 กิจกรรมการพนันและการเสี่ยงโชค', '', '', '', '', ''),
    ('R', 93, 'R 93 กิจกรรมด้านการกีฬา ความบันเทิงและนันทนาการ', 0, 1, 1, 0, 0),
    ('S', 94, 'S 94 กิจกรรมขององค์กรสมาชิก', 0, 1, 0, 1, 1),
    ('S', 95, 'S 95 การซ่อมคอมพิวเตอร์และของใช้ส่วนบุคคลและของใช้ในครัวเรือน', 0, 1, 1, 1, 0),
    ('S', 96, 'S 96 กิจกรรมการบริการอื่น ๆ ส่วนบุคคล', 0, 1, 1, 0, 0),
    ('T', 97, 'T 97 กิจกรรมการจ้างงานในครัวเรือนในฐานะที่เป็นนายจ้างส่วนบุคคล', 0, 1, 0, 0, 1),
    ('T', 98, 'T 98 กิจกรรมการผลิตสินค้าและบริการที่ทำขึ้นเองเพื่อใช้ในครัวเรือน ซึ่งไม่สามารถจำแนกกิจกรรมได้อย่างชัดเจน', 0, 1, 0, 1, 0),
    ('U', 99, 'U 99 กิจกรรมขององค์การระหว่างประเทศและภาคีสมาชิก', 0, 1, 0, 1, 0),
]
