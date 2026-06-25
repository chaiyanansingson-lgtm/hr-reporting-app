# lib/working_hours_db.py
# Ingests the "000_Actual_working_hours" template (Data sheet, per-employee)
# and produces the aggregates behind the ASM-style HR dashboard. Every
# aggregate accepts emp_nos -> scope to a manager's team (real-time, on the
# latest upload).  Emp ID in the sheet == employee master emp_no.
import datetime as dt
import io
from lib.db import get_conn, IS_POSTGRES, PH

SERIAL = "SERIAL PRIMARY KEY" if IS_POSTGRES else \
         "INTEGER PRIMARY KEY AUTOINCREMENT"
STD_DEFAULT = 167.0   # standard paid working hrs / head / month (editable)

# Data-sheet header -> our field (matched case-insensitively, "contains")
COLMAP = [
    ("emp id", "emp_id"), ("name", "name"), ("cost ctr", "cost_ctr"),
    ("department", "department"), ("function key", "function_key"),
    ("class", "klass"), ("func override", "func_override"),
    ("ot x1.5 (hr)", "ot15"), ("ot x3 (hr)", "ot3"), ("al (days)", "al_days"),
    ("other leave", "other_leave_days"), ("ot x1 keyed", "ot1"),
    ("ot x1.5 unsubmitted", "ot15_uns"), ("ot x3 holiday", "ot3_hol"),
]


def migrate():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""CREATE TABLE IF NOT EXISTS wh_uploads (
        id {SERIAL}, period TEXT, filename TEXT, n_rows INTEGER,
        std_hours REAL DEFAULT 167, uploaded_by TEXT, uploaded_at TEXT)""")
    cur.execute(f"""CREATE TABLE IF NOT EXISTS wh_rows (
        id {SERIAL}, upload_id INTEGER NOT NULL,
        emp_id TEXT, name TEXT, cost_ctr TEXT, department TEXT,
        function_key TEXT, klass TEXT, func_override TEXT,
        ot15 REAL, ot3 REAL, al_days REAL, other_leave_days REAL,
        ot1 REAL, ot15_uns REAL, ot3_hol REAL)""")
    conn.commit()


def _num(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def parse_upload(file_bytes):
    """Read the 'Data' sheet -> (rows, period_label). Falls back to the
    first sheet if no sheet literally named 'Data'."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
    # locate header row (the one containing "Emp ID")
    hdr_row = 1
    for r in range(1, 6):
        vals = [str(ws.cell(r, c).value or "").lower()
                for c in range(1, ws.max_column + 1)]
        if any("emp id" in v for v in vals):
            hdr_row = r; break
    headers = {c: str(ws.cell(hdr_row, c).value or "").lower().strip()
               for c in range(1, ws.max_column + 1)}
    colidx = {}
    for needle, field in COLMAP:
        for c, h in headers.items():
            if needle in h:
                colidx[field] = c; break
    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        emp = ws.cell(r, colidx.get("emp_id", 1)).value
        if emp in (None, ""):
            continue
        row = {"emp_id": str(emp).strip()}
        for field, c in colidx.items():
            if field == "emp_id":
                continue
            v = ws.cell(r, c).value
            if field in ("name", "cost_ctr", "department", "function_key",
                         "klass", "func_override"):
                row[field] = (str(v).strip() if v is not None else "")
            else:
                row[field] = _num(v)
        rows.append(row)
    # period label: try the summary sheet name (e.g. "May 26")
    period = next((s for s in wb.sheetnames if s not in ("Data", "Map")),
                  None) or dt.date.today().strftime("%b %y")
    return rows, period


def apply_upload(rows, period, filename, std_hours, actor):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""INSERT INTO wh_uploads (period, filename, n_rows,
                    std_hours, uploaded_by, uploaded_at)
                    VALUES ({PH},{PH},{PH},{PH},{PH},{PH})""",
                (period, filename, len(rows), std_hours, actor,
                 dt.datetime.now().isoformat(timespec="seconds")))
    conn.commit()
    if IS_POSTGRES:
        cur.execute("SELECT MAX(id) FROM wh_uploads"); uid = cur.fetchone()[0]
    else:
        uid = cur.lastrowid
    for x in rows:
        cur.execute(f"""INSERT INTO wh_rows (upload_id, emp_id, name,
            cost_ctr, department, function_key, klass, func_override,
            ot15, ot3, al_days, other_leave_days, ot1, ot15_uns, ot3_hol)
            VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},
            {PH},{PH},{PH},{PH})""",
                    (uid, x.get("emp_id"), x.get("name"), x.get("cost_ctr"),
                     x.get("department"), x.get("function_key"),
                     x.get("klass"), x.get("func_override"),
                     x.get("ot15", 0), x.get("ot3", 0), x.get("al_days", 0),
                     x.get("other_leave_days", 0), x.get("ot1", 0),
                     x.get("ot15_uns", 0), x.get("ot3_hol", 0)))
    conn.commit()
    return uid, len(rows)


def latest_upload():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT id, period, filename, n_rows, std_hours, "
                "uploaded_by, uploaded_at FROM wh_uploads "
                "ORDER BY id DESC LIMIT 1")
    r = cur.fetchone()
    if not r:
        return None
    return {"id": r[0], "period": r[1], "filename": r[2], "n_rows": r[3],
            "std_hours": r[4], "uploaded_by": r[5], "uploaded_at": r[6]}


def rows_for(upload_id, emp_nos=None):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM wh_rows WHERE upload_id={PH}", (upload_id,))
    cols = [d[0] for d in cur.description]
    out = [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
           for r in cur.fetchall()]
    if emp_nos is not None:
        keep = {str(e) for e in emp_nos}
        out = [x for x in out if str(x["emp_id"]) in keep]
    return out


def _ot(x):
    return ((x.get("ot15") or 0) + (x.get("ot3") or 0) + (x.get("ot1") or 0)
            + (x.get("ot15_uns") or 0) + (x.get("ot3_hol") or 0))


def dashboard(upload_id, emp_nos=None, std_hours=STD_DEFAULT):
    """All figures behind the ASM HR dashboard, scoped to emp_nos if given."""
    rows = rows_for(upload_id, emp_nos)
    hc = len(rows)
    perm = sum(1 for x in rows if (x.get("klass") or "").upper() == "PER")
    sub = sum(1 for x in rows if (x.get("klass") or "").upper() == "SUB")
    other = hc - perm - sub
    al_days = sum(x.get("al_days") or 0 for x in rows)
    oth_days = sum(x.get("other_leave_days") or 0 for x in rows)
    absent_hrs = (al_days + oth_days) * 8.0
    ot_total = sum(_ot(x) for x in rows)
    working_planned = hc * std_hours
    working_actual = max(working_planned - absent_hrs, 0)
    pct_abs = (absent_hrs / working_planned * 100) if working_planned else 0
    pct_ot = (ot_total / working_actual * 100) if working_actual else 0

    # by department (Working Hour vs OT + absenteeism by dept)
    byd = {}
    for x in rows:
        d = x.get("department") or x.get("function_key") or "—"
        a = byd.setdefault(d, {"dept": d, "hc": 0, "ot": 0.0,
                               "absent_hrs": 0.0})
        a["hc"] += 1
        a["ot"] += _ot(x)
        a["absent_hrs"] += ((x.get("al_days") or 0) +
                            (x.get("other_leave_days") or 0)) * 8.0
    dept_rows = []
    for a in byd.values():
        wp = a["hc"] * std_hours
        wa = max(wp - a["absent_hrs"], 0)
        dept_rows.append({
            "Department": a["dept"], "HC": a["hc"],
            "Working Hrs": round(wa, 0), "OT Hrs": round(a["ot"], 1),
            "% OT": round((a["ot"] / wa * 100) if wa else 0, 1),
            "Absent Hrs": round(a["absent_hrs"], 1),
            "% Absent": round((a["absent_hrs"] / wp * 100) if wp else 0, 2)})
    dept_rows.sort(key=lambda r: -r["OT Hrs"])

    return {
        "hc": hc, "perm": perm, "sub": sub, "other": other,
        "al_days": round(al_days, 1), "oth_days": round(oth_days, 1),
        "absent_hrs": round(absent_hrs, 1), "ot_total": round(ot_total, 1),
        "working_planned": round(working_planned, 0),
        "working_actual": round(working_actual, 0),
        "pct_abs": round(pct_abs, 2), "pct_ot": round(pct_ot, 1),
        "dept_rows": dept_rows,
        "criterion": {"ลาป่วย+ลากิจ (AL)": round(al_days, 1),
                      "ลาอื่นๆ (Other)": round(oth_days, 1)},
    }


def summary_table(upload_id, emp_nos=None, std_hours=STD_DEFAULT):
    """Working-hours summary by department (the 'May 26' style table)."""
    d = dashboard(upload_id, emp_nos, std_hours)
    return d["dept_rows"]


def list_uploads():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT id, period, filename, n_rows, std_hours, "
                "uploaded_at FROM wh_uploads ORDER BY id DESC")
    return [{"id": r[0], "period": r[1], "filename": r[2], "n_rows": r[3],
             "std_hours": r[4], "uploaded_at": r[5]} for r in cur.fetchall()]
