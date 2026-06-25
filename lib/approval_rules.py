# lib/approval_rules.py
# ---------------------------------------------------------------------------
# Configurable approval lines. Admin/Superadmin defines, per request kind and
# optional scope (a department, or "all"), an ordered list of approval steps.
# Each step is either a specific person (by staff number) or "the requester's
# next manager". When a request is submitted, the most specific matching line
# is used to build the approval chain; if no line is configured, the system
# falls back to the manager-walk (lib.approval_db.resolve_chain).
# ---------------------------------------------------------------------------
import datetime as dt
import json

from lib.db import get_conn, IS_POSTGRES, PH

SERIAL = "SERIAL PRIMARY KEY" if IS_POSTGRES else \
    "INTEGER PRIMARY KEY AUTOINCREMENT"

KINDS = ["leave", "ot", "shift", "time_edit", "resign", "car", "po",
         "permit_out", "permit_entry", "stock"]
KIND_LABEL = {
    "leave": "ใบลา · Leave",
    "ot": "OT · Overtime",
    "shift": "เปลี่ยนกะ · Shift change",
    "time_edit": "แก้เวลา · Time edit",
    "resign": "ลาออก · Resignation",
    "car": "จองรถ · Car booking",
    "po": "ใบสั่งซื้อ · Purchase order",
    "permit_out": "นำของออก · Permit out",
    "permit_entry": "นำเข้า · Permit entry",
    "stock": "เบิกของ · Stock",
}
# step types
STEP_MANAGER = "manager"   # the requester's next manager up the tree
STEP_EMP = "emp"           # a specific person, by staff number


def migrate():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""CREATE TABLE IF NOT EXISTS approval_lines (
        id {SERIAL},
        request_kind TEXT NOT NULL,
        scope_type TEXT NOT NULL DEFAULT 'all',   -- 'all' | 'department'
        scope_value TEXT NOT NULL DEFAULT '*',
        steps_json TEXT NOT NULL DEFAULT '[]',
        updated_by TEXT, updated_at TEXT,
        UNIQUE(request_kind, scope_type, scope_value)
    )""")
    conn.commit()


def _rows(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) if IS_POSTGRES else dict(r)
            for r in cur.fetchall()]


def get_line(kind, scope_type, scope_value):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"""SELECT * FROM approval_lines WHERE request_kind={PH}
        AND scope_type={PH} AND scope_value={PH}""",
                (kind, scope_type, scope_value))
    r = _rows(cur)
    return r[0] if r else None


def list_lines(kind=None):
    conn = get_conn(); cur = conn.cursor()
    if kind:
        cur.execute(f"SELECT * FROM approval_lines WHERE request_kind={PH} "
                    "ORDER BY scope_type, scope_value", (kind,))
    else:
        cur.execute("SELECT * FROM approval_lines "
                    "ORDER BY request_kind, scope_type, scope_value")
    return _rows(cur)


def set_line(kind, scope_type, scope_value, steps, actor="system"):
    """steps = [{'type':'manager'} | {'type':'emp','value':'E123'}, ...]"""
    conn = get_conn(); cur = conn.cursor()
    sj = json.dumps(steps, ensure_ascii=False)
    now = dt.datetime.now().isoformat(timespec="seconds")
    existing = get_line(kind, scope_type, scope_value)
    if existing:
        cur.execute(f"""UPDATE approval_lines SET steps_json={PH},
            updated_by={PH}, updated_at={PH} WHERE id={PH}""",
                    (sj, actor, now, existing["id"]))
    else:
        cur.execute(f"""INSERT INTO approval_lines (request_kind, scope_type,
            scope_value, steps_json, updated_by, updated_at)
            VALUES ({PH},{PH},{PH},{PH},{PH},{PH})""",
                    (kind, scope_type, scope_value, sj, actor, now))
    conn.commit()


def delete_line(line_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"DELETE FROM approval_lines WHERE id={PH}", (line_id,))
    conn.commit()


def _match_line(kind, rec):
    """Most specific line: department match first, then 'all'."""
    dept = (rec.get("dept_location") or "").strip()
    if dept:
        line = get_line(kind, "department", dept)
        if line:
            return line
    return get_line(kind, "all", "*")


def resolve_configured_chain(kind, rec, max_levels=5):
    """Build [(level, approver_record), ...] from the configured line, or None
    if no line is configured for this kind/scope."""
    line = _match_line(kind, rec)
    if not line:
        return None
    try:
        steps = json.loads(line.get("steps_json") or "[]")
    except Exception:
        steps = []
    if not steps:
        return None
    from lib import employee_db as edb
    from lib import approval_db as adb
    chain, seen = [], {str(rec.get("emp_no"))}
    mgr_cursor = rec
    lvl = 1
    for spec in steps:
        approver = None
        if spec.get("type") == STEP_EMP and spec.get("value"):
            approver = edb.get_record(emp_no=str(spec["value"]))
        elif spec.get("type") == STEP_MANAGER:
            approver = adb._find_by_name(mgr_cursor.get("mgr_name"))
            if approver:
                mgr_cursor = approver
        if approver and str(approver.get("emp_no")) not in seen:
            chain.append((lvl, approver))
            seen.add(str(approver.get("emp_no")))
            lvl += 1
        if lvl > max_levels:
            break
    return chain or None


def describe(rec_lookup, steps):
    """Human-readable preview of the steps for the admin UI."""
    out = []
    for i, s in enumerate(steps, 1):
        if s.get("type") == STEP_MANAGER:
            out.append(f"{i}. ผู้จัดการลำดับถัดไป · next manager")
        else:
            nm = rec_lookup.get(str(s.get("value")), s.get("value"))
            out.append(f"{i}. {s.get('value')} — {nm}")
    return out


# ---------------------------------------------------------------------------
# Routing-Request format (image): Applicant + Petitioner1/2 + Approver1/2/3 +
# Reviewer. Stored as ordered emp steps tagged with a role label, so the same
# resolver/queue/notification engine drives it. Applicant = the requester
# (not configured here). Empty slots are skipped.
# ---------------------------------------------------------------------------
ROUTING_SLOTS = [
    ("petitioner1", "Petitioner 1", "ผู้เสนอ 1"),
    ("petitioner2", "Petitioner 2", "ผู้เสนอ 2"),
    ("approver1", "Approver 1", "ผู้อนุมัติ 1"),
    ("approver2", "Approver 2", "ผู้อนุมัติ 2"),
    ("approver3", "Approver 3", "ผู้อนุมัติ 3"),
    ("reviewer", "Reviewer", "ผู้ตรวจสอบ"),
]


def set_routing(kind, scope_type, scope_value, slot_emps, actor="system"):
    """slot_emps: {'petitioner1':'E001', ...}. Builds ordered emp steps."""
    steps = []
    for key, en, _th in ROUTING_SLOTS:
        emp = str(slot_emps.get(key) or "").strip()
        if emp:
            steps.append({"type": STEP_EMP, "value": emp, "role": en})
    set_line(kind, scope_type, scope_value, steps, actor)


def get_routing(kind, scope_type, scope_value):
    """{'petitioner1':'E001', ...} reconstructed from the saved steps."""
    import json as _json
    out = {k: "" for k, _e, _t in ROUTING_SLOTS}
    line = get_line(kind, scope_type, scope_value)
    if not line:
        return out
    try:
        steps = _json.loads(line.get("steps_json") or "[]")
    except Exception:
        steps = []
    role_to_key = {en: k for k, en, _t in ROUTING_SLOTS}
    approver_slots = ["approver1", "approver2", "approver3"]
    ai = 0
    for s in steps:
        if s.get("type") != STEP_EMP:
            continue
        role = s.get("role")
        if role and role in role_to_key:
            out[role_to_key[role]] = str(s.get("value"))
        elif ai < len(approver_slots):
            out[approver_slots[ai]] = str(s.get("value")); ai += 1
    return out


def bulk_apply(rows, actor="system", valid_emp=None):
    """rows: dicts with kind, scope_type, scope_value + the 6 slot keys.
    valid_emp: optional set of known emp_nos to validate against.
    Returns (applied:int, errors:list[str])."""
    applied, errors = 0, []
    for i, r in enumerate(rows, 1):
        kind = str(r.get("kind") or "").strip().lower()
        if kind not in KINDS:
            errors.append(f"แถว {i}: ประเภทคำขอไม่ถูกต้อง '{kind}'"); continue
        st_ = str(r.get("scope_type") or "all").strip().lower()
        if st_ not in ("all", "department"):
            st_ = "all"
        sv = "*" if st_ == "all" else str(r.get("scope_value") or "").strip()
        if st_ == "department" and not sv:
            errors.append(f"แถว {i}: scope=department ต้องระบุชื่อแผนก"); continue
        slots = {k: str(r.get(k) or "").strip() for k, _e, _t in ROUTING_SLOTS}
        if valid_emp is not None:
            bad = [v for v in slots.values() if v and v not in valid_emp]
            if bad:
                errors.append(f"แถว {i}: ไม่พบรหัสพนักงาน {', '.join(bad)}")
                continue
        if not any(slots.values()):
            errors.append(f"แถว {i}: ไม่มีผู้อนุมัติเลย"); continue
        set_routing(kind, st_, sv, slots, actor)
        applied += 1
    return applied, errors


def routing_template_xlsx(employees=None):
    """Build the bulk-upload template (.xlsx) → bytes. Sheets:
    Instructions · Routing (fillable) · Employees (lookup)."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    NAVY = "15294D"; BLUE = "009ADE"; GREY = "EEF3FA"
    hdr = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hf = PatternFill("solid", fgColor=NAVY)
    sub = Font(name="Arial", bold=True, color=NAVY, size=12)
    body = Font(name="Arial", size=10)
    note = Font(name="Arial", size=10, italic=True, color="5B6472")
    thin = Side(style="thin", color="C9D3E3")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()

    # ---- Instructions ----
    ws = wb.active; ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "ANCA HRM · Approval Routing — Bulk Upload Template"
    ws["B2"].font = Font(name="Arial", bold=True, color=NAVY, size=15)
    ws["B3"] = "ตั้งค่าสายการอนุมัติทีละหลายรายการ · set many approval lines at once"
    ws["B3"].font = note
    lines = [
        ("1.", "ไปที่แท็บ “Routing” แล้วกรอกหนึ่งแถวต่อหนึ่ง (ประเภทคำขอ + ขอบเขต)."),
        ("",   "Go to the “Routing” tab and fill one row per (request kind + scope)."),
        ("2.", "Request kind: leave / ot / shift / time_edit / resign / car / po / "
               "permit_out / permit_entry / stock (ใช้ตัวพิมพ์เล็ก)."),
        ("3.", "Scope: ใส่ all = ทุกแผนก  หรือ  department = เฉพาะแผนก "
               "(แล้วระบุชื่อแผนกในคอลัมน์ Department ให้ตรงกับระบบ)."),
        ("",   "Scope: 'all' applies company-wide; 'department' needs the Department "
               "column filled with the exact department name."),
        ("4.", "Petitioner1–2 / Approver1–3 / Reviewer: ใส่ “รหัสพนักงาน” (Emp No) "
               "ของผู้ที่ต้องลงนามตามลำดับ. เว้นว่างได้ถ้าไม่มีชั้นนั้น."),
        ("",   "Put the staff number (Emp No) of each signer in order. Leave a cell "
               "blank to skip that step. The order signed is left → right."),
        ("5.", "ดูรหัสพนักงานได้ที่แท็บ “Employees”. Applicant = ผู้ยื่นเอง "
               "(ไม่ต้องกรอก)."),
        ("",   "Look up codes on the “Employees” tab. Applicant is the requester "
               "themself — no need to enter it."),
        ("6.", "บันทึกไฟล์ แล้วอัปโหลดที่ ผู้ดูแลระบบ → สายการอนุมัติ → แท็บ "
               "“อัปโหลดทีละมาก”. ระบบจะตรวจรหัสและสรุปผลให้."),
        ("",   "Save, then upload at Admin → Approval lines → “Bulk upload”. The "
               "system validates the codes and reports the result."),
        ("!",  "ถ้าตั้งซ้ำประเภท+ขอบเขตเดิม ระบบจะเขียนทับของเดิม · re-uploading the "
               "same kind+scope overwrites the previous line."),
    ]
    r = 5
    for tag, txt in lines:
        ws.cell(r, 2, tag).font = Font(name="Arial", bold=True, color=BLUE, size=10)
        c = ws.cell(r, 3, txt); c.font = body; c.alignment = left
        ws.row_dimensions[r].height = 26
        r += 1
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 96

    # ---- Routing (fillable) ----
    rt = wb.create_sheet("Routing")
    rt.sheet_view.showGridLines = False
    cols = ["Request kind", "Scope (all/department)", "Department",
            "Petitioner1", "Petitioner2", "Approver1", "Approver2",
            "Approver3", "Reviewer"]
    for j, c in enumerate(cols, 1):
        cell = rt.cell(1, j, c); cell.font = hdr; cell.fill = hf
        cell.alignment = center; cell.border = box
    examples = [
        ["leave", "all", "", "", "", "1010992", "", "", "HR001"],
        ["ot", "all", "", "", "", "1010992", "1020625", "", "HR001"],
        ["shift", "department", "Weld", "", "", "1020860", "", "", "HR001"],
        ["resign", "all", "", "", "", "1020860", "HR010", "GM001", "HR001"],
    ]
    for i, row in enumerate(examples, 2):
        for j, v in enumerate(row, 1):
            cell = rt.cell(i, j, v); cell.font = body; cell.border = box
            cell.alignment = center if j != 3 else left
    rt.cell(7, 1, "↑ แถว 2–5 เป็นตัวอย่าง — แก้ไข/ลบ/เพิ่มได้ · rows 2–5 are "
                  "examples; edit, delete, or add your own.").font = note
    widths = [16, 20, 16, 13, 13, 12, 12, 12, 12]
    from openpyxl.utils import get_column_letter
    for j, w in enumerate(widths, 1):
        rt.column_dimensions[get_column_letter(j)].width = w
    rt.freeze_panes = "A2"

    # ---- Employees (lookup) ----
    em = wb.create_sheet("Employees")
    em.sheet_view.showGridLines = False
    for j, c in enumerate(["Emp No", "Name", "Department"], 1):
        cell = em.cell(1, j, c); cell.font = hdr; cell.fill = hf
        cell.alignment = center; cell.border = box
    rows = employees or []
    if rows:
        for i, e in enumerate(rows, 2):
            em.cell(i, 1, str(e.get("emp_no") or "")).font = body
            em.cell(i, 2, e.get("emp_name_en") or e.get("emp_name_th") or "").font = body
            em.cell(i, 3, e.get("dept_location") or "").font = body
    else:
        em.cell(2, 1, "—").font = body
        em.cell(2, 2, "ดาวน์โหลดจากในแอปเพื่อได้รายชื่อพนักงานจริง · download "
                      "from the app for your live employee list").font = note
    for j, w in enumerate([16, 34, 22], 1):
        em.column_dimensions[get_column_letter(j)].width = w
    em.freeze_panes = "A2"

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def parse_routing_upload(file_bytes):
    """Read a filled template → list of row dicts for bulk_apply()."""
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(file_bytes), data_only=True)
    ws = wb["Routing"] if "Routing" in wb.sheetnames else wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        g = lambda i: (str(r[i]).strip() if i < len(r) and r[i] is not None else "")
        kind = g(0).lower()
        if kind.startswith("↑") or not kind:
            continue
        rows.append({
            "kind": kind, "scope_type": g(1).lower() or "all",
            "scope_value": g(2),
            "petitioner1": g(3), "petitioner2": g(4),
            "approver1": g(5), "approver2": g(6), "approver3": g(7),
            "reviewer": g(8),
        })
    return rows
